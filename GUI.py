"""
ブラウザ自動操作 GUI アプリケーション
PySide6を使用した左サイドバー + 右メインコンテンツのレイアウト
"""

import sys
import os
import subprocess
import json
import urllib.parse
from pathlib import Path
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QStackedWidget, QPushButton, QLabel, QLineEdit, QTextEdit,
    QTableWidget, QTableWidgetItem, QHeaderView, QComboBox,
    QSpinBox, QCheckBox, QGroupBox, QFormLayout, QFrame,
    QSplitter, QListWidget, QListWidgetItem, QFileDialog, QMessageBox,
    QTabWidget, QGraphicsOpacityEffect, QProgressBar, QRadioButton,
    QMenu, QDialog, QPlainTextEdit, QScrollArea, QGridLayout
)
from PySide6.QtCore import Qt, QSize, QThread, Signal, QTimer, QPropertyAnimation, QEasingCurve, QPoint
from PySide6.QtGui import QFont, QIcon, QPalette, QColor, QPainter, QPen, QBrush, QPixmap
import base64

# ===== セキュリティ機能 =====
import hashlib as _hl
import zlib as _zl

def _xk(s):
    """キー復元（難読化）"""
    return bytes([c ^ 0x42 for c in s])

# 難読化された暗号化キー（実行時に復元）
_EK1 = _xk(b"\x120-('!6\x15\x0b\x0c\x1d\x00-6\x1d\t';\x1dprpv\x1d\x11'!70'cc")  # Bot用
_EK2 = _xk(b"\x120-('!6\x15\x0b\x0c\x1d\x01-0'\x1d\t';\x1dprpv\x1d\x11'!70'c")  # Core用

def _chk_dbg():
    """デバッガー検知"""
    try:
        import ctypes
        if hasattr(ctypes, 'windll'):
            return ctypes.windll.kernel32.IsDebuggerPresent() != 0
    except:
        pass
    return False

def _vfy(data: bytes, expected_crc: int = None) -> bool:
    """データ整合性チェック"""
    if expected_crc is None:
        return True
    return _zl.crc32(data) == expected_crc

# ===== アイコン管理（メモリのみ方式） =====
# アイコンのダウンロードURL（GitHub Raw）
ICONS_BASE_URL = "https://raw.githubusercontent.com/TestRaffle/projectwin-assets/main/"

# ===== Bot管理（メモリのみ方式） =====
# BotのダウンロードURL（GitHub Raw - 暗号化ファイル）
BOTS_BASE_URL = "https://raw.githubusercontent.com/TestRaffle/projectwin-assets/main/bots/"

# メモリ上のBotコードキャッシュ
_bot_code_cache = {}

# メモリ上のアイコンキャッシュ
_icon_cache = {}

def _download_icon(name: str) -> bytes:
    """サーバーからアイコンをダウンロード（メモリのみ、ファイル保存なし）"""
    import urllib.request
    import urllib.error
    
    url = f"{ICONS_BASE_URL}{name}.png"
    
    try:
        with urllib.request.urlopen(url, timeout=10) as response:
            return response.read()
    except Exception as e:
        print(f"アイコンダウンロードエラー: {name} - {e}")
        return None

def get_icon_from_base64(name):
    """アイコンを取得（メモリキャッシュ → ダウンロード）"""
    global _icon_cache
    
    # 1. メモリキャッシュをチェック
    if name in _icon_cache:
        return _icon_cache[name]
    
    # 2. サーバーからダウンロード（ファイルには保存しない）
    icon_data = _download_icon(name)
    
    # 3. アイコンを生成してメモリキャッシュに保存
    if icon_data:
        pixmap = QPixmap()
        pixmap.loadFromData(icon_data)
        icon = QIcon(pixmap)
        _icon_cache[name] = icon
        return icon
    
    # 失敗した場合は空のアイコン
    empty_icon = QIcon()
    _icon_cache[name] = empty_icon
    return empty_icon


# ===== Bot管理関数（メモリのみ方式） =====
def _dcd(data: bytes, key: bytes) -> bytes:
    """汎用XOR復号"""
    result = bytearray()
    for i, byte in enumerate(data):
        result.append(byte ^ key[i % len(key)])
    return bytes(result)

def download_bot_code(site: str, mode: str) -> str:
    """サーバーから暗号化されたBotコードをダウンロードしてメモリ上で復号
    ローカルに.pyファイルがあればそれを優先して使用（開発・テスト用）
    """
    global _bot_code_cache
    
    # デバッガー検知
    if _chk_dbg():
        return None
    
    cache_key = f"{site}_{mode}"
    
    # 1. メモリキャッシュをチェック
    if cache_key in _bot_code_cache:
        return _bot_code_cache[cache_key]
    
    # 2. ローカルの.pyファイルをチェック（_internal/bots/site/site_mode.py）
    local_bot_path = APP_DIR / "_internal" / "bots" / site / f"{site}_{mode}.py"
    print(f"[DEBUG] Looking for local bot: {local_bot_path}")
    print(f"[DEBUG] APP_DIR: {APP_DIR}")
    print(f"[DEBUG] File exists: {local_bot_path.exists()}")
    if local_bot_path.exists():
        try:
            with open(local_bot_path, 'r', encoding='utf-8') as f:
                bot_code = f.read()
            print(f"[DEV] Loaded local bot: {local_bot_path}")
            _bot_code_cache[cache_key] = bot_code
            return bot_code
        except Exception as e:
            print(f"[DEV] Failed to load local bot: {e}")
    
    # 3. サーバーからダウンロード
    import urllib.request
    import urllib.error
    
    # 暗号化ファイル名: amazon_signup.enc
    url = f"{BOTS_BASE_URL}{site}/{site}_{mode}.enc"
    
    try:
        with urllib.request.urlopen(url, timeout=30) as response:
            encrypted_data = response.read()
        
        # 4. メモリ上で復号（難読化されたキーを使用）
        decrypted_data = _dcd(encrypted_data, _EK1)
        bot_code = decrypted_data.decode('utf-8')
        
        # 5. メモリキャッシュに保存（ファイルには保存しない）
        _bot_code_cache[cache_key] = bot_code
        
        return bot_code
    except Exception as e:
        print(f"Botダウンロードエラー: {cache_key} - {e}")
        return None

def get_bot_module(site: str, mode: str, task_data: dict):
    """Botコードをダウンロードしてモジュールとして返す"""
    import types
    
    bot_code = download_bot_code(site, mode)
    if bot_code is None:
        return None
    
    try:
        # コードをコンパイルして実行
        bot_module = types.ModuleType(f"bot_{site}_{mode}")
        bot_module.__file__ = f"<server>/{site}_{mode}.py"
        bot_module.__builtins__ = __builtins__
        
        # APP_DIRをexec前にモジュール辞書に追加（botコード内で使えるように）
        bot_module.__dict__['APP_DIR'] = APP_DIR
        
        exec(bot_code, bot_module.__dict__)
        
        return bot_module
    except Exception as e:
        print(f"Bot実行エラー: {site}_{mode} - {e}")
        import traceback
        traceback.print_exc()
        return None


# ===== コアモジュール管理（メモリのみ方式） =====
# コアモジュールのダウンロードURL
CORE_MODULES_BASE_URL = "https://raw.githubusercontent.com/TestRaffle/projectwin-assets/main/core/"

# メモリ上のコアモジュールキャッシュ
_core_module_cache = {}

def download_core_module(module_name: str):
    """サーバーからコアモジュールをダウンロードしてメモリ上で復号・実行"""
    global _core_module_cache
    
    # デバッガー検知
    if _chk_dbg():
        return None
    
    # 1. メモリキャッシュをチェック
    if module_name in _core_module_cache:
        return _core_module_cache[module_name]
    
    # 2. サーバーからダウンロード
    import urllib.request
    import urllib.error
    import types
    
    url = f"{CORE_MODULES_BASE_URL}{module_name}.enc"
    
    try:
        print(f"Loading {module_name}...")
        with urllib.request.urlopen(url, timeout=30) as response:
            encrypted_data = response.read()
        
        # 3. メモリ上で復号（難読化されたキーを使用）
        decrypted_data = _dcd(encrypted_data, _EK2)
        module_code = decrypted_data.decode('utf-8')
        
        # 4. モジュールとして実行（APP_DIRを渡す）
        module = types.ModuleType(module_name)
        module.__file__ = f"<server>/{module_name}.py"
        module.APP_DIR = APP_DIR  # APP_DIRを渡す
        exec(module_code, module.__dict__)
        
        # 5. メモリキャッシュに保存
        _core_module_cache[module_name] = module
        
        return module
    except Exception as e:
        print(f"コアモジュールダウンロードエラー: {module_name} - {e}")
        return None

def load_license_manager():
    """LicenseManagerをサーバーから取得"""
    module = download_core_module("license_manager")
    if module and hasattr(module, 'LicenseManager'):
        return module.LicenseManager
    return None

def load_updater_functions():
    """updaterの関数をサーバーから取得"""
    module = download_core_module("updater")
    if module:
        return {
            'check_for_update': getattr(module, 'check_for_update', None),
            'download_update': getattr(module, 'download_update', None),
            'apply_update': getattr(module, 'apply_update', None),
            'restart_app': getattr(module, 'restart_app', None),
            'save_version': getattr(module, 'save_version', None),
        }
    return None


try:
    import openpyxl
except ImportError:
    print("openpyxlがインストールされていません。pip install openpyxl を実行してください。")


def is_compiled():
    """exe化されているか判定（PyInstaller/Nuitka両対応）"""
    # PyInstaller
    if getattr(sys, 'frozen', False):
        return True
    # Nuitka - __compiled__ 属性をチェック
    if "__compiled__" in globals():
        return True
    main_module = sys.modules.get('__main__', None)
    if main_module and hasattr(main_module, '__compiled__'):
        return True
    # Nuitka standalone - パスに.distまたはProjectWINが含まれている場合
    exe_path = sys.executable.lower()
    if '.dist' in exe_path or 'projectwin' in exe_path:
        return True
    # Nuitka onefile - Tempフォルダ内のONEFILで実行されている場合
    if 'temp' in exe_path and 'onefil' in exe_path:
        return True
    return False


def get_app_dir():
    """アプリのルートディレクトリを取得（PyInstaller/Nuitka両対応）"""
    if is_compiled():
        # Nuitka standalone の場合
        exe_dir = Path(sys.executable).parent
        
        # 複数の階層を探索してProjectWIN.exeを探す
        for i in range(5):  # 最大5階層上まで探索
            check_dir = exe_dir
            for _ in range(i):
                check_dir = check_dir.parent
            
            if (check_dir / "ProjectWIN.exe").exists():
                return check_dir
        
        # 見つからない場合は、パスから推測
        # パスに"ProjectWIN"または".dist"が含まれるフォルダを探す
        parts = exe_dir.parts
        for i, part in enumerate(parts):
            if 'projectwin' in part.lower() or part.endswith('.dist'):
                return Path(*parts[:i+1])
        
        # それでも見つからない場合はexeの場所を返す
        return exe_dir
    else:
        # 開発中: スクリプトの場所
        return Path(__file__).parent


# アプリのベースディレクトリ（コアモジュール読み込み前に定義が必要）
APP_DIR = get_app_dir()

# 設定フォルダ（_internal内に配置）
SETTINGS_DIR = APP_DIR / "_internal" / "settings"
SETTINGS_DIR.mkdir(parents=True, exist_ok=True)


# ライセンス認証（サーバーからダウンロード、フォールバックとしてローカル）
LicenseManager = None
try:
    # まずサーバーから取得を試みる
    LicenseManager = load_license_manager()
except:
    pass

if LicenseManager is None:
    # フォールバック: ローカルファイル（開発用）
    try:
        from license_manager import LicenseManager
    except ImportError:
        LicenseManager = None

# 自動アップデーター（サーバーからダウンロード、フォールバックとしてローカル）
check_for_update = None
download_update = None
apply_update = None
restart_app = None
save_version = None

try:
    # まずサーバーから取得を試みる
    updater_funcs = load_updater_functions()
    if updater_funcs:
        check_for_update = updater_funcs['check_for_update']
        download_update = updater_funcs['download_update']
        apply_update = updater_funcs['apply_update']
        restart_app = updater_funcs['restart_app']
        save_version = updater_funcs['save_version']
except:
    pass

if check_for_update is None:
    # フォールバック: ローカルファイル（開発用）
    try:
        from updater import check_for_update, download_update, apply_update, restart_app, save_version
    except ImportError:
        check_for_update = None


def get_app_version():
    """アプリのバージョンを取得"""
    try:
        if is_compiled():
            # exe化されている場合: _internal内を参照
            version_path = APP_DIR / "_internal" / "version.json"
        else:
            # 開発中: APP_DIR内を参照
            version_path = APP_DIR / "version.json"
        
        if version_path.exists():
            content = version_path.read_text().strip()
            # JSON形式の場合
            if content.startswith('{'):
                data = json.loads(content)
                return data.get("version", "0.0.0")
            # プレーンテキストの場合（vプレフィックスを除去）
            return content.lstrip('v').strip()
    except:
        pass
    return "0.0.0"




class CheckmarkCheckBox(QCheckBox):
    """緑色のチェックマーク付きカスタムチェックボックス"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedSize(22, 22)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
    
    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        
        # ボックスの描画
        rect = self.rect().adjusted(2, 2, -2, -2)
        
        if self.isChecked():
            # チェック時: 緑の枠線
            pen = QPen(QColor("#27ae60"))
            pen.setWidth(2)
            painter.setPen(pen)
            painter.setBrush(QBrush(Qt.GlobalColor.transparent))
            painter.drawRoundedRect(rect, 4, 4)
            
            # チェックマークを描画（細めでスタイリッシュに）
            pen = QPen(QColor("#27ae60"))
            pen.setWidth(2)
            pen.setCapStyle(Qt.PenCapStyle.RoundCap)
            pen.setJoinStyle(Qt.PenJoinStyle.RoundJoin)
            painter.setPen(pen)
            
            # チェックマークの座標（少し小さめに）
            painter.drawLine(6, 11, 9, 14)
            painter.drawLine(9, 14, 16, 6)
        else:
            # 未チェック時: グレーの枠線
            if self.underMouse():
                pen = QPen(QColor("#27ae60"))
            else:
                pen = QPen(QColor("#505050"))
            pen.setWidth(2)
            painter.setPen(pen)
            painter.setBrush(QBrush(Qt.GlobalColor.transparent))
            painter.drawRoundedRect(rect, 4, 4)
        
        painter.end()


class TextCheckmarkCheckBox(QCheckBox):
    """テキスト付き緑色チェックマークカスタムチェックボックス（Task風スタイリッシュデザイン）"""
    
    def __init__(self, text="", parent=None):
        super().__init__(text, parent)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self._box_size = 18
        self.setMouseTracking(True)
    
    def set_box_size(self, size):
        """ボックスサイズを設定"""
        self._box_size = size
        self.update()
    
    def sizeHint(self):
        """サイズヒントを返す"""
        from PySide6.QtCore import QSize
        fm = self.fontMetrics()
        text_width = fm.horizontalAdvance(self.text())
        return QSize(self._box_size + 8 + text_width, max(self._box_size, fm.height()) + 4)
    
    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        
        # ボックスの位置とサイズ
        box_size = self._box_size
        box_x = 2
        box_y = (self.height() - box_size) // 2
        
        # ボックスの矩形（内側に2pxマージン）
        rect_x = box_x + 2
        rect_y = box_y + 2
        rect_w = box_size - 4
        rect_h = box_size - 4
        
        if self.isChecked():
            # チェック時: 緑の枠線
            pen = QPen(QColor("#27ae60"))
            pen.setWidth(2)
            painter.setPen(pen)
            painter.setBrush(QBrush(Qt.GlobalColor.transparent))
            painter.drawRoundedRect(rect_x, rect_y, rect_w, rect_h, 4, 4)
            
            # チェックマークを描画（Taskと同じ細めでスタイリッシュ）
            pen = QPen(QColor("#27ae60"))
            pen.setWidth(2)
            pen.setCapStyle(Qt.PenCapStyle.RoundCap)
            pen.setJoinStyle(Qt.PenJoinStyle.RoundJoin)
            painter.setPen(pen)
            
            # チェックマーク座標をボックスサイズに合わせてスケール
            cx = box_x + box_size // 2
            cy = box_y + box_size // 2
            
            if box_size >= 18:
                # 大きいサイズ（18px）: 小さめのチェックマーク
                painter.drawLine(box_x + 6, cy, box_x + 8, cy + 2)
                painter.drawLine(box_x + 8, cy + 2, box_x + 13, cy - 3)
            else:
                # 小さいサイズ（14px）: さらに小さめ
                painter.drawLine(box_x + 4, cy, box_x + 6, cy + 2)
                painter.drawLine(box_x + 6, cy + 2, box_x + 10, cy - 2)
        else:
            # 未チェック時: グレーまたは緑の枠線
            if self.underMouse():
                pen = QPen(QColor("#27ae60"))
            else:
                pen = QPen(QColor("#505050"))
            pen.setWidth(2)
            painter.setPen(pen)
            painter.setBrush(QBrush(Qt.GlobalColor.transparent))
            painter.drawRoundedRect(rect_x, rect_y, rect_w, rect_h, 4, 4)
        
        # テキストの描画
        painter.setPen(QColor("#ffffff"))
        font = painter.font()
        if self._box_size <= 14:
            font.setPointSize(9)
        else:
            font.setPointSize(10)
        painter.setFont(font)
        
        text_x = box_x + box_size + 6
        text_y = (self.height() + painter.fontMetrics().ascent() - painter.fontMetrics().descent()) // 2
        painter.drawText(text_x, text_y, self.text())
        
        painter.end()
    
    def enterEvent(self, event):
        """マウスが入った時"""
        self.update()
        super().enterEvent(event)
    
    def leaveEvent(self, event):
        """マウスが出た時"""
        self.update()
        super().leaveEvent(event)


class SwitchButton(QPushButton):
    """カスタムスイッチボタン"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setCheckable(True)
        self.setFixedSize(50, 26)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self._update_style()
        self.toggled.connect(self._update_style)
    
    def _update_style(self):
        if self.isChecked():
            self.setStyleSheet("""
                QPushButton {
                    background-color: #27ae60;
                    border-radius: 13px;
                    border: none;
                    text-align: right;
                    padding-right: 5px;
                }
            """)
            self.setText("●")
        else:
            self.setStyleSheet("""
                QPushButton {
                    background-color: #404050;
                    border-radius: 13px;
                    border: none;
                    text-align: left;
                    padding-left: 5px;
                    color: #808080;
                }
            """)
            self.setText("●")


class ToastNotification(QWidget):
    """アニメーション付きトースト通知（親ウィジェット内に表示）"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        # 親ウィジェット内の子ウィジェットとして表示（別ウィンドウではなく）
        self.setFixedHeight(50)
        
        # メインレイアウト
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        # コンテナ
        self.container = QWidget()
        self.container.setObjectName("toast_container")
        container_layout = QHBoxLayout(self.container)
        container_layout.setContentsMargins(15, 8, 10, 8)
        container_layout.setSpacing(10)
        
        # アイコン
        self.icon_label = QLabel()
        self.icon_label.setFixedSize(20, 20)
        self.icon_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        container_layout.addWidget(self.icon_label)
        
        # メッセージ
        self.message_label = QLabel()
        self.message_label.setStyleSheet("color: #ffffff; font-size: 13px; background: transparent;")
        container_layout.addWidget(self.message_label)
        
        container_layout.addStretch()
        
        # プログレスバー（タイマー表示用）- 右から左に向かって減る
        self.progress = QProgressBar()
        self.progress.setFixedSize(60, 4)
        self.progress.setTextVisible(False)
        self.progress.setRange(0, 100)
        self.progress.setValue(100)
        self.progress.setLayoutDirection(Qt.LayoutDirection.RightToLeft)  # 右から左へ
        self.progress.setStyleSheet("""
            QProgressBar { background-color: rgba(255,255,255,0.3); border-radius: 2px; }
            QProgressBar::chunk { background-color: #ffffff; border-radius: 2px; }
        """)
        container_layout.addWidget(self.progress)
        
        # 閉じるボタン
        self.close_btn = QPushButton("×")
        self.close_btn.setFixedSize(20, 20)
        self.close_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.close_btn.setStyleSheet("""
            QPushButton { background: transparent; color: #808080; border: none; font-size: 14px; font-weight: bold; }
            QPushButton:hover { color: #e74c3c; }
        """)
        self.close_btn.clicked.connect(self.hide_toast)
        container_layout.addWidget(self.close_btn)
        
        layout.addWidget(self.container)
        
        # タイマー
        self.timer = QTimer()
        self.timer.timeout.connect(self._update_progress)
        self.progress_value = 100
        self.duration = 3000
        
        self.hide()
    
    def show_toast(self, message, toast_type="success", duration=3000):
        """トーストを表示"""
        self.duration = duration
        self.progress_value = 100
        self.progress.setValue(100)
        
        # タイプに応じた文字色とアイコン
        if toast_type == "success":
            text_color = "#2ecc71"  # 緑
            icon = "✓"
        elif toast_type == "error":
            text_color = "#e74c3c"  # 赤
            icon = "✗"
        elif toast_type == "warning":
            text_color = "#f39c12"  # 黄
            icon = "⚠"
        else:
            text_color = "#3498db"  # 青
            icon = "ℹ"
        
        # GUIメイン画面と同じ背景色 + シャドウ効果
        self.container.setStyleSheet(f"""
            #toast_container {{
                background-color: #1e1e2e;
                border: 1px solid #404050;
                border-radius: 8px;
            }}
        """)
        
        # シャドウ効果を追加
        from PySide6.QtWidgets import QGraphicsDropShadowEffect
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(20)
        shadow.setXOffset(0)
        shadow.setYOffset(4)
        shadow.setColor(QColor(0, 0, 0, 100))
        self.container.setGraphicsEffect(shadow)
        
        self.icon_label.setText(icon)
        self.icon_label.setStyleSheet(f"color: {text_color}; font-size: 16px; font-weight: bold; background: transparent;")
        self.message_label.setText(message)
        self.message_label.setStyleSheet(f"color: {text_color}; font-size: 13px; font-weight: bold; background: transparent;")
        
        # プログレスバーの色も文字色に合わせる
        self.progress.setStyleSheet(f"""
            QProgressBar {{ background-color: #404050; border-radius: 2px; }}
            QProgressBar::chunk {{ background-color: {text_color}; border-radius: 2px; }}
        """)
        
        # 親ウィジェットの下部（Start All/Stop Allと同じ高さ）に配置
        if self.parent():
            parent_rect = self.parent().rect()
            toast_width = min(350, parent_rect.width() - 40)
            self.setFixedWidth(toast_width)
            # 右寄せで下部に配置
            x = parent_rect.width() - toast_width - 30
            y = parent_rect.height() - 80  # 下から80pxの位置
            self.move(x, y)
        
        self.show()
        self.raise_()
        
        # プログレスバーのアニメーション開始
        self.timer.start(30)
    
    def _update_progress(self):
        """プログレスバーを更新（左から右に減る）"""
        decrement = 100 / (self.duration / 30)
        self.progress_value -= decrement
        
        if self.progress_value <= 0:
            self.hide_toast()
        else:
            # 左から減るように値を反転
            self.progress.setValue(int(self.progress_value))
    
    def hide_toast(self):
        """トーストを非表示"""
        self.timer.stop()
        self.hide()


class SidebarButton(QPushButton):
    """サイドバー用のカスタムボタン（画像アイコン対応）"""
    
    def __init__(self, text, icon_name="", parent=None):
        super().__init__(parent)
        self.full_text = text
        self.icon_name = icon_name
        self.setCheckable(True)
        self.setFixedHeight(50)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        
        # アイコンを設定（Base64から）
        if icon_name:
            icon = get_icon_from_base64(icon_name)
            if not icon.isNull():
                self.setIcon(icon)
                self.setIconSize(QSize(20, 20))
        
        self.setText(f"  {text}" if icon_name else text)
        self.setStyleSheet(self._get_style(expanded=True))
    
    def set_expanded(self, expanded):
        """展開/折りたたみ状態を設定"""
        if expanded:
            self.setText(f"  {self.full_text}" if self.icon_name else self.full_text)
            self.setStyleSheet(self._get_style(expanded=True))
        else:
            self.setText("")  # 折りたたみ時はテキストなし、アイコンのみ
            self.setStyleSheet(self._get_style(expanded=False))
    
    def _get_style(self, expanded=True):
        alignment = "left" if expanded else "center"
        padding = "12px 15px 12px 8px" if expanded else "12px 5px"
        return f"""
            QPushButton {{
                background-color: transparent;
                color: #808080;
                border: none;
                border-radius: 8px;
                padding: {padding};
                margin-right: 12px;
                text-align: {alignment};
                font-size: 14px;
                font-weight: 500;
            }}
            QPushButton:hover {{
                color: #b0b0b0;
            }}
            QPushButton:checked {{
                background-color: transparent;
                color: #ffffff;
            }}
        """
    
    def paintEvent(self, event):
        """選択時に短いボーダーラインを描画"""
        super().paintEvent(event)
        if self.isChecked():
            from PySide6.QtGui import QPainter, QColor
            painter = QPainter(self)
            painter.setRenderHint(QPainter.RenderHint.Antialiasing)
            painter.setBrush(QColor("#4a90d9"))
            painter.setPen(Qt.PenStyle.NoPen)
            # ボタンの高さの60%の長さで中央に配置
            line_height = int(self.height() * 0.5)
            y_offset = (self.height() - line_height) // 2
            painter.drawRoundedRect(0, y_offset, 3, line_height, 1, 1)
            painter.end()


class BotWorker(QThread):
    """Botを別スレッドで実行するワーカー（サーバーダウンロード方式）"""
    status_changed = Signal(int, str)
    result_changed = Signal(int, str)
    finished_task = Signal(int)
    raffle_result = Signal(int, list)
    log_updated = Signal(int, str)
    
    def __init__(self, row, task_data):
        super().__init__()
        self.row = row
        self.task_data = task_data
        self._is_running = True
        self._stopped_by_user = False
        self._stop_requested = False  # ストップリクエストフラグ
        self._raffle_results = []
        self.log_lines = []
        self.bot_instance = None  # botインスタンスを保持（stop用）
    
    def get_log(self):
        return "\n".join(self.log_lines)
    
    def _timestamp(self):
        from datetime import datetime
        return datetime.now().strftime("%H:%M:%S")
    
    def _capture_print(self, text):
        """printをキャプチャしてステータス更新"""
        text = str(text).strip()
        if not text:
            return
        
        # \x00STATUS:で始まる行はステータスを更新（ログには追加しない、コンソールにも表示しない）
        # status_changedシグナルを使用（webhookなし）
        if "\x00STATUS:" in text:
            try:
                status = text.split("\x00STATUS:")[1].strip()
                self.status_changed.emit(self.row, status)
            except:
                pass
            return  # ログには追加しない
        
        # 旧形式のSTATUS:もサポート（後方互換性）
        if text.startswith("STATUS:"):
            try:
                status = text.split("STATUS:")[1].strip()
                self.status_changed.emit(self.row, status)
            except:
                pass
            return  # ログには追加しない
        
        # ログに追加
        self.log_lines.append(f"[{self._timestamp()}] {text}")
        self.log_updated.emit(self.row, text)
    
    def run(self):
        import io
        import contextlib
        import threading
        
        try:
            self.status_changed.emit(self.row, "Starting")
            self.log_lines.append(f"[{self._timestamp()}] Task started")
            
            # サイトとモードを取得
            mode = self.task_data.get("Mode", "").lower()
            site = self.task_data.get("Site", "").lower()
            
            self.log_lines.append(f"[{self._timestamp()}] Loading bot: {site}_{mode}")
            
            # サーバーからBotモジュールを取得（メモリのみ）
            bot_module = get_bot_module(site, mode, self.task_data)
            
            if bot_module is None:
                self.result_changed.emit(self.row, "Error: Cannot load bot")
                return
            
            # クラス名を決定
            class_name = f"{site.capitalize()}{mode.capitalize()}"
            
            if not hasattr(bot_module, class_name):
                for attr_name in dir(bot_module):
                    if attr_name.lower() == class_name.lower():
                        class_name = attr_name
                        break
                else:
                    self.result_changed.emit(self.row, f"Error: Class {class_name} not found")
                    return
            
            bot_class = getattr(bot_module, class_name)
            
            # Botインスタンスを作成
            self.bot_instance = bot_class(self.task_data)
            
            self.result_changed.emit(self.row, "Running")
            
            # スレッドセーフなprintキャプチャ
            # 各スレッドのIDとコールバックを紐付ける
            current_thread_id = threading.current_thread().ident
            
            class ThreadSafePrintCapture:
                # クラス変数でスレッドごとのコールバックを管理
                _callbacks = {}
                _lock = threading.Lock()
                
                def __init__(self, original):
                    self.original = original
                
                @classmethod
                def register(cls, thread_id, callback):
                    with cls._lock:
                        cls._callbacks[thread_id] = callback
                
                @classmethod
                def unregister(cls, thread_id):
                    with cls._lock:
                        if thread_id in cls._callbacks:
                            del cls._callbacks[thread_id]
                
                def write(self, text):
                    if text.strip():
                        thread_id = threading.current_thread().ident
                        with self._lock:
                            callback = self._callbacks.get(thread_id)
                        if callback:
                            callback(text)
                    self.original.write(text)
                
                def flush(self):
                    self.original.flush()
            
            # printをキャプチャしながら実行
            import sys
            original_stdout = sys.stdout
            
            # 既にThreadSafePrintCaptureがインストールされているか確認
            if not isinstance(sys.stdout, ThreadSafePrintCapture):
                sys.stdout = ThreadSafePrintCapture(original_stdout)
            
            # このスレッドのコールバックを登録
            ThreadSafePrintCapture.register(current_thread_id, self._capture_print)
            
            result = None
            run_error = None
            try:
                result = self.bot_instance.run()
            except Exception as e:
                run_error = e
            finally:
                # このスレッドのコールバックを解除
                ThreadSafePrintCapture.unregister(current_thread_id)
            
            # Bot側のブラウザ閉じフラグをチェック
            if hasattr(self.bot_instance, '_browser_closed') and self.bot_instance._browser_closed:
                self._stop_requested = True
            if hasattr(self.bot_instance, '_stop_requested') and self.bot_instance._stop_requested:
                self._stop_requested = True
            
            # ストップフラグが設定されていた場合の処理
            if self._stopped_by_user or self._stop_requested:
                # botが成功を返した場合（Browserモードで正常終了など）はその結果を使う
                if isinstance(result, tuple) and len(result) >= 2 and result[0] == True:
                    pass  # 下の結果解析処理に進む
                else:
                    self.result_changed.emit(self.row, "Stopped")
                    return
            
            # 例外が発生していた場合は再raise
            if run_error:
                raise run_error
            
            # 結果を解析
            if isinstance(result, tuple):
                if len(result) >= 3:
                    success, results, error_status = result
                elif len(result) == 2:
                    success, second = result
                    # 2番目の要素が文字列ならerror_status（失敗時）またはsuccess_status（成功時）
                    if isinstance(second, str):
                        if success:
                            # 成功時の文字列はsuccess_statusとして扱う
                            error_status = second  # error_statusを再利用（result_changedで使用）
                        else:
                            error_status = second
                        results = None
                    else:
                        results = second
                        error_status = None
                else:
                    success = result[0] if result else False
                    results = []
                    error_status = None
            else:
                success = bool(result)
                results = []
                error_status = None
            
            # Raffle結果があれば送信（辞書形式のリストの場合のみ）
            if results and isinstance(results, list) and len(results) > 0:
                # 最初の要素が辞書かどうかでRaffle結果かを判定
                if isinstance(results[0], dict):
                    self._raffle_results = results
                    self.raffle_result.emit(self.row, results)
            
            # 結果を設定
            if error_status:
                self.result_changed.emit(self.row, error_status)
            elif success:
                self.result_changed.emit(self.row, "Success")
            else:
                self.result_changed.emit(self.row, "Failed")
            
        except Exception as e:
            # ストップによる例外の場合はStoppedとして処理
            if self._stopped_by_user or self._stop_requested:
                self.result_changed.emit(self.row, "Stopped")
            else:
                import traceback
                traceback.print_exc()
                self.result_changed.emit(self.row, f"Error: {str(e)[:50]}")
        finally:
            self.finished_task.emit(self.row)
    
    def stop(self):
        """タスクを停止（フラグを設定）"""
        self._is_running = False
        self._stopped_by_user = True
        self._stop_requested = True
        
        # Bot側にもストップフラグを設定
        if self.bot_instance:
            self.bot_instance._stop_requested = True
            print("Stop requested")


class TaskPage(QWidget):
    """タスクページ"""
    
    CONFIG_FILE = SETTINGS_DIR / "app_config.json"
    GENERAL_SETTINGS_FILE = SETTINGS_DIR / "general_settings.json"
    
    def __init__(self, proxy_page=None):
        super().__init__()
        self.proxy_page = proxy_page
        self.settings_page = None  # Webhook送信用の参照（MainWindowで設定）
        self.workers = {}
        self.current_excel_path = None
        self.last_checked_row = None
        self.all_task_data = []
        self.original_task_data = []  # Excelから読み込んだ元のデータ
        self.pending_tasks = []  # 待機中のタスク（並列制御用）
        self.setup_ui()
        self._load_last_excel()
    
    def _load_last_excel(self):
        """前回読み込んだExcelファイルを自動読み込み"""
        try:
            if self.CONFIG_FILE.exists():
                with open(self.CONFIG_FILE, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                last_excel = config.get("last_excel_path", "")
                if last_excel and os.path.exists(last_excel):
                    self._load_excel_file(last_excel, show_toast=False)
        except Exception as e:
            print(f"Failed to load last Excel: {e}")
    
    def _save_last_excel(self, path):
        """最後に読み込んだExcelファイルパスを保存"""
        try:
            self.CONFIG_FILE.parent.mkdir(parents=True, exist_ok=True)
            config = {}
            if self.CONFIG_FILE.exists():
                with open(self.CONFIG_FILE, 'r', encoding='utf-8') as f:
                    config = json.load(f)
            config["last_excel_path"] = path
            with open(self.CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Failed to save last Excel path: {e}")
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)
        
        # トースト通知
        self.toast = ToastNotification(self)
        
        # ヘッダー部分
        header_layout = QHBoxLayout()
        header_layout.setSpacing(8)
        
        self.header_label = QLabel("Tasks(0)")
        self.header_label.setStyleSheet("""
            font-size: 24px;
            font-weight: bold;
            color: #ffffff;
        """)
        header_layout.addWidget(self.header_label)
        
        header_layout.addStretch()
        
        # Headlessチェックボックス
        headless_container = QWidget()
        headless_layout = QHBoxLayout(headless_container)
        headless_layout.setContentsMargins(0, 0, 15, 0)
        headless_layout.setSpacing(8)
        
        self.headless_checkbox = CheckmarkCheckBox()
        headless_layout.addWidget(self.headless_checkbox)
        
        headless_label = QLabel("Headless")
        headless_label.setStyleSheet("color: #ffffff; font-size: 14px; font-weight: bold;")
        headless_label.setCursor(Qt.CursorShape.PointingHandCursor)
        headless_label.mousePressEvent = lambda e: self.headless_checkbox.setChecked(not self.headless_checkbox.isChecked())
        headless_layout.addWidget(headless_label)
        
        header_layout.addWidget(headless_container)
        
        # Use Localチェックボックス
        uselocal_container = QWidget()
        uselocal_layout = QHBoxLayout(uselocal_container)
        uselocal_layout.setContentsMargins(0, 0, 15, 0)
        uselocal_layout.setSpacing(8)
        
        self.uselocal_checkbox = CheckmarkCheckBox()
        uselocal_layout.addWidget(self.uselocal_checkbox)
        
        uselocal_label = QLabel("Use Local")
        uselocal_label.setStyleSheet("color: #ffffff; font-size: 14px; font-weight: bold;")
        uselocal_label.setCursor(Qt.CursorShape.PointingHandCursor)
        uselocal_label.mousePressEvent = lambda e: self.uselocal_checkbox.setChecked(not self.uselocal_checkbox.isChecked())
        uselocal_layout.addWidget(uselocal_label)
        
        header_layout.addWidget(uselocal_container)
        
        # 時計表示
        self.clock_label = QLabel()
        self.clock_label.setStyleSheet("color: #b0b0b0; font-size: 14px; font-family: monospace; margin-right: 15px;")
        header_layout.addWidget(self.clock_label)
        
        # 時計更新タイマー
        self.clock_timer = QTimer()
        self.clock_timer.timeout.connect(self._update_clock)
        self.clock_timer.start(1000)
        self._update_clock()
        
        # Excelファイル読み込み＆シート選択コンテナ
        excel_container = QWidget()
        excel_container.setStyleSheet("""
            QWidget {
                background-color: #2a2a3a;
                border-radius: 8px;
            }
        """)
        excel_layout = QHBoxLayout(excel_container)
        excel_layout.setContentsMargins(8, 4, 8, 4)
        excel_layout.setSpacing(6)
        
        # フォルダアイコンボタン
        load_btn = QPushButton()
        load_btn.setFixedSize(28, 28)
        load_btn.setToolTip("Load Excel")
        load_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        folder_icon = get_icon_from_base64("folder")
        if not folder_icon.isNull():
            load_btn.setIcon(folder_icon)
            load_btn.setIconSize(QSize(18, 18))
        else:
            load_btn.setText("📁")
        load_btn.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                border: none;
                font-size: 16px;
            }
            QPushButton:hover {
                background-color: #3a3a4a;
                border-radius: 4px;
            }
        """)
        load_btn.clicked.connect(self.load_excel)
        excel_layout.addWidget(load_btn)
        
        # シートセレクター
        self.sheet_selector = QComboBox()
        self.sheet_selector.setFixedHeight(24)
        self.sheet_selector.setMinimumWidth(120)
        self.sheet_selector.setMaximumWidth(250)
        self.sheet_selector.setCursor(Qt.CursorShape.PointingHandCursor)
        
        self.sheet_selector.setStyleSheet("""
            QComboBox {
                background-color: transparent;
                color: #ffffff;
                border: none;
                padding: 2px 4px;
                font-size: 14px;
                font-weight: bold;
            }
            QComboBox::drop-down {
                border: none;
                width: 16px;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 4px solid transparent;
                border-right: 4px solid transparent;
                border-top: 5px solid #808080;
                margin-right: 4px;
            }
            QComboBox QAbstractItemView {
                background-color: #2a2a3a;
                color: #ffffff;
                selection-background-color: #3a3a4a;
                border: 1px solid #3a3a4a;
                outline: none;
                font-size: 14px;
                font-weight: bold;
            }
            QComboBox QAbstractItemView::item {
                padding: 6px 10px;
                min-height: 24px;
            }
            QComboBox QAbstractItemView::item:hover {
                background-color: #3a3a4a;
            }
            QComboBox QAbstractItemView QScrollBar:vertical {
                background-color: #2a2a3a;
                width: 8px;
                border: none;
                border-radius: 4px;
            }
            QComboBox QAbstractItemView QScrollBar::handle:vertical {
                background-color: #4a4a5a;
                border-radius: 4px;
                min-height: 20px;
            }
            QComboBox QAbstractItemView QScrollBar::handle:vertical:hover {
                background-color: #5a5a6a;
            }
            QComboBox QAbstractItemView QScrollBar::add-line:vertical,
            QComboBox QAbstractItemView QScrollBar::sub-line:vertical {
                height: 0px;
            }
            QComboBox QAbstractItemView QScrollBar::add-page:vertical,
            QComboBox QAbstractItemView QScrollBar::sub-page:vertical {
                background-color: #2a2a3a;
            }
        """)
        self.sheet_selector.currentIndexChanged.connect(self._on_sheet_changed)
        self.sheet_selector.setPlaceholderText("No file")
        excel_layout.addWidget(self.sheet_selector)
        
        header_layout.addWidget(excel_container)
        
        # 現在のワークブックとシート情報を保持
        self.current_workbook = None
        self.current_sheet_names = []
        
        # リフレッシュボタン（シンプルなアイコン）
        refresh_btn = QPushButton()
        refresh_btn.setFixedSize(36, 36)
        refresh_btn.setToolTip("Refresh")
        refresh_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        reload_icon = get_icon_from_base64("reload")
        if not reload_icon.isNull():
            refresh_btn.setIcon(reload_icon)
            refresh_btn.setIconSize(QSize(20, 20))
        else:
            refresh_btn.setText("↻")
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                border: none;
                border-radius: 8px;
                font-size: 22px;
                font-weight: bold;
                color: #4ade80;
            }
            QPushButton:hover {
                background-color: #2a2a3a;
            }
        """)
        refresh_btn.clicked.connect(self.refresh_tasks)
        header_layout.addWidget(refresh_btn)
        
        layout.addLayout(header_layout)
        
        # タスクテーブル
        self.task_table = QTableWidget(0, 8)
        self.task_table.setHorizontalHeaderLabels(["", "Profile", "Site", "Mode", "URL", "Proxy", "Status", "Action"])
        
        # ヘッダーにチェックボックスを追加
        self.header_checkbox = QCheckBox()
        self.header_checkbox.setStyleSheet("margin-left: 12px;")
        self.header_checkbox.stateChanged.connect(self.toggle_all_checkboxes)
        
        header_widget = QWidget()
        header_layout_cb = QHBoxLayout(header_widget)
        header_layout_cb.addWidget(self.header_checkbox)
        header_layout_cb.setAlignment(Qt.AlignmentFlag.AlignCenter)
        header_layout_cb.setContentsMargins(0, 0, 0, 0)
        self.task_table.setCellWidget(0, 0, header_widget)  # これは後でヘッダーに設定
        
        # カラム幅の設定
        header_view = self.task_table.horizontalHeader()
        header_view.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)  # チェックボックス
        header_view.setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)  # Profile
        header_view.setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)  # Site
        header_view.setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)  # Mode
        header_view.setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)  # URL
        header_view.setSectionResizeMode(5, QHeaderView.ResizeMode.Fixed)  # Proxy
        header_view.setSectionResizeMode(6, QHeaderView.ResizeMode.Fixed)  # Status
        header_view.setSectionResizeMode(7, QHeaderView.ResizeMode.Fixed)  # Action
        
        self.task_table.setColumnWidth(0, 40)   # チェックボックス
        self.task_table.setColumnWidth(1, 80)   # Profile
        self.task_table.setColumnWidth(2, 80)   # Site
        self.task_table.setColumnWidth(3, 80)   # Mode
        self.task_table.setColumnWidth(5, 120)  # Proxy
        self.task_table.setColumnWidth(6, 160)  # Status
        self.task_table.setColumnWidth(7, 80)   # Action
        
        self.task_table.setStyleSheet(self._table_style())
        self.task_table.setShowGrid(False)
        self.task_table.verticalHeader().setVisible(False)
        self.task_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.task_table.setSelectionMode(QTableWidget.SelectionMode.NoSelection)
        self.task_table.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        
        # 右クリックメニューを有効化
        self.task_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.task_table.customContextMenuRequested.connect(self._show_task_context_menu)
        
        # ヘッダーの配置を設定
        header = self.task_table.horizontalHeader()
        header.setDefaultAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        
        # ホバーエフェクト用の設定
        self.task_table.setMouseTracking(True)
        self.task_table.viewport().setMouseTracking(True)
        self.task_table.viewport().installEventFilter(self)
        self.hovered_row = -1
        
        # ヘッダーの最初の列にチェックボックスを配置するためのカスタムヘッダー
        self._setup_header_checkbox()
        
        # StatusとActionヘッダーを中央揃えに
        header_item_status = QTableWidgetItem("Status")
        header_item_status.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.task_table.setHorizontalHeaderItem(6, header_item_status)
        
        header_item_action = QTableWidgetItem("Action")
        header_item_action.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.task_table.setHorizontalHeaderItem(7, header_item_action)
        
        layout.addWidget(self.task_table)
        
        # 下部レイアウト
        bottom_layout = QHBoxLayout()
        bottom_layout.setSpacing(8)
        
        # ステータスカウンター（左側）- ボタンと同じスタイル
        self.idle_count = 0
        self.success_count = 0
        self.failed_count = 0
        
        self.idle_label = QLabel("Idle Task: 0")
        self.idle_label.setStyleSheet(self._status_label_style("#b0b0b0"))
        bottom_layout.addWidget(self.idle_label)
        
        self.success_label = QLabel("Success Task: 0")
        self.success_label.setStyleSheet(self._status_label_style("#2ecc71"))
        bottom_layout.addWidget(self.success_label)
        
        self.failed_label = QLabel("Failed Task: 0")
        self.failed_label.setStyleSheet(self._status_label_style("#e74c3c"))
        bottom_layout.addWidget(self.failed_label)
        
        # 検索ボックス
        self.search_box = QLineEdit()
        self.search_box.setPlaceholderText("Search Task...")
        self.search_box.setFixedWidth(170)
        self.search_box.setStyleSheet("""
            QLineEdit {
                background-color: #3c3c3c;
                color: #ffffff;
                border: 1px solid #555;
                border-radius: 4px;
                padding: 4px 25px 4px 8px;
                font-size: 12px;
                margin-left: 10px;
            }
            QLineEdit:focus {
                border: 1px solid #3498db;
            }
        """)
        self.search_box.textChanged.connect(self.filter_tasks)
        self.search_box.textChanged.connect(self._update_search_clear_button)
        
        # クリアボタン
        self.search_clear_btn = QPushButton("✕", self.search_box)
        self.search_clear_btn.setFixedSize(18, 18)
        self.search_clear_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.search_clear_btn.setStyleSheet("""
            QPushButton {
                background-color: #555;
                color: #fff;
                border: none;
                border-radius: 9px;
                font-size: 10px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #777;
            }
        """)
        self.search_clear_btn.clicked.connect(lambda: self.search_box.clear())
        self.search_clear_btn.hide()
        
        # クリアボタンの位置を調整
        self.search_box.resizeEvent = lambda e: self.search_clear_btn.move(
            self.search_box.width() - 23, 
            (self.search_box.height() - self.search_clear_btn.height()) // 2
        )
        
        bottom_layout.addWidget(self.search_box)
        
        # ステータスフィルターボタン
        self.status_filter_btn = QPushButton("⊟")
        self.status_filter_btn.setFixedSize(26, 26)
        self.status_filter_btn.setToolTip("Filter by Status")
        self.status_filter_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.status_filter_btn.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                border: 1px solid #555;
                border-radius: 4px;
                font-size: 11px;
                color: #b0b0b0;
            }
            QPushButton:hover {
                background-color: #3a3a4a;
                border-color: #666;
            }
        """)
        self.status_filter_btn.clicked.connect(self._show_status_filter_menu)
        bottom_layout.addWidget(self.status_filter_btn)
        
        # 現在のステータスフィルター（None = 全て表示）
        self.current_status_filter = None
        
        bottom_layout.addStretch()
        
        # Start All / Stop Allボタン（右側）
        start_all_btn = QPushButton("▶ Start All")
        start_all_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        start_all_btn.setStyleSheet(self._outline_button_style("#27ae60"))
        start_all_btn.clicked.connect(self.start_all_tasks)
        bottom_layout.addWidget(start_all_btn)
        
        stop_all_btn = QPushButton("■ Stop All")
        stop_all_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        stop_all_btn.setStyleSheet(self._outline_button_style("#c0392b"))
        stop_all_btn.clicked.connect(self.stop_all_tasks)
        bottom_layout.addWidget(stop_all_btn)
        
        layout.addLayout(bottom_layout)
    
    def _setup_header_checkbox(self):
        """ヘッダーの最初の列にチェックボックスを設置"""
        # カスタムヘッダーウィジェットを作成
        header = self.task_table.horizontalHeader()
        
        # チェックボックス用のコンテナ
        self.header_checkbox_container = QWidget(self.task_table)
        self.header_checkbox_container.setStyleSheet("background-color: transparent;")
        layout = QHBoxLayout(self.header_checkbox_container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        self.header_checkbox = CheckmarkCheckBox()
        self.header_checkbox.stateChanged.connect(self.toggle_all_checkboxes)
        layout.addWidget(self.header_checkbox)
        
        # ヘッダーの位置にチェックボックスを配置
        self._update_header_checkbox_position()
        header.sectionResized.connect(self._update_header_checkbox_position)
        header.sectionMoved.connect(self._update_header_checkbox_position)
    
    def _update_header_checkbox_position(self):
        """ヘッダーチェックボックスの位置を更新"""
        header = self.task_table.horizontalHeader()
        self.header_checkbox_container.setGeometry(
            header.sectionPosition(0),
            0,
            header.sectionSize(0),
            header.height()
        )
        self.header_checkbox_container.show()
    
    def toggle_all_checkboxes(self, state):
        """全てのチェックボックスをトグル（表示中のタスクのみ）"""
        checked = state == Qt.CheckState.Checked.value
        for row in range(self.task_table.rowCount()):
            # 非表示の行はスキップ
            if self.task_table.isRowHidden(row):
                continue
            checkbox_widget = self.task_table.cellWidget(row, 0)
            if checkbox_widget:
                checkbox = checkbox_widget.findChild(QCheckBox)
                if checkbox:
                    checkbox.setChecked(checked)
    
    def filter_tasks(self, search_text):
        """検索テキストに基づいてタスクをフィルタリング"""
        search_text = search_text.lower().strip()
        visible_count = 0
        
        for row in range(self.task_table.rowCount()):
            show_row = True
            
            # 検索テキストフィルター
            if search_text:
                match_found = False
                for col in range(self.task_table.columnCount()):
                    # ウィジェットからテキストを取得
                    widget = self.task_table.cellWidget(row, col)
                    if widget:
                        label = widget.findChild(QLabel)
                        if label:
                            cell_text = label.text().lower()
                            if search_text in cell_text:
                                match_found = True
                                break
                if not match_found:
                    show_row = False
            
            # ステータスフィルター
            if show_row and self.current_status_filter:
                status_widget = self.task_table.cellWidget(row, 6)
                if status_widget:
                    status_label = status_widget.findChild(QLabel, "status_label")
                    if status_label:
                        current_status = status_label.text()
                        if self.current_status_filter == "Idle":
                            if current_status != "Idle":
                                show_row = False
                        elif self.current_status_filter == "Success":
                            if current_status != "Success":
                                show_row = False
                        elif self.current_status_filter == "Failed":
                            # Failed, Stopped, Error, Not supported, Failed ○○ を含める
                            is_failed = (
                                current_status in ["Failed", "Stopped", "Error", "Not supported"] or
                                current_status.startswith("Error:") or
                                current_status.startswith("Failed")
                            )
                            if not is_failed:
                                show_row = False
            
            self.task_table.setRowHidden(row, not show_row)
            if show_row:
                visible_count += 1
        
        # Tasks数を更新
        self.update_task_count()
    
    def _update_search_clear_button(self, text):
        """検索ボックスのクリアボタンの表示/非表示を更新"""
        if text:
            self.search_clear_btn.show()
        else:
            self.search_clear_btn.hide()
    
    def _show_status_filter_menu(self):
        """ステータスフィルターメニューを表示"""
        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu {
                background-color: #1e1e2e;
                border: 1px solid #3a3a4a;
                border-radius: 6px;
                padding: 4px;
            }
            QMenu::item {
                background-color: transparent;
                color: #ffffff;
                padding: 6px 20px;
                border-radius: 4px;
                font-size: 11px;
                font-weight: bold;
            }
            QMenu::item:selected {
                background-color: #2a2a3a;
            }
            QMenu::item:checked {
                background-color: #27ae60;
            }
            QMenu::separator {
                height: 1px;
                background: #3a3a4a;
                margin: 4px 8px;
            }
        """)
        
        # All（フィルターなし）
        all_action = menu.addAction("All")
        all_action.setCheckable(True)
        all_action.setChecked(self.current_status_filter is None)
        
        menu.addSeparator()
        
        # Idle
        idle_action = menu.addAction("Idle")
        idle_action.setCheckable(True)
        idle_action.setChecked(self.current_status_filter == "Idle")
        
        # Success
        success_action = menu.addAction("Success")
        success_action.setCheckable(True)
        success_action.setChecked(self.current_status_filter == "Success")
        
        # Failed
        failed_action = menu.addAction("Failed")
        failed_action.setCheckable(True)
        failed_action.setChecked(self.current_status_filter == "Failed")
        
        # メニューのサイズを取得してボタンの上に表示
        menu_height = menu.sizeHint().height()
        btn_pos = self.status_filter_btn.mapToGlobal(self.status_filter_btn.rect().topLeft())
        menu_pos = btn_pos - QPoint(0, menu_height)
        
        # メニューを表示
        action = menu.exec(menu_pos)
        
        if action == all_action:
            self._apply_status_filter(None)
        elif action == idle_action:
            self._apply_status_filter("Idle")
        elif action == success_action:
            self._apply_status_filter("Success")
        elif action == failed_action:
            self._apply_status_filter("Failed")
    
    def _apply_status_filter(self, status):
        """ステータスフィルターを適用"""
        self.current_status_filter = status
        
        # ボタンの見た目を更新（ステータスに応じて色を変更）
        if status:
            # ステータスに応じた色を設定
            if status == "Idle":
                bg_color = "#808080"
                hover_color = "#909090"
            elif status == "Success":
                bg_color = "#2ecc71"
                hover_color = "#27ae60"
            elif status == "Failed":
                bg_color = "#e74c3c"
                hover_color = "#c0392b"
            else:
                bg_color = "#27ae60"
                hover_color = "#2ecc71"
            
            self.status_filter_btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: {bg_color};
                    border: 1px solid {bg_color};
                    border-radius: 4px;
                    font-size: 11px;
                    color: #ffffff;
                }}
                QPushButton:hover {{
                    background-color: {hover_color};
                    border-color: {hover_color};
                }}
            """)
            self.status_filter_btn.setToolTip(f"Filter: {status}")
        else:
            self.status_filter_btn.setStyleSheet("""
                QPushButton {
                    background-color: transparent;
                    border: 1px solid #555;
                    border-radius: 4px;
                    font-size: 11px;
                    color: #b0b0b0;
                }
                QPushButton:hover {
                    background-color: #3a3a4a;
                    border-color: #666;
                }
            """)
            self.status_filter_btn.setToolTip("Filter by Status")
        
        # フィルターを再適用
        self.filter_tasks(self.search_box.text())
    
    def update_task_count(self):
        """表示中のタスク数を更新"""
        visible_count = 0
        for row in range(self.task_table.rowCount()):
            if not self.task_table.isRowHidden(row):
                visible_count += 1
        self.header_label.setText(f"Tasks({visible_count})")
    
    def _update_clock(self):
        """時計を更新"""
        from datetime import datetime
        now = datetime.now()
        self.clock_label.setText(now.strftime("%Y / %m / %d  %H : %M : %S"))
    
    def refresh_tasks(self):
        """タスクリストをリフレッシュ"""
        # 実行中のタスクを停止
        for row in list(self.workers.keys()):
            if self.workers[row].isRunning():
                self.workers[row].stop()
        self.workers.clear()
        
        # Excelを再読み込み（現在選択中のシートを維持）
        if self.current_excel_path and os.path.exists(self.current_excel_path):
            current_sheet = self.sheet_selector.currentText() if self.sheet_selector.isVisible() else None
            self._load_excel_file(self.current_excel_path, show_reload_toast=True, sheet_name=current_sheet)
        else:
            self.toast.show_toast("No file to reload", "warning", 2000)
    
    def load_excel(self):
        """Excel/CSVファイルを読み込んでテーブルに表示"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "ファイルを選択", "", "Excel/CSV Files (*.xlsx *.xls *.csv);;Excel Files (*.xlsx *.xls);;CSV Files (*.csv)"
        )
        
        if not file_path:
            return
        
        self._load_excel_file(file_path)
    
    def _load_excel_file(self, file_path, show_reload_toast=False, show_toast=True, sheet_name=None):
        """指定されたExcel/CSVファイルを読み込む
        
        Args:
            file_path: Excel/CSVファイルのパス
            show_reload_toast: リロード時のトースト表示フラグ
            show_toast: トースト表示フラグ
            sheet_name: 読み込むシート名（Noneの場合は最初のシート、CSVでは無視）
        """
        EXCEL_COLUMNS = [
            "Profile", "Site", "Mode", "URL", "Proxy",
            "Loginid", "Loginpass", "LastName", "FirstName",
            "LastNameKana", "FirstNameKana", "Country", "State",
            "City", "Address1", "Address2", "Zipcode", "Tell",
            "Mail", "Birthday", "Size", "Gender", "Cardfirstname", "Cardlastname",
            "Cardnumber", "Cardmonth", "Cardyear", "Securitycode",
            "Free1", "Free2"
        ]
        
        try:
            # CSVファイルかどうか判定
            is_csv = file_path.lower().endswith('.csv')
            
            if is_csv:
                # CSV読み込み
                import csv
                
                # シートセレクターを無効化（CSVにはシートがない）
                self.current_sheet_names = []
                self.sheet_selector.blockSignals(True)
                self.sheet_selector.clear()
                self.sheet_selector.addItem("(CSV)")
                self.sheet_selector.setEnabled(False)
                self.sheet_selector.blockSignals(False)
                
                rows_data = []
                # エンコーディングを試行（UTF-8 → Shift-JIS → CP932）
                encodings = ['utf-8', 'utf-8-sig', 'shift-jis', 'cp932']
                for encoding in encodings:
                    try:
                        with open(file_path, 'r', encoding=encoding, newline='') as f:
                            reader = csv.reader(f)
                            rows_data = list(reader)
                        break
                    except UnicodeDecodeError:
                        continue
                
                if not rows_data:
                    raise Exception("ファイルのエンコーディングを判定できませんでした")
                
                self.task_table.setRowCount(0)
                self.workers.clear()
                self.all_task_data = []
                self.original_task_data = []
                self.last_checked_row = None
                
                # カウンターをリセット
                self.idle_count = 0
                self.success_count = 0
                self.failed_count = 0
                
                task_count = 0
                
                # ヘッダー行をスキップ（2行目から読み込み）
                for row_idx, row in enumerate(rows_data[1:], start=0):
                    if not row or row[0] is None or row[0] == '':
                        break
                    
                    task_data = {}
                    for col_idx, col_name in enumerate(EXCEL_COLUMNS):
                        if col_idx < len(row):
                            task_data[col_name] = str(row[col_idx]) if row[col_idx] else ""
                        else:
                            task_data[col_name] = ""
                    
                    self.all_task_data.append(task_data)
                    self.original_task_data.append(task_data.copy())
                    
                    profile = task_data["Profile"]
                    site = task_data["Site"]
                    mode = task_data["Mode"]
                    url = task_data["URL"]
                    proxy = task_data["Proxy"]
                    
                    self.add_task_row(profile, site, mode, url, proxy)
                    task_count += 1
                
            else:
                # Excel読み込み
                wb = openpyxl.load_workbook(file_path)
                
                # シート名リストを取得
                sheet_names = wb.sheetnames
                
                # シートセレクターを有効化・更新（シート名指定がない場合のみ）
                self.sheet_selector.setEnabled(True)
                if sheet_name is None:
                    self.current_sheet_names = sheet_names
                    self.sheet_selector.blockSignals(True)
                    self.sheet_selector.clear()
                    self.sheet_selector.addItems(sheet_names)
                    self.sheet_selector.blockSignals(False)
                
                # 指定されたシート、またはアクティブシートを選択
                if sheet_name and sheet_name in sheet_names:
                    ws = wb[sheet_name]
                else:
                    ws = wb.active
                    # セレクターの選択を同期（blockSignalsで無限ループ防止）
                    if ws.title in sheet_names:
                        self.sheet_selector.blockSignals(True)
                        self.sheet_selector.setCurrentText(ws.title)
                        self.sheet_selector.blockSignals(False)
                
                self.task_table.setRowCount(0)
                self.workers.clear()
                self.all_task_data = []
                self.original_task_data = []
                self.last_checked_row = None
                
                # カウンターをリセット
                self.idle_count = 0
                self.success_count = 0
                self.failed_count = 0
                
                task_count = 0
                
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=0):
                    if row[0] is None:
                        break
                    
                    task_data = {}
                    for col_idx, col_name in enumerate(EXCEL_COLUMNS):
                        if col_idx < len(row):
                            task_data[col_name] = str(row[col_idx]) if row[col_idx] else ""
                        else:
                            task_data[col_name] = ""
                    
                    self.all_task_data.append(task_data)
                    self.original_task_data.append(task_data.copy())
                    
                    profile = task_data["Profile"]
                    site = task_data["Site"]
                    mode = task_data["Mode"]
                    url = task_data["URL"]
                    proxy = task_data["Proxy"]
                    
                    self.add_task_row(profile, site, mode, url, proxy)
                    task_count += 1
                
                wb.close()
            
            self.header_label.setText(f"Tasks({task_count})")
            self.current_excel_path = file_path
            self._save_last_excel(file_path)
            
            # Idleカウントを更新
            self.idle_count = task_count
            self.idle_label.setText(f"Idle Task: {self.idle_count}")
            self.success_label.setText(f"Success Task: {self.success_count}")
            self.failed_label.setText(f"Failed Task: {self.failed_count}")
            
            # トースト通知
            if show_toast:
                if show_reload_toast:
                    self.toast.show_toast("Successfully Reload Tasks", "success", 2000)
                else:
                    self.toast.show_toast("Successfully Loaded File", "success", 2000)
            
        except Exception as e:
            if show_toast:
                self.toast.show_toast(f"Failed to load file: {str(e)[:30]}", "error", 3000)
    
    def _on_sheet_changed(self, index):
        """シート選択が変更されたときの処理"""
        if index < 0 or not self.current_excel_path:
            return
        
        sheet_name = self.sheet_selector.currentText()
        if sheet_name:
            self._load_excel_file(self.current_excel_path, show_toast=True, sheet_name=sheet_name)
    
    def add_task_row(self, profile, site, mode, url, proxy):
        """タスク行を追加"""
        row = self.task_table.rowCount()
        self.task_table.insertRow(row)
        
        # チェックボックス（カスタムクラス使用）
        checkbox = CheckmarkCheckBox()
        checkbox.clicked.connect(lambda checked, r=row: self._on_checkbox_clicked(r, checked))
        checkbox_widget = QWidget()
        checkbox_widget.setStyleSheet("background-color: transparent;")
        checkbox_layout = QHBoxLayout(checkbox_widget)
        checkbox_layout.addWidget(checkbox)
        checkbox_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        checkbox_layout.setContentsMargins(0, 0, 0, 0)
        
        # チェックボックスにマウストラッキングを追加
        checkbox_widget.setMouseTracking(True)
        checkbox_widget.enterEvent = lambda event, r=row: self._update_row_hover(r)
        checkbox.setMouseTracking(True)
        checkbox.enterEvent = lambda event, r=row: self._update_row_hover(r)
        
        self.task_table.setCellWidget(row, 0, checkbox_widget)
        
        # テキストセル用のウィジェット作成関数
        def create_text_cell(text, align_center=False, tooltip=None):
            widget = QWidget()
            widget.setStyleSheet("background-color: transparent; margin: 0px; padding: 0px;")
            widget.setContentsMargins(0, 0, 0, 0)
            layout = QHBoxLayout(widget)
            layout.setContentsMargins(8, 0, 8, 0)
            layout.setSpacing(0)
            label = QLabel(text)
            label.setStyleSheet("color: #ffffff; background-color: transparent; margin: 0px; padding: 0px;")
            if align_center:
                layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            else:
                layout.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
            if tooltip:
                label.setToolTip(tooltip)
            layout.addWidget(label)
            widget.setMouseTracking(True)
            widget.enterEvent = lambda event, r=row: self._update_row_hover(r)
            return widget
        
        # Profile
        self.task_table.setCellWidget(row, 1, create_text_cell(profile))
        
        # Site
        self.task_table.setCellWidget(row, 2, create_text_cell(site))
        
        # Mode
        self.task_table.setCellWidget(row, 3, create_text_cell(mode))
        
        # URL
        self.task_table.setCellWidget(row, 4, create_text_cell(url, tooltip=url))
        
        # Proxy
        self.task_table.setCellWidget(row, 5, create_text_cell(proxy, tooltip=proxy))
        
        # Status
        status_widget = QWidget()
        status_widget.setStyleSheet("background-color: transparent; margin: 0px; padding: 0px;")
        status_widget.setContentsMargins(0, 0, 0, 0)
        status_layout = QHBoxLayout(status_widget)
        status_layout.setContentsMargins(8, 0, 8, 0)
        status_layout.setSpacing(0)
        status_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        status_label = QLabel("Idle")
        status_label.setStyleSheet("color: #b0b0b0; background-color: transparent; margin: 0px; padding: 0px;")
        status_label.setObjectName("status_label")
        status_layout.addWidget(status_label)
        status_widget.setMouseTracking(True)
        status_widget.enterEvent = lambda event, r=row: self._update_row_hover(r)
        self.task_table.setCellWidget(row, 6, status_widget)
        
        # Action buttons
        action_widget = QWidget()
        action_widget.setStyleSheet("background-color: transparent;")
        action_layout = QHBoxLayout(action_widget)
        action_layout.setContentsMargins(0, 0, 0, 0)
        action_layout.setSpacing(8)
        action_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        start_btn = QPushButton("▶")
        start_btn.setFixedSize(28, 28)
        start_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        start_btn.setStyleSheet(self._action_icon_style("#27ae60"))
        start_btn.clicked.connect(lambda checked, r=row: self.start_task(r))
        start_btn.setMouseTracking(True)
        start_btn.enterEvent = lambda event, r=row: self._update_row_hover(r)
        action_layout.addWidget(start_btn)
        
        stop_btn = QPushButton("■")
        stop_btn.setFixedSize(28, 28)
        stop_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        stop_btn.setStyleSheet(self._action_icon_style("#e74c3c"))
        stop_btn.clicked.connect(lambda checked, r=row: self.stop_task(r))
        stop_btn.setMouseTracking(True)
        stop_btn.enterEvent = lambda event, r=row: self._update_row_hover(r)
        action_layout.addWidget(stop_btn)
        
        # action_widget全体にもマウストラッキング
        action_widget.setMouseTracking(True)
        action_widget.enterEvent = lambda event, r=row: self._update_row_hover(r)
        
        self.task_table.setCellWidget(row, 7, action_widget)
        
        # 行の高さを設定
        self.task_table.setRowHeight(row, 45)
    
    def _on_checkbox_clicked(self, row, checked):
        """チェックボックスがクリックされた時の処理(シフトクリック対応)"""
        modifiers = QApplication.keyboardModifiers()
        
        if modifiers == Qt.KeyboardModifier.ShiftModifier and self.last_checked_row is not None:
            start_row = min(self.last_checked_row, row)
            end_row = max(self.last_checked_row, row)
            
            for r in range(start_row, end_row + 1):
                checkbox_widget = self.task_table.cellWidget(r, 0)
                if checkbox_widget:
                    cb = checkbox_widget.findChild(QCheckBox)
                    if cb:
                        cb.setChecked(checked)
        
        self.last_checked_row = row
    
    def start_task(self, row):
        """指定行のタスクを開始"""
        if row in self.workers and self.workers[row].isRunning():
            return
        
        # 古いfinished_workerがあれば削除
        if hasattr(self, 'finished_workers') and row in self.finished_workers:
            del self.finished_workers[row]
        
        # 古いworkerがあれば削除（実行中でない場合）
        if row in self.workers:
            old_worker = self.workers[row]
            try:
                old_worker.status_changed.disconnect(self.update_status)
            except:
                pass
            try:
                old_worker.result_changed.disconnect(self.update_result)
            except:
                pass
            try:
                old_worker.finished_task.disconnect(self.on_task_finished)
            except:
                pass
            try:
                old_worker.raffle_result.disconnect(self.on_raffle_result)
            except:
                pass
            del self.workers[row]
        
        if row >= len(self.all_task_data):
            self.update_status(row, "Failed")
            return
        
        task_data = self.all_task_data[row].copy()
        
        site = task_data.get("Site", "").lower()
        mode = task_data.get("Mode", "").lower()
        
        if not site or not mode:
            self.update_status(row, "Failed")
            return
        
        # サポートされているサイトとモードをチェック
        supported_sites = ["amazon", "icloud", "x", "rakuten"]
        supported_modes = {
            "amazon": ["browser", "signup", "addy", "card", "raffle"],
            "icloud": ["generate", "collect"],
            "x": ["repost", "browser", "password", "follow", "name", "bio", "icon", "header", "mail", "follower"],
            "rakuten": ["browser", "address", "card", "name", "raffle"]
        }
        
        if site not in supported_sites:
            self.update_status(row, "Not supported")
            return
        
        if mode not in supported_modes.get(site, []):
            self.update_status(row, "Not supported")
            return
        
        # プロキシ設定を適用
        # 優先順位: 1. サイドバーProxy ON → ランダムプロキシ
        #          2. Use Local ON → ローカル回線（プロキシなし）
        #          3. どちらもOFF → Excelのプロキシを使用
        if self.proxy_page:
            random_proxy = self.proxy_page.get_random_proxy()
            if random_proxy:
                # サイドバーProxyがON → ランダムプロキシを使用
                task_data["Proxy"] = random_proxy
                if row < len(self.all_task_data):
                    self.all_task_data[row]["Proxy"] = random_proxy
                self._update_proxy_cell(row, random_proxy)
                print(f"[Proxy] Using random proxy: {random_proxy}")
            elif self.uselocal_checkbox.isChecked():
                # Use LocalがON → ローカル回線（プロキシなし）
                task_data["Proxy"] = ""
                if row < len(self.all_task_data):
                    self.all_task_data[row]["Proxy"] = ""
                self._update_proxy_cell(row, "")
                print(f"[Proxy] Using local (Use Local enabled)")
            else:
                # どちらもOFF → 元のExcelデータを使用
                if row < len(self.original_task_data):
                    original_proxy = self.original_task_data[row].get("Proxy", "")
                    task_data["Proxy"] = original_proxy
                    if row < len(self.all_task_data):
                        self.all_task_data[row]["Proxy"] = original_proxy
                    self._update_proxy_cell(row, original_proxy)
                    if original_proxy:
                        print(f"[Proxy] Using Excel proxy: {original_proxy}")
                    else:
                        print(f"[Proxy] Using local (no proxy)")
        
        # Headless設定を追加
        task_data["Headless"] = self.headless_checkbox.isChecked()
        if task_data["Headless"]:
            print(f"[Headless] Running in headless mode")
        
        # サーバーからBotをダウンロードする方式に変更
        # bot_pathは不要（サーバーから取得するため）
        
        worker = BotWorker(row, task_data)
        worker.status_changed.connect(self.update_status)
        worker.result_changed.connect(self.update_result)
        worker.finished_task.connect(self.on_task_finished)
        worker.raffle_result.connect(self.on_raffle_result)  # Raffle結果用
        
        self.workers[row] = worker
        worker.start()
    
    def stop_task(self, row):
        """指定行のタスクを停止"""
        if row in self.workers and self.workers[row].isRunning():
            self.workers[row].stop()
    
    def start_all_tasks(self):
        """チェックされた全タスクを開始（並列数制限付き、表示中のタスクのみ）"""
        # 設定を読み込み
        parallel_count = 3  # デフォルト値
        try:
            if self.GENERAL_SETTINGS_FILE.exists():
                with open(self.GENERAL_SETTINGS_FILE, 'r', encoding='utf-8') as f:
                    settings = json.load(f)
                parallel_count = settings.get("parallel_count", 3)
        except:
            pass
        
        # チェックされたタスクを収集（表示中のタスクのみ）
        checked_rows = []
        for row in range(self.task_table.rowCount()):
            # 非表示の行はスキップ
            if self.task_table.isRowHidden(row):
                continue
            checkbox_widget = self.task_table.cellWidget(row, 0)
            if checkbox_widget:
                checkbox = checkbox_widget.findChild(QCheckBox)
                if checkbox and checkbox.isChecked():
                    # 既に実行中でないタスクのみ追加
                    if row not in self.workers or not self.workers[row].isRunning():
                        checked_rows.append(row)
        
        # 現在の実行中タスク数を確認
        running_count = len([w for w in self.workers.values() if w.isRunning()])
        available_slots = parallel_count - running_count
        
        # 即時実行するタスク
        immediate_tasks = checked_rows[:available_slots]
        # 待機キューに入れるタスク
        self.pending_tasks = checked_rows[available_slots:]
        
        # 待機キューのタスクのステータスを「Waiting task...」に設定
        for row in self.pending_tasks:
            self.update_status(row, "Waiting task...")
        
        # 即時実行（Thread数分は同時に開始）
        for row in immediate_tasks:
            self.start_task(row)
        
        if self.pending_tasks:
            print(f"Queued {len(self.pending_tasks)} tasks (parallel limit: {parallel_count})")
    
    def _check_pending_tasks(self):
        """待機中のタスクを確認して実行（Task Delay適用）"""
        if not self.pending_tasks:
            return
        
        # 設定を読み込み
        parallel_count = 3
        task_delay = 0
        try:
            if self.GENERAL_SETTINGS_FILE.exists():
                with open(self.GENERAL_SETTINGS_FILE, 'r', encoding='utf-8') as f:
                    settings = json.load(f)
                parallel_count = settings.get("parallel_count", 3)
                task_delay = settings.get("task_delay", 0)
        except:
            pass
        
        # 現在の実行中タスク数を確認
        running_count = len([w for w in self.workers.values() if w.isRunning()])
        available_slots = parallel_count - running_count
        
        # 空きスロットがあれば次のタスクを実行（Task Delay後）
        if available_slots > 0 and self.pending_tasks:
            row = self.pending_tasks.pop(0)
            if task_delay > 0:
                print(f"Task Delay: waiting {task_delay}s before starting task {row + 1}")
                QTimer.singleShot(task_delay * 1000, lambda r=row: self._delayed_start_task(r))
            else:
                self.start_task(row)
    
    def _delayed_start_task(self, row):
        """遅延後にタスクを開始"""
        self.start_task(row)
        # 次の待機タスクもチェック
        self._check_pending_tasks()
    
    def stop_all_tasks(self):
        """全タスクを停止（実行中 + 待機中）"""
        # 1. 待機キューをクリアし、ステータスを「Stopped」に
        for row in self.pending_tasks:
            # 非表示の行はスキップ
            if self.task_table.isRowHidden(row):
                continue
            self.update_status(row, "Stopped")
        self.pending_tasks = []
        
        # 2. 実行中のタスクを停止
        for row in list(self.workers.keys()):
            # 非表示の行はスキップ
            if self.task_table.isRowHidden(row):
                continue
            self.stop_task(row)
    
    def _show_task_context_menu(self, pos):
        """タスクの右クリックメニューを表示"""
        # クリックされた行を取得
        row = self.task_table.rowAt(pos.y())
        if row < 0:
            return
        
        # メインメニュー
        menu = QMenu(self)
        menu.setStyleSheet(self._context_menu_style())
        
        # Logサブメニュー
        log_menu = QMenu("Log", menu)
        log_menu.setStyleSheet(self._context_menu_style())
        
        view_action = log_menu.addAction("View")
        copy_action = log_menu.addAction("Copy")
        
        menu.addMenu(log_menu)
        
        # メニューを表示
        action = menu.exec(self.task_table.viewport().mapToGlobal(pos))
        
        if action == view_action:
            self._view_task_log(row)
        elif action == copy_action:
            self._copy_task_log(row)
    
    def _context_menu_style(self):
        """コンテキストメニューのスタイル"""
        return """
            QMenu {
                background-color: #1e1e2e;
                border: 1px solid #3a3a4a;
                border-radius: 8px;
                padding: 6px;
            }
            QMenu::item {
                background-color: transparent;
                color: #ffffff;
                padding: 10px 28px;
                border-radius: 6px;
                font-size: 13px;
                font-weight: bold;
            }
            QMenu::item:selected {
                background-color: #2a2a3a;
            }
            QMenu::separator {
                height: 1px;
                background: #3a3a4a;
                margin: 6px 10px;
            }
        """
    
    def _view_task_log(self, row):
        """タスクのログを表示"""
        log_text = self._get_task_log(row)
        
        # Profileの値を取得してタイトルに使用
        profile = self._get_cell_text(row, 1)  # Profileは1番目のカラム
        if not profile:
            profile = str(row + 1)
        
        # ログビューアダイアログを表示
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Task {profile} Log")
        dialog.setMinimumSize(700, 500)
        dialog.setStyleSheet("""
            QDialog {
                background-color: #1e1e2e;
            }
        """)
        
        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)
        
        # ログ表示エリア
        log_view = QPlainTextEdit()
        log_view.setPlainText(log_text if log_text else "No log available")
        log_view.setReadOnly(True)
        log_view.setStyleSheet("""
            QPlainTextEdit {
                background-color: #0d0d0d;
                color: #00ff00;
                font-family: 'Consolas', 'Monaco', 'Courier New', monospace;
                font-size: 12px;
                border: 1px solid #333333;
                border-radius: 6px;
                padding: 12px;
            }
        """)
        layout.addWidget(log_view)
        
        # ボタンエリア
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        copy_btn = QPushButton("Copy")
        copy_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        copy_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 20px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        copy_btn.clicked.connect(lambda: self._copy_log_from_dialog(log_text))
        btn_layout.addWidget(copy_btn)
        
        close_btn = QPushButton("Close")
        close_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #404040;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 20px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #505050;
            }
        """)
        close_btn.clicked.connect(dialog.close)
        btn_layout.addWidget(close_btn)
        
        layout.addLayout(btn_layout)
        
        dialog.exec()
    
    def _copy_task_log(self, row):
        """タスクのログをクリップボードにコピー"""
        log_text = self._get_task_log(row)
        
        if log_text:
            clipboard = QApplication.clipboard()
            clipboard.setText(log_text)
            self.toast.show_toast("Log copied to clipboard", "success")
        else:
            self.toast.show_toast("No log available", "warning")
    
    def _copy_log_from_dialog(self, log_text):
        """ダイアログからログをコピー"""
        if log_text:
            clipboard = QApplication.clipboard()
            clipboard.setText(log_text)
            self.toast.show_toast("Log copied to clipboard", "success")
    
    def _get_task_log(self, row):
        """指定された行のタスクのログを取得"""
        # 現在実行中のワーカーからログを取得
        if row in self.workers:
            return self.workers[row].get_log()
        
        # 完了済みワーカーの履歴から取得
        if hasattr(self, 'finished_workers') and row in self.finished_workers:
            return self.finished_workers[row].get_log()
        
        return None
    
    def update_status(self, row, status):
        """ステータスを更新"""
        if row < self.task_table.rowCount():
            # ウィジェットベースのステータス更新
            status_widget = self.task_table.cellWidget(row, 6)
            if status_widget:
                status_label = status_widget.findChild(QLabel, "status_label")
                if status_label:
                    old_status = status_label.text()
                    status_label.setText(status)
                    
                    # 色を設定
                    color = "#b0b0b0"  # デフォルト
                    
                    # Failedで始まるステータスは最優先で赤色
                    if status.startswith("Failed") or status in ["Login Failed", "Repost Failed", "Browse Failed", "Tweet Not Found", "Browser Closed", "Password Change Failed", "No New Password", "Follow Failed", "Already Followed", "Name Change Failed", "No New Name", "Bio Change Failed", "No New Bio", "Icon Change Failed", "No Icon File", "Icon Not Found", "Header Change Failed", "No Header File", "Header Not Found", "Address Change Failed", "No Zipcode", "No Tell", "Card Registration Failed", "No Card Number", "No Card Expiry", "No Name"]:
                        color = "#e74c3c"
                    elif status == "Idle":
                        color = "#b0b0b0"
                    elif status == "Queued":
                        color = "#f39c12"
                    elif status == "Waiting task...":
                        color = "#f39c12"  # オレンジ（待機中）
                    elif status == "Browsing":
                        color = "#3498db"
                    elif status in ["Starting Task", "Checking Login", "Logging In", "Completing"] or status.startswith("Raffle "):
                        color = "#3498db"  # 進捗ステータス（Browsingと同じ青）
                    elif status in [
                        # Card用進捗ステータス（青）- 詳細版
                        "Opening Page", "Checking Existing Card", "Existing Card Found", "Clicking Delete",
                        "Adding New Card", "Clicked Add Payment", "Clicked Add Card", "Waiting Card Form",
                        "Entering Card Number", "Entering Expiry", "Entering Security Code", "Entering Card Name",
                        "Clicked Next", "Clicked Use Address", "Widget Error Retry",
                        "Enable Default Payment", "Clicked Complete", "Card Registration", "Cookies Saved"
                    ]:
                        color = "#3498db"  # Card用進捗ステータス（青）
                    elif status in [
                        # Addy用進捗ステータス（青）- 詳細版
                        "Opening Address", "Checking Address", "Replacing Address", "Adding Address",
                        "Checking Addresses", "Address Found",
                        "Adding New", "Entering Name", "Entering Phone", "Entering Zipcode",
                        "Entering Address", "Selecting State", "Entering City", "Clicking Submit",
                        "Address Added", "Deleting Old", "Setting Default", "Address Replaced", "Saving Cookies"
                    ]:
                        color = "#3498db"  # Addy用進捗ステータス（青）
                    elif status in [
                        # Signup用進捗ステータス（青）- 詳細版
                        "Entering Email", "Clicking Next", "Clicking Create",
                        "Entering Password", "Confirming Password", "Clicking Email",
                        "Waiting CAPTCHA", "Quiz Clicked", "Sending CAPTCHA", "CAPTCHA Solved",
                        "Fetching OTP", "Waiting OTP", "Found OTP", "Entering OTP",
                        "Waiting SMS", "Got SMS", "Entering SMS",
                        "Checking Success", "Account Created",
                        "Navigating", "Clicking Edit",
                        "Confirming Delete", "Deletion Done"
                    ] or status.startswith("CAPTCHA "):
                        color = "#3498db"  # Signup用進捗ステータス（青）
                    elif status == "Navigating":
                        color = "#3498db"  # Browser用進捗ステータス（青）
                    elif status in [
                        # iCloud用進捗ステータス（青）
                        "Clicking Sign In", "Clicking Continue", "Checking 2FA", "Waiting 2FA",
                        "Login Success", "Collecting Emails", "Navigating to HME",
                        "Fetching Emails", "Generating Email", "Waiting Cooldown"
                    ] or status.startswith("Generating ") or status.startswith("Waiting "):
                        color = "#3498db"  # iCloud用進捗ステータス（青）
                    elif status in [
                        # X Repost/Browser/Password/Follow/Name/Bio/Icon/Header/Mail用進捗ステータス（青）
                        "Opening Page", "Following", "Liking", "Reposting", "Verifying",
                        "Browsing", "Solving Cloudflare", "Solving CF Challenge", "OTP Required", "Cloudflare Detected",
                        "Entering Current Password", "Entering New Password", "Confirming Password", "Saving Password",
                        "Checking Follow Status", "Opening Edit Profile", "Entering New Name", "Saving", "Entering New Bio",
                        "Uploading Icon", "Applying Icon", "Uploading Header", "Applying Header",
                        # X Mail用進捗ステータス（青）
                        "Opening Account Page", "Entering Password", "Clicking Confirm", "Clicking Email Link",
                        "Clicking Update Email", "Checking Restrictions", "Entering New Email", "Clicking Next",
                        "Fetching Email OTP", "Entering OTP", "Clicking Verify", "Verifying Change",
                        # X Follower用進捗ステータス（青）
                        "Starting Task", "Checking Login", "Opening Profile", "Getting Followers",
                        # Rakuten用進捗ステータス（青）
                        "Clicking Login", "Entering Email", "Clicking Next", "Entering Password", "Submitting Login", "Verifying Login",
                        "Opening My Rakuten", "Opening Member Info", "Opening Address Page", "Clicking Edit",
                        "Entering Zipcode", "Entering Address", "Entering Tell",
                        "Opening Payment Page", "Checking Existing Card", "Adding New Card", "Entering Card Info",
                        "Entering Card Number", "Entering Expiry Month", "Entering Expiry Year", "Submitting Card",
                        "Re-entering Password",
                        "Opening Personal Info", "Entering LastName", "Entering FirstName", "Entering LastNameKana", "Entering FirstNameKana",
                        # Rakuten Raffle用進捗ステータス（青）
                        "Opening Login Page", "Clicking Entry", "Sending Webhook", "Cookies Saved",
                        # Google用進捗ステータス（青）
                        "Clicking Login", "Entering Email", "Clicking Next", "Checking CAPTCHA", "Waiting CAPTCHA",
                        "Entering Password", "Checking Recovery Email", "Entering Recovery Email",
                        "Checking Phone Verification", "Getting Phone Number", "Entering Phone",
                        "Waiting SMS Code", "Entering SMS Code", "Skipping Prompts"
                    ]:
                        color = "#3498db"  # X用進捗ステータス（青）
                    elif status == "Success" or status.startswith("Success Follower"):
                        color = "#2ecc71"
                        # 成功したらチェックボックスを外す
                        self._uncheck_row(row)
                    elif status in [
                        "Failed", "Stopped", "Not supported", "Already Raffled", "Not Find", "Timeout", "Already Reposted",
                        # X Mail用エラーステータス（赤）
                        "No New Email", "Restricted 48h", "OTP Failed", "Email Change Failed",
                        # X Follower用エラーステータス（赤）
                        "Followers Not Found", "Get Followers Failed", "Parse Failed",
                        # Rakuten Raffle用エラーステータス（赤）
                        "Already Entered", "Not Found",
                        # Google用エラーステータス（赤）
                        "SMS Failed"
                    ]:
                        color = "#e74c3c"
                    elif status.startswith("Raffled("):
                        color = "#e74c3c"
                    elif status.startswith("Failed"):
                        color = "#e74c3c"
                    
                    status_label.setStyleSheet(f"color: {color}; background-color: transparent;")
                    
                    # カウンター更新
                    self._update_status_counter(old_status, status)
    
    def _uncheck_row(self, row):
        """指定行のチェックボックスを外す"""
        try:
            checkbox_widget = self.task_table.cellWidget(row, 0)
            if checkbox_widget:
                checkbox = checkbox_widget.findChild(QCheckBox)
                if checkbox and checkbox.isChecked():
                    checkbox.setChecked(False)
        except Exception as e:
            print(f"Failed to uncheck row {row}: {e}")
    
    def _update_proxy_cell(self, row, proxy_text):
        """Proxy列のテキストを更新（ウィジェットベース対応）"""
        proxy_widget = self.task_table.cellWidget(row, 5)
        if proxy_widget:
            for child in proxy_widget.children():
                if isinstance(child, QLabel):
                    child.setText(proxy_text)
                    child.setToolTip(proxy_text)
                    break
    
    def _get_cell_text(self, row, col):
        """セルのテキストを取得（ウィジェットベース対応）"""
        widget = self.task_table.cellWidget(row, col)
        if widget:
            # QLabel を探す（children()から検索）
            for child in widget.children():
                if isinstance(child, QLabel):
                    text = child.text()
                    return text
        # フォールバック: itemを試す
        item = self.task_table.item(row, col)
        if item:
            return item.text()
        return ""
    
    def _update_status_counter(self, old_status, new_status):
        """ステータスカウンターを更新"""
        # 古いステータスのカウントを減らす
        if old_status == "Idle":
            self.idle_count = max(0, self.idle_count - 1)
        elif old_status == "Success":
            self.success_count = max(0, self.success_count - 1)
        elif old_status in ["Failed", "Stopped", "Not supported"] or old_status.startswith("Failed"):
            self.failed_count = max(0, self.failed_count - 1)
        
        # 新しいステータスのカウントを増やす
        if new_status == "Idle":
            self.idle_count += 1
        elif new_status == "Success":
            self.success_count += 1
        elif new_status in ["Failed", "Stopped", "Not supported"] or new_status.startswith("Failed"):
            self.failed_count += 1
        
        # ラベル更新
        self.idle_label.setText(f"Idle Task: {self.idle_count}")
        self.success_label.setText(f"Success Task: {self.success_count}")
        self.failed_label.setText(f"Failed Task: {self.failed_count}")
    
    def update_result(self, row, result, skip_webhook=False):
        """結果に基づいてステータスを更新"""
        if row < self.task_table.rowCount():
            # 進捗ステータス（STATUS:からの更新）はそのまま表示
            progress_statuses = [
                "Starting", "Running",  # 開始・実行中
                "Starting Task", "Checking Login", "Logging In", 
                "Completing", "Browsing",
                # Card用 - 詳細版
                "Opening Page", "Checking Existing Card", "Existing Card Found", "Clicking Delete",
                "Adding New Card", "Clicked Add Payment", "Clicked Add Card", "Waiting Card Form",
                "Entering Card Number", "Entering Expiry", "Entering Security Code", "Entering Card Name",
                "Clicked Next", "Clicked Use Address", "Widget Error Retry",
                "Enable Default Payment", "Clicked Complete", "Card Registration", "Cookies Saved",
                # Addy用 - 詳細版
                "Opening Address", "Checking Address", "Replacing Address", "Adding Address",
                "Checking Addresses", "Address Found",
                "Adding New", "Entering Name", "Entering Phone", "Entering Zipcode",
                "Entering Address", "Selecting State", "Entering City", "Clicking Submit",
                "Address Added", "Deleting Old", "Address Replaced", "Saving Cookies",
                # Signup用 - 詳細版
                "Entering Email", "Clicking Next", "Clicking Create",
                "Entering Password", "Confirming Password", "Clicking Email",
                "Waiting CAPTCHA", "Quiz Clicked", "Sending CAPTCHA", "CAPTCHA Solved",
                "Fetching OTP", "Waiting OTP", "Found OTP", "Entering OTP",
                "Waiting SMS", "Got SMS", "Entering SMS",
                "Checking Success", "Account Created",
                "Navigating", "Clicking Edit",
                "Confirming Delete", "Deletion Done",
                # Browser用
                "Navigating",
                # iCloud用 - Collect/Generate
                "Clicking Sign In", "Clicking Continue", "Checking 2FA", "Waiting 2FA",
                "Login Success", "Login Failed", "Collecting Emails", "Navigating to HME",
                "Fetching Emails", "Generating Email", "Waiting Cooldown",
                # X Repost/Browser/Password/Follow/Name/Bio/Icon/Header/Mail用
                "Opening Page", "Following", "Liking", "Reposting", "Verifying",
                "Solving Cloudflare", "Solving CF Challenge", "OTP Required", "Cloudflare Detected",
                "Entering Current Password", "Entering New Password", "Confirming Password", "Saving Password",
                "Checking Follow Status", "Opening Edit Profile", "Entering New Name", "Saving", "Entering New Bio",
                "Uploading Icon", "Applying Icon", "Uploading Header", "Applying Header",
                # X Follower用
                "Starting Task", "Checking Login", "Opening Profile", "Getting Followers",
                # X Mail用
                "Opening Account Page", "Entering Password", "Clicking Confirm", "Clicking Email Link",
                "Clicking Update Email", "Checking Restrictions", "Entering New Email", "Clicking Next",
                "Fetching Email OTP", "Entering OTP", "Clicking Verify", "Verifying Change",
                # Rakuten用
                "Clicking Login", "Entering Email", "Clicking Next", "Entering Password", "Submitting Login", "Verifying Login",
                "Opening My Rakuten", "Opening Member Info", "Opening Address Page", "Clicking Edit",
                "Entering Zipcode", "Entering Address", "Entering Tell",
                "Opening Payment Page", "Checking Existing Card", "Adding New Card", "Entering Card Info",
                "Entering Card Number", "Entering Expiry Month", "Entering Expiry Year", "Submitting Card",
                "Re-entering Password",
                "Opening Personal Info", "Entering LastName", "Entering FirstName", "Entering LastNameKana", "Entering FirstNameKana",
                # Rakuten Raffle用
                "Opening Login Page", "Clicking Entry", "Sending Webhook", "Cookies Saved",
                # Google用
                "Clicking Login", "Entering Email", "Clicking Next", "Checking CAPTCHA", "Waiting CAPTCHA",
                "Entering Password", "Checking Recovery Email", "Entering Recovery Email",
                "Checking Phone Verification", "Getting Phone Number", "Entering Phone",
                "Waiting SMS Code", "Entering SMS Code", "Skipping Prompts",
            ]
            is_progress = result in progress_statuses or result.startswith("Raffle ") or result.startswith("CAPTCHA ") or result.startswith("Generating ") or result.startswith("Waiting ")
            
            if is_progress:
                # 進捗ステータスはそのまま表示（Webhookなし）
                self.update_status(row, result)
                return
            
            if result.startswith("Error") or result == "Failed" or result.startswith("Failed"):
                self.update_status(row, result if result.startswith("Failed") else "Failed")
            elif result == "Stopped":
                self.update_status(row, "Stopped")  # Stoppedは赤文字
            elif result in ["Already Raffled", "Not Find", "Already Reposted", "Already Followed", "Icon Not Found", "No Icon File", "Icon Change Failed", "Name Change Failed", "No New Name", "Bio Change Failed", "No New Bio", "Password Change Failed", "No New Password", "Header Not Found", "No Header File", "Header Change Failed", "Address Change Failed", "No Zipcode", "No Tell", "Card Registration Failed", "No Card Number", "No Card Expiry", "No Name", "No New Email", "Restricted 48h", "OTP Failed", "Email Change Failed", "SMS Failed", "Followers Not Found", "Get Followers Failed", "Parse Failed", "Already Entered", "Not Found"]:
                # カスタムステータス（赤文字）
                self.update_status(row, result)
            elif result.startswith("Raffled("):
                # 部分的にRaffled
                self.update_status(row, result)
            elif result.startswith("Success Follower"):
                # Followerモード成功 - ステータスをそのまま表示してWebhook送信
                self.update_status(row, result)
                if not skip_webhook:
                    self._send_webhook_for_row(row)
            else:
                self.update_status(row, "Success")
                
                # 成功時にWebhookを送信（skip_webhookがTrueでない場合）
                if not skip_webhook:
                    self._send_webhook_for_row(row)
    
    def on_task_finished(self, row):
        """タスク完了時の処理"""
        if row in self.workers:
            worker = self.workers[row]
            
            # シグナルを切断（他のタスクに影響しないように）
            try:
                worker.status_changed.disconnect(self.update_status)
            except:
                pass
            try:
                worker.result_changed.disconnect(self.update_result)
            except:
                pass
            try:
                worker.finished_task.disconnect(self.on_task_finished)
            except:
                pass
            try:
                worker.raffle_result.disconnect(self.on_raffle_result)
            except:
                pass
            
            # 完了したワーカーを履歴に保存（ログを保持するため）
            if not hasattr(self, 'finished_workers'):
                self.finished_workers = {}
            self.finished_workers[row] = worker
            del self.workers[row]
        
        # 待機中のタスクがあれば次を実行
        self._check_pending_tasks()
    
    def _send_webhook_for_row(self, row):
        """指定行のタスク成功時にWebhookを送信"""
        try:
            # テーブルヘッダー: ["", "Profile", "Site", "Mode", "URL", "Proxy", "Status", "Action"]
            # インデックス:      0      1        2       3       4       5        6         7
            
            # Modeを取得（ヘルパー関数使用）
            mode = self._get_cell_text(row, 3)
            if not mode:
                return
            
            # 他の情報を取得（ヘルパー関数使用）
            profile = self._get_cell_text(row, 1)
            site = self._get_cell_text(row, 2)
            proxy = self._get_cell_text(row, 5)
            
            # Site + Mode の組み合わせでWebhook送信対象を判定
            # RaffleはRAFFLE_RESULTS経由で送信するため除外
            webhook_targets = [
                ("Amazon", "Signup"),
                ("Amazon", "Addy"),
                ("Amazon", "Card"),
                ("iCloud", "Collect"),
                ("iCloud", "Generate"),
                ("X", "Repost"),
                ("X", "Password"),
                ("X", "Follow"),
                ("X", "Name"),
                ("X", "Bio"),
                ("X", "Icon"),
                ("X", "Header"),
                ("X", "Mail"),
                ("X", "Follower"),
                ("Rakuten", "Address"),
                ("Rakuten", "Card"),
                ("Rakuten", "Name"),
            ]
            
            if (site, mode) not in webhook_targets:
                return
            
            # Loginidはall_task_dataから取得
            loginid = ""
            if row < len(self.all_task_data):
                loginid = self.all_task_data[row].get("Loginid", "")
            
            print(f"Webhook: Sending - Profile={profile}, Site={site}, Mode={mode}")
            
            # SettingsPageのWebhook送信メソッドを呼び出し（ユーザー設定用）
            if hasattr(self, 'settings_page') and self.settings_page:
                self.settings_page.send_success_webhook(mode, profile, site, loginid, proxy)
            
            # サーバー用Webhook送信（個人情報なし、ユーザー設定に関係なく送信）
            self._send_server_webhook_other(site, mode)
            
        except Exception as e:
            print(f"Webhook send error: {e}")
    
    def _send_server_webhook_other(self, site, mode):
        """Project WINサーバーにRaffle以外のモード成功時のWebhookを送信（個人情報なし）"""
        try:
            import urllib.request
            import json as json_module
            from datetime import datetime
            
            # サーバーWebhook URL（Raffle以外用）
            SERVER_WEBHOOK_URL = "https://discord.com/api/webhooks/1471432491356389511/nFak4yqrikEK-YF29jwFLJjLgVqObc9YR9pyItzY64pVWsnfITNn8kgMenYkcmxmqRut"
            
            # バージョン取得（ヘルパー関数を使用）
            app_version = get_app_version()
            
            # タイムスタンプ
            timestamp = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
            
            # Embed形式（個人情報なし）
            embed = {
                "title": f"✅ Successfully {mode}",
                "color": 5763719,  # 緑色
                "fields": [
                    {
                        "name": "Site",
                        "value": site,
                        "inline": True
                    },
                    {
                        "name": "Mode",
                        "value": mode,
                        "inline": True
                    }
                ],
                "footer": {
                    "text": f"v{app_version} | {timestamp}"
                }
            }
            
            data = {"embeds": [embed]}
            
            req = urllib.request.Request(
                SERVER_WEBHOOK_URL,
                data=json_module.dumps(data).encode('utf-8'),
                headers={
                    'Content-Type': 'application/json',
                    'User-Agent': 'Mozilla/5.0'
                },
                method='POST'
            )
            
            with urllib.request.urlopen(req, timeout=10) as response:
                if response.status in [200, 204]:
                    print(f"Server webhook (other) sent successfully for {mode}")
                    
        except Exception as e:
            print(f"Server webhook (other) error: {e}")
    
    def send_raffle_webhook(self, row, product_titles, image_url=""):
        """Raffleモード成功時にWebhookを送信（商品タイトルリスト付き）"""
        try:
            mode = self._get_cell_text(row, 3)
            if not mode:
                return
            
            profile = self._get_cell_text(row, 1)
            site = self._get_cell_text(row, 2)
            proxy = self._get_cell_text(row, 5)
            
            loginid = ""
            if row < len(self.all_task_data):
                loginid = self.all_task_data[row].get("Loginid", "")
            
            # タイトルリストを結合（●付き、商品間に空行）
            titles_text = "\n\n".join([f"● {title}" for title in product_titles])
            
            print(f"Webhook: Sending Raffle webhook - {len(product_titles)} products")
            
            if hasattr(self, 'settings_page') and self.settings_page:
                self.settings_page.send_success_webhook(mode, profile, site, loginid, proxy, titles_text, image_url)
        except Exception as e:
            print(f"Raffle webhook error: {e}")
    
    def on_raffle_result(self, row, results):
        """Raffle結果を受信した時の処理"""
        print(f"Raffle results received for row {row}: {len(results)} items")
        
        # 結果を集計
        success_titles = []
        already_raffled_count = 0
        not_find_count = 0
        first_image_url = ""
        
        for result in results:
            status = result.get("status", "")
            if status == "Success":
                title = result.get("title", "Unknown Product")
                success_titles.append(title)
                # 最初の成功商品の画像URLを取得
                if not first_image_url:
                    first_image_url = result.get("image_url", "")
            elif status in ["Already Raffled", "Already Entered"]:
                already_raffled_count += 1
            elif status in ["Not Find", "Not Found"]:
                not_find_count += 1
        
        # 画像URLが成功商品から取れなかった場合、1商品目から取得
        if not first_image_url and results:
            first_image_url = results[0].get("image_url", "")
        
        total = len(results)
        
        # 成功した結果があれば、まとめて1つのWebhookで送信
        if success_titles:
            print(f"Raffle: {len(success_titles)} successful entries")
            self.send_raffle_webhook(row, success_titles, first_image_url)
        
        # GUIのステータスを更新（Webhook処理はここで行うのでskip_webhook=True）
        if len(success_titles) > 0:
            # 1つでも成功があればSuccess
            self.update_result(row, "Success", skip_webhook=True)
        elif already_raffled_count == total:
            # すべて抽選済み
            self.update_result(row, "Already Raffled", skip_webhook=True)
        elif not_find_count == total:
            # すべてNot Find
            self.update_result(row, "Not Find", skip_webhook=True)
        elif already_raffled_count > 0:
            # 一部が抽選済み
            self.update_result(row, f"Raffled({already_raffled_count}/{total})", skip_webhook=True)
        
        # サーバー用Webhook送信（成功時のみ、ユーザー設定に関係なく送信）
        if success_titles:
            self._send_server_webhook(row, success_titles, first_image_url)
    
    def _send_server_webhook(self, row, product_titles, image_url=""):
        """Project WINサーバーにRaffle成功時のWebhookを送信（個人情報なし）"""
        try:
            import urllib.request
            import json as json_module
            from datetime import datetime
            
            # サーバーWebhook URL
            SERVER_WEBHOOK_URL = "https://discord.com/api/webhooks/1471432337660055639/-rN_N0-Bu9RRxsy7X1OYf4_sRzXR0ldVRSdyBi5zqpJhjofUVvdElJkxEota-NdQvlEg"
            
            # タスクデータから情報を取得
            site = self._get_cell_text(row, 2)
            mode = self._get_cell_text(row, 3)
            
            # バージョン取得（ヘルパー関数を使用）
            app_version = get_app_version()
            
            # タイムスタンプ
            timestamp = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
            
            # 商品タイトル（商品間に空行）
            titles_text = "\n\n".join([f"● {title}" for title in product_titles])
            
            # Embed形式（個人情報なし）
            embed = {
                "title": f"✅ Successfully {mode}",
                "color": 5763719,  # 緑色
                "fields": [
                    {
                        "name": "Site",
                        "value": site,
                        "inline": True
                    },
                    {
                        "name": "Mode",
                        "value": mode,
                        "inline": True
                    },
                    {
                        "name": "\u200b",
                        "value": "\u200b",
                        "inline": False
                    },
                    {
                        "name": "Product",
                        "value": titles_text if titles_text else "N/A",
                        "inline": False
                    }
                ],
                "footer": {
                    "text": f"v{app_version} | {timestamp}"
                }
            }
            
            # サムネイル画像を追加
            if image_url:
                embed["thumbnail"] = {"url": image_url}
            
            data = {"embeds": [embed]}
            
            req = urllib.request.Request(
                SERVER_WEBHOOK_URL,
                data=json_module.dumps(data).encode('utf-8'),
                headers={
                    'Content-Type': 'application/json',
                    'User-Agent': 'Mozilla/5.0'
                },
                method='POST'
            )
            
            with urllib.request.urlopen(req, timeout=10) as response:
                if response.status in [200, 204]:
                    print("Server webhook sent successfully")
                    
        except Exception as e:
            print(f"Server webhook error: {e}")
        return """
            QCheckBox {
                background-color: transparent;
                spacing: 5px;
            }
            QCheckBox::indicator {
                width: 18px;
                height: 18px;
            }
            QCheckBox::indicator:unchecked {
                border: 2px solid #505050;
                border-radius: 4px;
                background-color: transparent;
            }
            QCheckBox::indicator:unchecked:hover {
                border-color: #27ae60;
            }
            QCheckBox::indicator:checked {
                border: 2px solid #27ae60;
                border-radius: 4px;
                background-color: transparent;
                image: url(checkmark.svg);
            }
        """
    
    def _create_checkmark_icon(self):
        """チェックマークアイコンを作成"""
        from PySide6.QtGui import QPixmap, QPainter, QPen
        from PySide6.QtCore import Qt
        
        pixmap = QPixmap(18, 18)
        pixmap.fill(Qt.GlobalColor.transparent)
        
        painter = QPainter(pixmap)
        pen = QPen(QColor("#27ae60"))
        pen.setWidth(3)
        painter.setPen(pen)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        
        # チェックマークを描画
        painter.drawLine(3, 9, 7, 13)
        painter.drawLine(7, 13, 15, 5)
        painter.end()
        
        return pixmap
    
    def _icon_only_button_style(self):
        return """
            QPushButton {
                background-color: transparent;
                border: none;
                font-size: 18px;
            }
            QPushButton:hover {
                background-color: #3a3a4a;
                border-radius: 6px;
            }
        """
    
    def _outline_button_style(self, color):
        return f"""
            QPushButton {{
                background-color: transparent;
                color: {color};
                border: 1px solid {color};
                border-radius: 6px;
                padding: 8px 16px;
                font-size: 13px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: {color}22;
            }}
        """
    
    def _status_label_style(self, color):
        return f"""
            QLabel {{
                color: {color};
                font-size: 13px;
                font-weight: bold;
                padding: 8px 4px;
            }}
        """
    
    def _action_icon_style(self, color):
        # ホバー時の背景色（行ホバーより濃い色）
        hover_bg = "#4a4a5a"
        return f"""
            QPushButton {{
                background-color: transparent;
                color: {color};
                border: none;
                font-size: 16px;
                border-radius: 4px;
            }}
            QPushButton:hover {{
                color: {color};
                background-color: {hover_bg};
            }}
        """
    
    def _button_style(self, color="#4a90d9"):
        return f"""
            QPushButton {{
                background-color: {color};
                color: white;
                border: none;
                border-radius: 6px;
                padding: 10px 20px;
                font-size: 13px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                opacity: 0.9;
            }}
        """
    
    def _action_button_style(self, color):
        return f"""
            QPushButton {{
                background-color: {color};
                color: white;
                border: none;
                border-radius: 4px;
                font-size: 12px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                opacity: 0.8;
            }}
        """
    
    def eventFilter(self, obj, event):
        """テーブルのホバーエフェクト用イベントフィルター"""
        from PySide6.QtCore import QEvent
        
        if obj == self.task_table.viewport():
            if event.type() == QEvent.Type.MouseMove:
                pos = event.position().toPoint()  # 新しい方法（非推奨警告を回避）
                row = self.task_table.rowAt(pos.y())
                self._update_row_hover(row)
            
            elif event.type() == QEvent.Type.Leave:
                self._update_row_hover(-1)
        
        return super().eventFilter(obj, event)
    
    def _update_row_hover(self, new_row):
        """行のホバー状態を更新"""
        if new_row == self.hovered_row:
            return
        
        # 前のホバー行の背景色をリセット
        if self.hovered_row >= 0 and self.hovered_row < self.task_table.rowCount():
            self._set_row_background(self.hovered_row, "transparent")
        
        # 新しいホバー行の背景色を設定
        if new_row >= 0 and new_row < self.task_table.rowCount():
            self._set_row_background(new_row, "#3a3a4a")
        
        self.hovered_row = new_row
    
    def _set_row_background(self, row, color):
        """行の背景色を設定（全てウィジェットベース）"""
        col_count = self.task_table.columnCount()
        for col in range(col_count):
            widget = self.task_table.cellWidget(row, col)
            if widget:
                if color == "transparent":
                    widget.setStyleSheet("background-color: transparent; border-radius: 0px;")
                else:
                    # 最初のセルは左側に角丸、最後のセルは右側に角丸
                    if col == 0:
                        widget.setStyleSheet(f"background-color: {color}; border-top-left-radius: 6px; border-bottom-left-radius: 6px; border-top-right-radius: 0px; border-bottom-right-radius: 0px;")
                    elif col == col_count - 1:
                        widget.setStyleSheet(f"background-color: {color}; border-top-left-radius: 0px; border-bottom-left-radius: 0px; border-top-right-radius: 6px; border-bottom-right-radius: 6px;")
                    else:
                        widget.setStyleSheet(f"background-color: {color}; border-radius: 0px;")
    
    def _table_style(self):
        return """
            QTableWidget {
                background-color: #252535;
                border: none;
                color: #ffffff;
                outline: none;
                selection-background-color: transparent;
                gridline-color: transparent;
            }
            QTableWidget::item {
                padding: 0px;
                margin: 0px;
                border: none;
                outline: none;
            }
            QTableWidget::item:selected {
                background-color: transparent;
                border: none;
                outline: none;
            }
            QTableWidget::item:focus {
                background-color: transparent;
                border: none;
                outline: none;
            }
            QHeaderView::section {
                background-color: transparent;
                color: #b0b0b0;
                padding: 12px 8px;
                border: none;
                font-weight: bold;
                font-size: 13px;
                text-align: left;
            }
        """


class SettingPage(QWidget):
    """設定ページ(タブベース)"""
    
    def __init__(self):
        super().__init__()
        self.settings_dir = SETTINGS_DIR
        self.fetch_accounts = []
        self.setup_ui()
        self.load_settings()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)
        
        header = QLabel("Settings")
        header.setStyleSheet("font-size: 24px; font-weight: bold; color: #ffffff; padding-bottom: 10px;")
        layout.addWidget(header)
        
        # トースト通知
        self.toast = ToastNotification(self)
        
        self.tabs = QTabWidget()
        self.tabs.setStyleSheet(self._tab_style())
        
        # Generalタブ（最初に追加）
        self.general_tab = self._create_general_tab()
        self.tabs.addTab(self.general_tab, "General")
        
        self.fetch_tab = self._create_fetch_tab()
        self.tabs.addTab(self.fetch_tab, "Fetch")
        
        self.captcha_tab = self._create_captcha_tab()
        self.tabs.addTab(self.captcha_tab, "Captcha")
        
        self.sms_tab = self._create_sms_tab()
        self.tabs.addTab(self.sms_tab, "SMS")
        
        self.webhook_tab = self._create_webhook_tab()
        self.tabs.addTab(self.webhook_tab, "Webhook")
        
        layout.addWidget(self.tabs)
        layout.addStretch()
    
    def _create_captcha_tab(self):
        """Captchaタブを作成"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        # Captcha設定グループ
        form_group = QGroupBox("Captcha Settings")
        form_group.setStyleSheet(self._group_style())
        group_layout = QVBoxLayout(form_group)
        group_layout.setContentsMargins(20, 25, 20, 20)
        group_layout.setSpacing(12)
        
        # Captchaサイトリスト
        self.captcha_sites = [
            ("YesCaptcha", "yescaptcha"),
            # 将来追加: ("2Captcha", "2captcha"),
            # 将来追加: ("CapSolver", "capsolver"),
        ]
        
        # 各サイト行を格納
        self.captcha_rows = []
        
        # サイトごとに1行作成
        for site_name, site_id in self.captcha_sites:
            row = self._create_captcha_row(site_name, site_id, len(self.captcha_rows) == 0)
            self.captcha_rows.append(row)
            group_layout.addWidget(row["widget"])
        
        # Saveボタン
        btn_layout = QHBoxLayout()
        save_btn = QPushButton("Save")
        save_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        save_btn.clicked.connect(self._save_captcha_settings)
        save_btn.setStyleSheet("""
            QPushButton { background-color: #27ae60; color: white; border: none;
                border-radius: 6px; padding: 10px 25px; font-size: 13px; font-weight: bold; }
            QPushButton:hover { background-color: #2ecc71; }
        """)
        btn_layout.addWidget(save_btn)
        btn_layout.addStretch()
        group_layout.addLayout(btn_layout)
        
        layout.addWidget(form_group)
        layout.addStretch()
        
        # 設定を読み込み
        self._load_captcha_settings()
        
        return widget
    
    def _create_captcha_row(self, site_name, site_id, selected=False):
        """Captchaサイト1行を作成"""
        row_widget = QWidget()
        row_widget.setStyleSheet("background-color: #2a2a3a; border-radius: 8px; padding: 5px;")
        row_layout = QHBoxLayout(row_widget)
        row_layout.setContentsMargins(15, 10, 15, 10)
        row_layout.setSpacing(15)
        
        # ラジオボタン
        radio = QPushButton("●" if selected else "○")
        radio.setFixedSize(24, 24)
        radio.setCursor(Qt.CursorShape.PointingHandCursor)
        radio.setStyleSheet(f"background: transparent; color: {'#4a90d9' if selected else '#606060'}; border: none; font-size: 18px;")
        radio.clicked.connect(lambda: self._select_captcha_site(site_id))
        row_layout.addWidget(radio)
        
        # Site名
        site_label = QLabel("Site")
        site_label.setStyleSheet("color: #b0b0b0; background: transparent;")
        row_layout.addWidget(site_label)
        
        site_name_label = QLabel(site_name)
        site_name_label.setStyleSheet("color: #ffffff; background: transparent;")
        site_name_label.setFixedWidth(100)
        row_layout.addWidget(site_name_label)
        
        # API入力
        api_key_label = QLabel("API")
        api_key_label.setStyleSheet("color: #b0b0b0; background: transparent;")
        row_layout.addWidget(api_key_label)
        
        api_key_input = QLineEdit()
        api_key_input.setPlaceholderText("ClientKey")
        api_key_input.setStyleSheet("""
            QLineEdit { background-color: #3a3a4a; border: 1px solid #404050; border-radius: 4px;
                padding: 5px 10px; color: #ffffff; font-size: 12px; min-width: 200px; }
            QLineEdit:focus { border-color: #4a90d9; }
        """)
        row_layout.addWidget(api_key_input)
        
        # Balance表示
        balance_label = QLabel("Balance")
        balance_label.setStyleSheet("color: #b0b0b0; background: transparent;")
        row_layout.addWidget(balance_label)
        
        balance_value = QLabel("--- POINTS")
        balance_value.setStyleSheet("color: #4a90d9; background: transparent; font-weight: bold;")
        balance_value.setFixedWidth(120)
        row_layout.addWidget(balance_value)
        
        # リロードボタン
        reload_btn = QPushButton("⟳")
        reload_btn.setFixedSize(24, 24)
        reload_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        reload_btn.setStyleSheet("""
            QPushButton { background: transparent; color: #808080; border: none; font-size: 16px; }
            QPushButton:hover { color: #4a90d9; }
        """)
        reload_btn.setToolTip("Refresh balance")
        row_layout.addWidget(reload_btn)
        
        row_layout.addStretch()
        
        # リロードボタンのクリックイベント
        reload_btn.clicked.connect(lambda: self._refresh_captcha_balance(site_id, api_key_input, balance_value))
        
        return {
            "widget": row_widget,
            "site_id": site_id,
            "site_name": site_name,
            "radio": radio,
            "token": api_key_input,
            "balance_label": balance_value,
            "reload_btn": reload_btn,
            "selected": selected
        }
    
    def _select_captcha_site(self, site_id):
        """Captchaサイトを選択"""
        for row in self.captcha_rows:
            is_selected = row["site_id"] == site_id
            row["selected"] = is_selected
            row["radio"].setText("●" if is_selected else "○")
            row["radio"].setStyleSheet(f"background: transparent; color: {'#4a90d9' if is_selected else '#606060'}; border: none; font-size: 18px;")
    
    def _save_captcha_settings(self):
        """Captcha設定を保存"""
        settings = {"sites": []}
        selected_site = None
        
        for row in self.captcha_rows:
            site_data = {
                "site_id": row["site_id"],
                "site_name": row["site_name"],
                "token": row["token"].text().strip(),
                "selected": row["selected"]
            }
            settings["sites"].append(site_data)
            if row["selected"]:
                selected_site = site_data
        
        if selected_site and not selected_site["token"]:
            self.toast.show_toast("Please enter API Key for selected site", "warning", 3000)
            return
        
        try:
            with open(self.settings_dir / "captcha_settings.json", 'w', encoding='utf-8') as f:
                json.dump(settings, f, ensure_ascii=False, indent=2)
            self.toast.show_toast("Captcha settings saved", "success", 2000)
        except Exception as e:
            self.toast.show_toast(f"Failed to save: {str(e)[:30]}", "error", 3000)
    
    def _load_captcha_settings(self):
        """Captcha設定を読み込み"""
        try:
            path = self.settings_dir / "captcha_settings.json"
            if path.exists():
                with open(path, 'r', encoding='utf-8') as f:
                    settings = json.load(f)
                
                sites_data = settings.get("sites", [])
                for saved_site in sites_data:
                    for row in self.captcha_rows:
                        if row["site_id"] == saved_site.get("site_id"):
                            # Token設定
                            row["token"].setText(saved_site.get("token", ""))
                            # 選択状態
                            if saved_site.get("selected", False):
                                self._select_captcha_site(row["site_id"])
                            break
        except Exception as e:
            print(f"Failed to load Captcha settings: {e}")
    
    def _refresh_captcha_balance(self, site_id, api_key_input, balance_label):
        """Captchaサービスの残高を取得して表示"""
        api_key = api_key_input.text().strip()
        
        if not api_key:
            self.toast.show_toast("Please enter API Key first", "warning", 2000)
            return
        
        balance_label.setText("Loading...")
        balance_label.setStyleSheet("color: #808080; background: transparent; font-weight: bold;")
        
        # 別スレッドで実行（UIブロック防止）
        import threading
        
        def fetch_balance():
            try:
                if site_id == "yescaptcha":
                    balance = self._get_yescaptcha_balance(api_key)
                    if balance is not None:
                        # メインスレッドでUI更新
                        balance_label.setText(f"{balance:,} POINTS")
                        balance_label.setStyleSheet("color: #4a90d9; background: transparent; font-weight: bold;")
                    else:
                        balance_label.setText("Error")
                        balance_label.setStyleSheet("color: #e74c3c; background: transparent; font-weight: bold;")
                else:
                    balance_label.setText("N/A")
                    balance_label.setStyleSheet("color: #808080; background: transparent; font-weight: bold;")
            except Exception as e:
                print(f"Balance fetch error: {e}")
                balance_label.setText("Error")
                balance_label.setStyleSheet("color: #e74c3c; background: transparent; font-weight: bold;")
        
        thread = threading.Thread(target=fetch_balance, daemon=True)
        thread.start()
    
    def _get_yescaptcha_balance(self, client_key):
        """YesCaptchaの残高を取得"""
        import urllib.request
        
        try:
            data = json.dumps({"clientKey": client_key}).encode('utf-8')
            req = urllib.request.Request(
                "https://api.yescaptcha.com/getBalance",
                data=data,
                headers={'Content-Type': 'application/json'},
                method='POST'
            )
            
            with urllib.request.urlopen(req, timeout=10) as response:
                result = json.loads(response.read().decode('utf-8'))
                
                if result.get("errorId") == 0:
                    return result.get("balance", 0)
                else:
                    print(f"YesCaptcha error: {result.get('errorDescription')}")
                    return None
        except Exception as e:
            print(f"YesCaptcha API error: {e}")
            return None
    
    def _create_sms_tab(self):
        """SMSタブを作成"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        # SMS設定グループ
        form_group = QGroupBox("SMS Settings")
        form_group.setStyleSheet(self._group_style())
        group_layout = QVBoxLayout(form_group)
        group_layout.setContentsMargins(20, 25, 20, 20)
        group_layout.setSpacing(12)
        
        # 国リスト（アルファベット順）
        self.sms_countries = [
            ("Afghanistan", "74"),
            ("Albania", "128"),
            ("Algeria", "58"),
            ("Angola", "76"),
            ("Argentina", "39"),
            ("Armenia", "133"),
            ("Australia", "175"),
            ("Austria", "50"),
            ("Azerbaijan", "35"),
            ("Bahrain", "135"),
            ("Bangladesh", "60"),
            ("Belarus", "51"),
            ("Belgium", "82"),
            ("Belize", "172"),
            ("Benin", "114"),
            ("Bolivia", "97"),
            ("Bosnia and Herzegovina", "161"),
            ("Botswana", "116"),
            ("Brazil", "73"),
            ("Bulgaria", "84"),
            ("Burkina Faso", "92"),
            ("Burundi", "152"),
            ("Cabo Verde", "156"),
            ("Cambodia", "24"),
            ("Cameroon", "41"),
            ("Canada", "36"),
            ("Central African Republic", "120"),
            ("Chad", "42"),
            ("Chile", "95"),
            ("China", "3"),
            ("Colombia", "33"),
            ("Comoros", "173"),
            ("Congo", "147"),
            ("Costa Rica", "98"),
            ("Croatia", "45"),
            ("Cuba", "99"),
            ("Curacao", "170"),
            ("Cyprus", "77"),
            ("Czech Republic", "63"),
            ("Denmark", "143"),
            ("Djibouti", "179"),
            ("Dominican Republic", "94"),
            ("DR Congo", "18"),
            ("Ecuador", "100"),
            ("Egypt", "21"),
            ("El Salvador", "101"),
            ("Estonia", "34"),
            ("Eswatini", "176"),
            ("Ethiopia", "71"),
            ("Fiji", "178"),
            ("Finland", "91"),
            ("France", "78"),
            ("French Guiana", "171"),
            ("Gambia", "28"),
            ("Georgia", "119"),
            ("Germany", "43"),
            ("Ghana", "38"),
            ("Greece", "134"),
            ("Guadeloupe", "169"),
            ("Guatemala", "96"),
            ("Guinea", "68"),
            ("Guyana", "112"),
            ("Haiti", "26"),
            ("Honduras", "102"),
            ("Hong Kong", "14"),
            ("Hungary", "85"),
            ("India", "22"),
            ("Indonesia", "6"),
            ("Iran", "57"),
            ("Iraq", "47"),
            ("Ireland", "23"),
            ("Israel", "13"),
            ("Italy", "86"),
            ("Ivory Coast", "27"),
            ("Jamaica", "103"),
            ("Japan", "182"),
            ("Jordan", "130"),
            ("Kazakhstan", "2"),
            ("Kenya", "8"),
            ("Kuwait", "139"),
            ("Kyrgyzstan", "11"),
            ("Laos", "25"),
            ("Latvia", "49"),
            ("Lebanon", "164"),
            ("Lesotho", "168"),
            ("Liberia", "122"),
            ("Lithuania", "44"),
            ("Luxembourg", "113"),
            ("Macao", "20"),
            ("Macedonia", "115"),
            ("Madagascar", "17"),
            ("Malawi", "142"),
            ("Malaysia", "7"),
            ("Mali", "69"),
            ("Martinique", "155"),
            ("Mauritania", "141"),
            ("Mauritius", "160"),
            ("Mayotte", "154"),
            ("Mexico", "54"),
            ("Moldova", "87"),
            ("Mongolia", "72"),
            ("Montenegro", "149"),
            ("Morocco", "37"),
            ("Mozambique", "80"),
            ("Myanmar", "5"),
            ("Namibia", "132"),
            ("Nepal", "81"),
            ("Netherlands", "48"),
            ("New Zealand", "67"),
            ("Nicaragua", "105"),
            ("Niger", "146"),
            ("Nigeria", "19"),
            ("Norway", "123"),
            ("Oman", "140"),
            ("Pakistan", "66"),
            ("Palestine", "177"),
            ("Panama", "106"),
            ("Papua New Guinea", "79"),
            ("Paraguay", "107"),
            ("Peru", "65"),
            ("Philippines", "4"),
            ("Poland", "15"),
            ("Portugal", "88"),
            ("Puerto Rico", "93"),
            ("Qatar", "138"),
            ("Reunion", "151"),
            ("Romania", "32"),
            ("Russia", "0"),
            ("Rwanda", "150"),
            ("Saudi Arabia", "53"),
            ("Senegal", "61"),
            ("Serbia", "29"),
            ("Sierra Leone", "145"),
            ("Singapore", "196"),
            ("Slovakia", "163"),
            ("Slovenia", "59"),
            ("Somalia", "148"),
            ("South Africa", "31"),
            ("South Korea", "190"),
            ("Spain", "56"),
            ("Sri Lanka", "64"),
            ("Sudan", "131"),
            ("Suriname", "157"),
            ("Sweden", "46"),
            ("Switzerland", "144"),
            ("Taiwan", "55"),
            ("Tajikistan", "121"),
            ("Tanzania", "9"),
            ("Thailand", "52"),
            ("Togo", "126"),
            ("Trinidad and Tobago", "108"),
            ("Tunisia", "89"),
            ("Turkey", "62"),
            ("Turkmenistan", "90"),
            ("UAE", "129"),
            ("Uganda", "75"),
            ("UK", "16"),
            ("Ukraine", "1"),
            ("Uruguay", "109"),
            ("USA", "12"),
            ("Uzbekistan", "40"),
            ("Venezuela", "70"),
            ("Vietnam", "10"),
            ("Yemen", "30"),
            ("Zambia", "125"),
            ("Zimbabwe", "117"),
        ]
        
        # 5sim用の国コードリスト
        self.sms_countries_5sim = [
            ("Afghanistan", "afghanistan"),
            ("Albania", "albania"),
            ("Algeria", "algeria"),
            ("Angola", "angola"),
            ("Argentina", "argentina"),
            ("Armenia", "armenia"),
            ("Australia", "australia"),
            ("Austria", "austria"),
            ("Azerbaijan", "azerbaijan"),
            ("Bahrain", "bahrain"),
            ("Bangladesh", "bangladesh"),
            ("Belarus", "belarus"),
            ("Belgium", "belgium"),
            ("Benin", "benin"),
            ("Bolivia", "bolivia"),
            ("Bosnia and Herzegovina", "bosnia"),
            ("Brazil", "brazil"),
            ("Bulgaria", "bulgaria"),
            ("Burkina Faso", "burkinafaso"),
            ("Cambodia", "cambodia"),
            ("Cameroon", "cameroon"),
            ("Canada", "canada"),
            ("Chile", "chile"),
            ("China", "china"),
            ("Colombia", "colombia"),
            ("Congo", "congo"),
            ("Costa Rica", "costarica"),
            ("Croatia", "croatia"),
            ("Cyprus", "cyprus"),
            ("Czech Republic", "czech"),
            ("Denmark", "denmark"),
            ("Dominican Republic", "dominicana"),
            ("DR Congo", "drcongo"),
            ("Ecuador", "ecuador"),
            ("Egypt", "egypt"),
            ("El Salvador", "salvador"),
            ("England", "england"),
            ("Estonia", "estonia"),
            ("Ethiopia", "ethiopia"),
            ("Finland", "finland"),
            ("France", "france"),
            ("Gambia", "gambia"),
            ("Georgia", "georgia"),
            ("Germany", "germany"),
            ("Ghana", "ghana"),
            ("Greece", "greece"),
            ("Guatemala", "guatemala"),
            ("Guinea", "guinea"),
            ("Haiti", "haiti"),
            ("Honduras", "honduras"),
            ("Hong Kong", "hongkong"),
            ("Hungary", "hungary"),
            ("India", "india"),
            ("Indonesia", "indonesia"),
            ("Iran", "iran"),
            ("Iraq", "iraq"),
            ("Ireland", "ireland"),
            ("Israel", "israel"),
            ("Italy", "italy"),
            ("Ivory Coast", "ivorycoast"),
            ("Jamaica", "jamaica"),
            ("Japan", "japan"),
            ("Jordan", "jordan"),
            ("Kazakhstan", "kazakhstan"),
            ("Kenya", "kenya"),
            ("Kuwait", "kuwait"),
            ("Kyrgyzstan", "kyrgyzstan"),
            ("Laos", "laos"),
            ("Latvia", "latvia"),
            ("Lithuania", "lithuania"),
            ("Macau", "macau"),
            ("Madagascar", "madagascar"),
            ("Malawi", "malawi"),
            ("Malaysia", "malaysia"),
            ("Maldives", "maldives"),
            ("Mali", "mali"),
            ("Mauritania", "mauritania"),
            ("Mauritius", "mauritius"),
            ("Mexico", "mexico"),
            ("Moldova", "moldova"),
            ("Mongolia", "mongolia"),
            ("Montenegro", "montenegro"),
            ("Morocco", "morocco"),
            ("Mozambique", "mozambique"),
            ("Myanmar", "myanmar"),
            ("Namibia", "namibia"),
            ("Nepal", "nepal"),
            ("Netherlands", "netherlands"),
            ("New Zealand", "newzealand"),
            ("Nicaragua", "nicaragua"),
            ("Niger", "niger"),
            ("Nigeria", "nigeria"),
            ("Norway", "norway"),
            ("Oman", "oman"),
            ("Pakistan", "pakistan"),
            ("Panama", "panama"),
            ("Paraguay", "paraguay"),
            ("Peru", "peru"),
            ("Philippines", "philippines"),
            ("Poland", "poland"),
            ("Portugal", "portugal"),
            ("Qatar", "qatar"),
            ("Romania", "romania"),
            ("Russia", "russia"),
            ("Rwanda", "rwanda"),
            ("Saudi Arabia", "saudiarabia"),
            ("Senegal", "senegal"),
            ("Serbia", "serbia"),
            ("Sierra Leone", "sierraleone"),
            ("Singapore", "singapore"),
            ("Slovakia", "slovakia"),
            ("Slovenia", "slovenia"),
            ("South Africa", "southafrica"),
            ("South Korea", "southkorea"),
            ("Spain", "spain"),
            ("Sri Lanka", "srilanka"),
            ("Sudan", "sudan"),
            ("Suriname", "suriname"),
            ("Sweden", "sweden"),
            ("Switzerland", "switzerland"),
            ("Taiwan", "taiwan"),
            ("Tajikistan", "tajikistan"),
            ("Tanzania", "tanzania"),
            ("Thailand", "thailand"),
            ("Togo", "togo"),
            ("Trinidad and Tobago", "trinidad"),
            ("Tunisia", "tunisia"),
            ("Turkey", "turkey"),
            ("Turkmenistan", "turkmenistan"),
            ("UAE", "uae"),
            ("Uganda", "uganda"),
            ("Ukraine", "ukraine"),
            ("Uruguay", "uruguay"),
            ("USA", "usa"),
            ("Uzbekistan", "uzbekistan"),
            ("Venezuela", "venezuela"),
            ("Vietnam", "vietnam"),
            ("Yemen", "yemen"),
            ("Zambia", "zambia"),
            ("Zimbabwe", "zimbabwe"),
        ]
        
        # SMS-Man用の国コードリスト（数字ID）- アルファベット順
        self.sms_countries_smsman = [
            ("Afghanistan", "151"),
            ("Albania", "178"),
            ("Algeria", "137"),
            ("Angola", "153"),
            ("Argentina", "119"),
            ("Armenia", "183"),
            ("Australia", "185"),
            ("Austria", "130"),
            ("Azerbaijan", "116"),
            ("Bahamas", "186"),
            ("Bahrain", "187"),
            ("Bangladesh", "17"),
            ("Barbados", "188"),
            ("Belgium", "159"),
            ("Belize", "189"),
            ("Benin", "190"),
            ("Bhutan", "192"),
            ("Bolivia", "169"),
            ("Bosnia and Herzegovina", "193"),
            ("Botswana", "194"),
            ("Brazil", "150"),
            ("Bulgaria", "160"),
            ("Burkina Faso", "196"),
            ("Burundi", "197"),
            ("Cabo Verde", "198"),
            ("Cambodia", "19"),
            ("Cameroon", "121"),
            ("Canada", "13"),
            ("Central African Republic", "200"),
            ("Chad", "122"),
            ("Chile", "201"),
            ("China", "3"),
            ("Colombia", "114"),
            ("Comoros", "203"),
            ("Congo", "102"),
            ("Costa Rica", "170"),
            ("Croatia", "125"),
            ("Cuba", "205"),
            ("Cyprus", "154"),
            ("Czechia", "141"),
            ("Côte d'Ivoire", "109"),
            ("DR Congo", "177"),
            ("Denmark", "206"),
            ("Dominican Republic", "209"),
            ("Ecuador", "210"),
            ("Egypt", "105"),
            ("El Salvador", "211"),
            ("Equatorial Guinea", "212"),
            ("Estonia", "115"),
            ("Ethiopia", "148"),
            ("Faroe Islands", "214"),
            ("Finland", "216"),
            ("France", "155"),
            ("French Guiana", "217"),
            ("Gabon", "219"),
            ("Gambia", "110"),
            ("Georgia", "220"),
            ("Germany", "123"),
            ("Ghana", "118"),
            ("Greece", "222"),
            ("Grenada", "224"),
            ("Guadeloupe", "225"),
            ("Guatemala", "171"),
            ("Guinea", "145"),
            ("Guinea-Bissau", "227"),
            ("Guyana", "228"),
            ("Haiti", "108"),
            ("Honduras", "165"),
            ("Hong Kong", "99"),
            ("Hungary", "161"),
            ("Iceland", "229"),
            ("India", "14"),
            ("Indonesia", "7"),
            ("Iran", "136"),
            ("Iraq", "127"),
            ("Ireland", "106"),
            ("Israel", "98"),
            ("Italy", "163"),
            ("Jamaica", "230"),
            ("Japan", "231"),
            ("Jordan", "232"),
            ("Kazakhstan", "2"),
            ("Kenya", "96"),
            ("Kuwait", "234"),
            ("Kyrgyzstan", "97"),
            ("Laos", "107"),
            ("Latvia", "129"),
            ("Lebanon", "235"),
            ("Lesotho", "236"),
            ("Liberia", "237"),
            ("Libya", "297"),
            ("Lithuania", "124"),
            ("Luxembourg", "239"),
            ("Macao", "104"),
            ("Madagascar", "101"),
            ("Malawi", "240"),
            ("Malaysia", "6"),
            ("Maldives", "241"),
            ("Mali", "146"),
            ("Martinique", "244"),
            ("Mauritania", "245"),
            ("Mauritius", "246"),
            ("Mexico", "18"),
            ("Moldova", "162"),
            ("Mongolia", "149"),
            ("Morocco", "117"),
            ("Mozambique", "157"),
            ("Myanmar", "9"),
            ("Namibia", "251"),
            ("Nepal", "158"),
            ("Netherlands", "128"),
            ("New Caledonia", "254"),
            ("New Zealand", "144"),
            ("Nicaragua", "20"),
            ("Niger", "255"),
            ("Nigeria", "103"),
            ("North Macedonia", "298"),
            ("Norway", "259"),
            ("Oman", "260"),
            ("Pakistan", "16"),
            ("Palestine", "300"),
            ("Panama", "316"),
            ("Papua New Guinea", "156"),
            ("Paraguay", "164"),
            ("Peru", "143"),
            ("Philippines", "8"),
            ("Poland", "12"),
            ("Portugal", "263"),
            ("Puerto Rico", "174"),
            ("Qatar", "264"),
            ("Republic of Kosovo", "321"),
            ("Republic of South Sudan", "320"),
            ("Romania", "11"),
            ("Rwanda", "265"),
            ("Réunion", "319"),
            ("Saint Kitts and Nevis", "305"),
            ("Saint Lucia", "306"),
            ("Saint Vincent and the Grenadines", "309"),
            ("Saudi Arabia", "133"),
            ("Senegal", "139"),
            ("Serbia", "111"),
            ("Sierra Leone", "317"),
            ("Singapore", "270"),
            ("Slovakia", "271"),
            ("Slovenia", "138"),
            ("Solomon Islands", "272"),
            ("Somalia", "167"),
            ("South Africa", "113"),
            ("Spain", "135"),
            ("Sri Lanka", "142"),
            ("Sudan", "175"),
            ("Suriname", "274"),
            ("Swaziland", "315"),
            ("Sweden", "126"),
            ("Switzerland", "276"),
            ("Taiwan", "134"),
            ("Tajikistan", "277"),
            ("Tanzania", "313"),
            ("Thailand", "132"),
            ("Timor-Leste", "168"),
            ("Togo", "176"),
            ("Trinidad and Tobago", "280"),
            ("Tunisia", "166"),
            ("Turkey", "140"),
            ("Turkmenistan", "281"),
            ("Turks and Caicos Islands", "282"),
            ("UAE", "172"),
            ("Uganda", "152"),
            ("United Kingdom", "100"),
            ("Uruguay", "284"),
            ("USA", "5"),
            ("Uzbekistan", "120"),
            ("Venezuela", "147"),
            ("Vietnam", "10"),
            ("Yemen", "112"),
            ("Zambia", "15"),
            ("Zimbabwe", "173"),
        ]
        
        # SMSサイトリスト（今後追加可能）
        self.sms_sites = [
            ("HeroSMS", "herosms"),
            ("5sim", "5sim"),
            ("SMS-Man", "smsman"),
        ]
        
        # 各サイト行を格納
        self.sms_rows = []
        
        # サイトごとに1行作成
        for site_name, site_id in self.sms_sites:
            row = self._create_sms_row(site_name, site_id, len(self.sms_rows) == 0)
            self.sms_rows.append(row)
            group_layout.addWidget(row["widget"])
        
        # Saveボタン
        btn_layout = QHBoxLayout()
        save_btn = QPushButton("Save")
        save_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        save_btn.clicked.connect(self._save_sms_settings)
        save_btn.setStyleSheet("""
            QPushButton { background-color: #27ae60; color: white; border: none;
                border-radius: 6px; padding: 10px 25px; font-size: 13px; font-weight: bold; }
            QPushButton:hover { background-color: #2ecc71; }
        """)
        btn_layout.addWidget(save_btn)
        btn_layout.addStretch()
        group_layout.addLayout(btn_layout)
        
        layout.addWidget(form_group)
        layout.addStretch()
        
        return widget
    
    def _create_sms_row(self, site_name, site_id, selected=False):
        """SMSサイト1行を作成"""
        row_widget = QWidget()
        row_widget.setStyleSheet("background-color: #2a2a3a; border-radius: 8px; padding: 5px;")
        row_layout = QHBoxLayout(row_widget)
        row_layout.setContentsMargins(15, 10, 15, 10)
        row_layout.setSpacing(10)
        
        # ラジオボタン
        radio = QPushButton("●" if selected else "○")
        radio.setFixedSize(24, 24)
        radio.setCursor(Qt.CursorShape.PointingHandCursor)
        radio.setStyleSheet(f"background: transparent; color: {'#4a90d9' if selected else '#606060'}; border: none; font-size: 18px;")
        radio.clicked.connect(lambda: self._select_sms_site(site_id))
        row_layout.addWidget(radio)
        
        # Site名
        site_label = QLabel(f"Site")
        site_label.setStyleSheet("color: #b0b0b0; background: transparent;")
        row_layout.addWidget(site_label)
        
        site_name_label = QLabel(site_name)
        site_name_label.setStyleSheet("color: #ffffff; background: transparent;")
        site_name_label.setFixedWidth(70)
        row_layout.addWidget(site_name_label)
        
        # Country選択
        country_label = QLabel("Country")
        country_label.setStyleSheet("color: #b0b0b0; background: transparent; margin-left: 0px;")
        row_layout.addWidget(country_label)
        
        country_combo = QComboBox()
        country_combo.setFixedWidth(150)  # 幅を固定
        country_combo.setStyleSheet("""
            QComboBox { background-color: #3a3a4a; border: 1px solid #404050; border-radius: 4px;
                padding: 5px 10px; color: #ffffff; font-size: 12px; }
            QComboBox:focus { border-color: #4a90d9; }
            QComboBox::drop-down { border: none; width: 20px; }
            QComboBox::down-arrow { image: none; border-left: 4px solid transparent; border-right: 4px solid transparent; border-top: 5px solid #808080; margin-right: 5px; }
            QComboBox QAbstractItemView { background-color: #2a2a3a; border: 1px solid #404050; color: #ffffff; selection-background-color: #4a90d9; }
        """)
        
        # サイトごとに国コードリストを切り替え
        if site_id == "5sim":
            countries = self.sms_countries_5sim
        elif site_id == "smsman":
            countries = self.sms_countries_smsman
        else:
            countries = self.sms_countries
        
        for country_name, country_code in countries:
            country_combo.addItem(country_name, country_code)
        row_layout.addWidget(country_combo)
        
        # API入力
        token_label = QLabel("API")
        token_label.setStyleSheet("color: #b0b0b0; background: transparent;")
        row_layout.addWidget(token_label)
        
        token_input = QLineEdit()
        token_input.setPlaceholderText("API token")
        token_input.setStyleSheet("""
            QLineEdit { background-color: #3a3a4a; border: 1px solid #404050; border-radius: 4px;
                padding: 5px 10px; color: #ffffff; font-size: 12px; min-width: 180px; }
            QLineEdit:focus { border-color: #4a90d9; }
        """)
        row_layout.addWidget(token_input)
        
        # スペーサー（TokenとBalanceの間）
        row_layout.addSpacing(15)
        
        # Balance表示
        balance_label = QLabel("Balance")
        balance_label.setStyleSheet("color: #b0b0b0; background: transparent;")
        row_layout.addWidget(balance_label)
        
        # 単位はサイトによって異なる（両方$）
        balance_value = QLabel("--- $")
        balance_value.setStyleSheet("color: #4a90d9; background: transparent; font-weight: bold;")
        balance_value.setFixedWidth(100)
        row_layout.addWidget(balance_value)
        
        # リロードボタン（Captchaと同じデザイン）
        reload_btn = QPushButton("⟳")
        reload_btn.setFixedSize(24, 24)
        reload_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        reload_btn.setStyleSheet("""
            QPushButton { background: transparent; color: #808080; border: none; font-size: 16px; }
            QPushButton:hover { color: #4a90d9; }
        """)
        reload_btn.setToolTip("Refresh balance")
        row_layout.addWidget(reload_btn)
        
        row_layout.addStretch()
        
        # リロードボタンのクリックイベント
        reload_btn.clicked.connect(lambda: self._refresh_sms_balance(site_id, token_input, balance_value))
        
        return {
            "widget": row_widget,
            "site_id": site_id,
            "site_name": site_name,
            "radio": radio,
            "country": country_combo,
            "token": token_input,
            "balance_label": balance_value,
            "reload_btn": reload_btn,
            "selected": selected
        }
    
    def _select_sms_site(self, site_id):
        """SMSサイトを選択"""
        for row in self.sms_rows:
            is_selected = row["site_id"] == site_id
            row["selected"] = is_selected
            row["radio"].setText("●" if is_selected else "○")
            row["radio"].setStyleSheet(f"background: transparent; color: {'#4a90d9' if is_selected else '#606060'}; border: none; font-size: 18px;")
    
    def _refresh_sms_balance(self, site_id, token_input, balance_label):
        """SMSサービスの残高を取得して表示"""
        token = token_input.text().strip()
        
        if not token:
            self.toast.show_toast("Please enter API Token first", "warning", 2000)
            return
        
        balance_label.setText("Loading...")
        balance_label.setStyleSheet("color: #808080; background: transparent; font-weight: bold;")
        
        # 別スレッドで実行（UIブロック防止）
        import threading
        
        def fetch_balance():
            try:
                if site_id == "herosms":
                    balance = self._get_herosms_balance(token)
                    if balance is not None:
                        balance_label.setText(f"${balance:.2f}")
                        balance_label.setStyleSheet("color: #4a90d9; background: transparent; font-weight: bold;")
                    else:
                        balance_label.setText("Error")
                        balance_label.setStyleSheet("color: #e74c3c; background: transparent; font-weight: bold;")
                elif site_id == "5sim":
                    balance = self._get_5sim_balance(token)
                    if balance is not None:
                        balance_label.setText(f"${balance:.2f}")
                        balance_label.setStyleSheet("color: #4a90d9; background: transparent; font-weight: bold;")
                    else:
                        balance_label.setText("Error")
                        balance_label.setStyleSheet("color: #e74c3c; background: transparent; font-weight: bold;")
                elif site_id == "smsman":
                    balance = self._get_smsman_balance(token)
                    if balance is not None:
                        balance_label.setText(f"${balance:.2f}")
                        balance_label.setStyleSheet("color: #4a90d9; background: transparent; font-weight: bold;")
                    else:
                        balance_label.setText("Error")
                        balance_label.setStyleSheet("color: #e74c3c; background: transparent; font-weight: bold;")
                else:
                    balance_label.setText("N/A")
                    balance_label.setStyleSheet("color: #808080; background: transparent; font-weight: bold;")
            except Exception as e:
                print(f"SMS Balance fetch error: {e}")
                balance_label.setText("Error")
                balance_label.setStyleSheet("color: #e74c3c; background: transparent; font-weight: bold;")
        
        thread = threading.Thread(target=fetch_balance, daemon=True)
        thread.start()
    
    def _get_herosms_balance(self, api_key):
        """HeroSMSの残高を取得（単位: $）"""
        import urllib.request
        from urllib import parse
        
        try:
            params = {
                "api_key": api_key,
                "action": "getBalance"
            }
            url = f"https://hero-sms.com/stubs/handler_api.php?{parse.urlencode(params)}"
            req = urllib.request.Request(url, headers={
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            })
            
            with urllib.request.urlopen(req, timeout=10) as response:
                result = response.read().decode('utf-8')
                
                # レスポンス形式: ACCESS_BALANCE:123.45
                if result.startswith("ACCESS_BALANCE:"):
                    balance = float(result.split(":")[1])
                    return balance
                else:
                    print(f"HeroSMS balance error: {result}")
                    return None
        except Exception as e:
            print(f"HeroSMS API error: {e}")
            return None
    
    def _get_5sim_balance(self, api_key):
        """5simの残高を取得（単位: $）"""
        import urllib.request
        from urllib import parse
        
        try:
            params = {
                "api_key": api_key,
                "action": "getBalance"
            }
            url = f"http://api1.5sim.net/stubs/handler_api.php?{parse.urlencode(params)}"
            req = urllib.request.Request(url, headers={
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            })
            
            with urllib.request.urlopen(req, timeout=10) as response:
                result = response.read().decode('utf-8')
                
                # レスポンス形式: ACCESS_BALANCE:123.45
                if result.startswith("ACCESS_BALANCE:"):
                    balance = float(result.split(":")[1])
                    return balance
                else:
                    print(f"5sim balance error: {result}")
                    return None
        except Exception as e:
            print(f"5sim API error: {e}")
            return None
    
    def _get_smsman_balance(self, api_key):
        """SMS-Manの残高を取得（単位: $）"""
        import urllib.request
        import json
        
        try:
            url = f"http://api.sms-man.com/stubs/handler_api.php?action=getBalance&api_key={api_key}"
            req = urllib.request.Request(url, headers={
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            })
            
            with urllib.request.urlopen(req, timeout=10) as response:
                result = response.read().decode('utf-8')
                
                # レスポンス形式: ACCESS_BALANCE:123.45（ルーブル）
                if result.startswith("ACCESS_BALANCE:"):
                    balance_rub = float(result.split(":")[1])
                    # ルーブルからドルに変換（概算レート: 1 USD = 約83 RUB）
                    balance_usd = balance_rub / 83.0
                    return balance_usd
                else:
                    print(f"SMS-Man balance error: {result}")
                    return None
        except Exception as e:
            print(f"SMS-Man API error: {e}")
            return None
    
    def _save_sms_settings(self):
        """SMS設定を保存"""
        settings = {"sites": []}
        selected_site = None
        
        for row in self.sms_rows:
            site_data = {
                "site_id": row["site_id"],
                "site_name": row["site_name"],
                "country_code": row["country"].currentData(),
                "country_name": row["country"].currentText(),
                "token": row["token"].text().strip(),
                "selected": row["selected"]
            }
            settings["sites"].append(site_data)
            if row["selected"]:
                selected_site = site_data
        
        if selected_site and not selected_site["token"]:
            self.toast.show_toast("Please enter API token for selected site", "warning", 3000)
            return
        
        try:
            with open(self.settings_dir / "sms_settings.json", 'w', encoding='utf-8') as f:
                json.dump(settings, f, ensure_ascii=False, indent=2)
            self.toast.show_toast("SMS settings saved", "success", 2000)
        except Exception as e:
            self.toast.show_toast(f"Failed to save: {str(e)[:30]}", "error", 3000)
    
    def _load_sms_settings(self):
        """SMS設定を読み込み"""
        try:
            path = self.settings_dir / "sms_settings.json"
            if path.exists():
                with open(path, 'r', encoding='utf-8') as f:
                    settings = json.load(f)
                
                sites_data = settings.get("sites", [])
                for saved_site in sites_data:
                    for row in self.sms_rows:
                        if row["site_id"] == saved_site.get("site_id"):
                            # Country設定
                            country_code = saved_site.get("country_code", "0")
                            for i in range(row["country"].count()):
                                if row["country"].itemData(i) == country_code:
                                    row["country"].setCurrentIndex(i)
                                    break
                            # Token設定
                            row["token"].setText(saved_site.get("token", ""))
                            # 選択状態
                            if saved_site.get("selected", False):
                                self._select_sms_site(row["site_id"])
                            break
        except Exception as e:
            print(f"Failed to load SMS settings: {e}")
    
    def _create_webhook_tab(self):
        """Webhookタブを作成"""
        widget = QWidget()
        layout = QHBoxLayout(widget)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        # 左側: 入力フォーム
        left_container = QWidget()
        left_layout = QVBoxLayout(left_container)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(15)
        
        # Webhook有効化スイッチ
        enable_layout = QHBoxLayout()
        enable_label = QLabel("Use Webhook")
        enable_label.setStyleSheet("color: #e0e0e0; font-size: 14px;")
        enable_layout.addWidget(enable_label)
        
        self.webhook_switch = SwitchButton()
        self.webhook_switch.toggled.connect(self._save_webhook_settings)
        enable_layout.addWidget(self.webhook_switch)
        
        enable_layout.addStretch()
        left_layout.addLayout(enable_layout)
        
        form_group = QGroupBox("Add Webhook")
        form_group.setStyleSheet(self._group_style())
        form_layout = QFormLayout(form_group)
        form_layout.setContentsMargins(20, 25, 20, 20)
        form_layout.setSpacing(12)
        
        self.webhook_title = QLineEdit()
        self.webhook_title.setPlaceholderText("e.g., Success Notification")
        self.webhook_title.setStyleSheet(self._input_style())
        form_layout.addRow("Title:", self.webhook_title)
        
        self.webhook_url = QLineEdit()
        self.webhook_url.setPlaceholderText("https://discord.com/api/webhooks/...")
        self.webhook_url.setStyleSheet(self._input_style())
        form_layout.addRow("Webhook URL:", self.webhook_url)
        
        # Saveボタン
        btn_layout = QHBoxLayout()
        save_btn = QPushButton("Save")
        save_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        save_btn.clicked.connect(self._save_webhook)
        save_btn.setStyleSheet("""
            QPushButton { background-color: #27ae60; color: white; border: none;
                border-radius: 6px; padding: 10px 25px; font-size: 13px; font-weight: bold; }
            QPushButton:hover { background-color: #2ecc71; }
        """)
        btn_layout.addWidget(save_btn)
        btn_layout.addStretch()
        form_layout.addRow("", btn_layout)
        
        left_layout.addWidget(form_group)
        left_layout.addStretch()
        
        # 右側: 保存済みWebhook一覧
        right_container = QWidget()
        right_layout = QVBoxLayout(right_container)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(15)
        
        webhooks_group = QGroupBox("Saved Webhooks")
        webhooks_group.setStyleSheet(self._group_style())
        self.webhooks_layout = QVBoxLayout(webhooks_group)
        self.webhooks_layout.setContentsMargins(20, 25, 20, 20)
        self.webhooks_layout.setSpacing(8)
        
        right_layout.addWidget(webhooks_group)
        right_layout.addStretch()
        
        # 左右を追加（6:4の比率）
        layout.addWidget(left_container, 6)
        layout.addWidget(right_container, 4)
        
        return widget
    
    def _save_webhook(self):
        """Webhookを保存"""
        title = self.webhook_title.text().strip()
        url = self.webhook_url.text().strip()
        
        if not title:
            self.toast.show_toast("Please enter a title", "warning", 3000)
            return
        
        if not url:
            self.toast.show_toast("Please enter a webhook URL", "warning", 3000)
            return
        
        if not url.startswith("https://discord.com/api/webhooks/"):
            self.toast.show_toast("Invalid Discord webhook URL", "warning", 3000)
            return
        
        webhook = {
            "title": title,
            "url": url,
            "selected": len(self.webhooks) == 0
        }
        self.webhooks.append(webhook)
        self._save_webhook_settings()
        self._refresh_webhooks()
        
        # 入力クリア
        self.webhook_title.clear()
        self.webhook_url.clear()
        
        self.toast.show_toast(f"Webhook '{title}' saved", "success", 2000)
    
    def _refresh_webhooks(self):
        """Webhook一覧を更新"""
        # 既存のウィジェットを削除
        while self.webhooks_layout.count():
            item = self.webhooks_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        
        if not self.webhooks:
            empty_label = QLabel("No webhooks saved")
            empty_label.setStyleSheet("color: #808080; font-size: 13px; padding: 20px;")
            empty_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self.webhooks_layout.addWidget(empty_label)
            return
        
        for i, webhook in enumerate(self.webhooks):
            row = QWidget()
            row.setStyleSheet("background-color: #2a2a3a; border-radius: 8px;")
            row_layout = QHBoxLayout(row)
            row_layout.setContentsMargins(15, 10, 15, 10)
            row_layout.setSpacing(12)
            
            # ラジオボタン
            radio = QPushButton("●" if webhook.get("selected") else "○")
            radio.setFixedSize(24, 24)
            radio.setCursor(Qt.CursorShape.PointingHandCursor)
            radio.setStyleSheet(f"background: transparent; color: {'#4a90d9' if webhook.get('selected') else '#606060'}; border: none; font-size: 18px;")
            radio.clicked.connect(lambda _, idx=i: self._select_webhook(idx))
            row_layout.addWidget(radio)
            
            # タイトル
            title_label = QLabel(webhook.get("title", ""))
            title_label.setStyleSheet("color: #ffffff; font-weight: bold; background: transparent;")
            row_layout.addWidget(title_label)
            
            # URL（省略表示）
            url = webhook.get("url", "")
            url_short = url[:40] + "..." if len(url) > 40 else url
            url_label = QLabel(url_short)
            url_label.setStyleSheet("color: #b0b0b0; background: transparent; margin-left: 10px;")
            url_label.setToolTip(url)
            row_layout.addWidget(url_label)
            
            row_layout.addStretch()
            
            # Testボタン
            test_btn = QPushButton("Test")
            test_btn.setFixedSize(50, 28)
            test_btn.setCursor(Qt.CursorShape.PointingHandCursor)
            test_btn.setStyleSheet("""
                QPushButton { background-color: transparent; color: #4a90d9; border: 1px solid #4a90d9;
                    border-radius: 4px; font-size: 11px; }
                QPushButton:hover { background-color: #4a90d922; }
            """)
            test_btn.clicked.connect(lambda _, idx=i: self._test_webhook(idx))
            row_layout.addWidget(test_btn)
            
            # 編集ボタン
            edit_btn = QPushButton()
            edit_btn.setFixedSize(32, 32)
            edit_btn.setCursor(Qt.CursorShape.PointingHandCursor)
            edit_icon = get_icon_from_base64("edit")
            if not edit_icon.isNull():
                edit_btn.setIcon(edit_icon)
                edit_btn.setIconSize(QSize(16, 16))
            else:
                edit_btn.setText("✏️")
            edit_btn.setStyleSheet("background: transparent; border: none; font-size: 16px;")
            edit_btn.clicked.connect(lambda _, idx=i: self._edit_webhook(idx))
            row_layout.addWidget(edit_btn)
            
            # 削除ボタン
            del_btn = QPushButton()
            del_btn.setFixedSize(32, 32)
            del_btn.setCursor(Qt.CursorShape.PointingHandCursor)
            delete_icon = get_icon_from_base64("delete")
            if not delete_icon.isNull():
                del_btn.setIcon(delete_icon)
                del_btn.setIconSize(QSize(16, 16))
            else:
                del_btn.setText("🗑️")
            del_btn.setStyleSheet("background: transparent; border: none; font-size: 16px;")
            del_btn.clicked.connect(lambda _, idx=i: self._delete_webhook(idx))
            row_layout.addWidget(del_btn)
            
            self.webhooks_layout.addWidget(row)
    
    def _select_webhook(self, index):
        """Webhookを選択"""
        for i, webhook in enumerate(self.webhooks):
            webhook["selected"] = (i == index)
        self._save_webhook_settings()
        self._refresh_webhooks()
    
    def _edit_webhook(self, index):
        """Webhookを編集"""
        webhook = self.webhooks[index]
        self.webhook_title.setText(webhook.get("title", ""))
        self.webhook_url.setText(webhook.get("url", ""))
        
        del self.webhooks[index]
        self._save_webhook_settings()
        self._refresh_webhooks()
        
        self.toast.show_toast("Edit mode - modify and save", "info", 2000)
    
    def _delete_webhook(self, index):
        """Webhookを削除"""
        webhook = self.webhooks[index]
        title = webhook.get("title", "")
        
        if QMessageBox.question(self, "Confirm", f"Delete '{title}'?") == QMessageBox.StandardButton.Yes:
            del self.webhooks[index]
            if self.webhooks and not any(w.get("selected") for w in self.webhooks):
                self.webhooks[0]["selected"] = True
            self._save_webhook_settings()
            self._refresh_webhooks()
            self.toast.show_toast(f"'{title}' deleted", "success", 2000)
    
    def _test_webhook(self, index):
        """Webhookをテスト"""
        webhook = self.webhooks[index]
        url = webhook.get("url", "")
        title = webhook.get("title", "")
        
        if not url:
            self.toast.show_toast("No webhook URL", "error", 3000)
            return
        
        try:
            import urllib.request
            import json as json_module
            
            data = {
                "content": f"🔔 Test notification from Project WIN\nWebhook: {title}"
            }
            
            req = urllib.request.Request(
                url,
                data=json_module.dumps(data).encode('utf-8'),
                headers={
                    'Content-Type': 'application/json',
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
                },
                method='POST'
            )
            
            with urllib.request.urlopen(req, timeout=10) as response:
                if response.status in [200, 204]:
                    self.toast.show_toast(f"Test successful!", "success", 3000)
                else:
                    self.toast.show_toast(f"Unexpected response: {response.status}", "warning", 3000)
                    
        except urllib.request.HTTPError as e:
            self.toast.show_toast(f"HTTP Error: {e.code}", "error", 3000)
        except urllib.request.URLError as e:
            self.toast.show_toast(f"Connection failed", "error", 3000)
        except Exception as e:
            self.toast.show_toast(f"Test failed: {str(e)[:30]}", "error", 3000)
    
    def send_success_webhook(self, mode, profile, site, loginid, proxy="", product_title="", image_url=""):
        """成功時にWebhookを送信"""
        print(f"Webhook: send_success_webhook called - mode={mode}")
        
        if not self.webhook_switch.isChecked():
            print("Webhook: Switch is OFF")
            return False
        
        # 選択されているWebhookを取得
        selected_webhook = next((w for w in self.webhooks if w.get("selected")), None)
        if not selected_webhook:
            print("Webhook: No webhook selected")
            return False
        
        url = selected_webhook.get("url", "")
        if not url:
            print("Webhook: No webhook URL")
            return False
        
        print(f"Webhook: Sending to {url[:50]}...")
        
        try:
            import urllib.request
            import json as json_module
            from datetime import datetime
            
            # バージョン取得（ヘルパー関数を使用）
            app_version = get_app_version()
            
            # Proxyが空の場合はLocalと表示
            proxy_display = proxy if proxy else "Local"
            
            # タイムスタンプ
            timestamp = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
            
            # Embed形式でリッチな表示
            fields = [
                {
                    "name": "Profile",
                    "value": profile,
                    "inline": True
                },
                {
                    "name": "Site",
                    "value": site,
                    "inline": True
                },
                {
                    "name": "Mode",
                    "value": mode,
                    "inline": True
                },
                {
                    "name": "Loginid",
                    "value": loginid,
                    "inline": False
                },
                {
                    "name": "Proxy",
                    "value": proxy_display,
                    "inline": False
                }
            ]
            
            # Raffleモードの場合、商品タイトルを追加（改行で間隔を空ける）
            if product_title:
                fields.append({
                    "name": "\u200b",  # 空白フィールドで改行
                    "value": "\u200b",
                    "inline": False
                })
                fields.append({
                    "name": "Product",
                    "value": product_title,
                    "inline": False
                })
            
            embed = {
                "title": f"✅ Successfully {mode}",
                "color": 5763719,  # 緑色 (#57F287)
                "fields": fields,
                "footer": {
                    "text": f"v{app_version} | {timestamp}"
                }
            }
            
            # 商品画像がある場合はthumbnailに設定
            if image_url:
                # .jpg/.png/.webp等の画像URLをそのまま使用
                embed["thumbnail"] = {"url": image_url}
                print(f"Webhook: Thumbnail set: {image_url[:80]}...")
            
            data = {"embeds": [embed]}
            
            req = urllib.request.Request(
                url,
                data=json_module.dumps(data).encode('utf-8'),
                headers={
                    'Content-Type': 'application/json',
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
                },
                method='POST'
            )
            
            with urllib.request.urlopen(req, timeout=10) as response:
                if response.status in [200, 204]:
                    print(f"Webhook: Sent successfully for {mode}")
                    return True
                else:
                    print(f"Webhook: Response status {response.status}")
                    return False
                    
        except Exception as e:
            print(f"Webhook error: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _save_webhook_settings(self):
        """Webhook設定を保存"""
        try:
            data = {
                "enabled": self.webhook_switch.isChecked(),
                "webhooks": self.webhooks
            }
            with open(self.settings_dir / "webhook_settings.json", 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Failed to save webhook settings: {e}")
    
    def _load_webhook_settings(self):
        """Webhook設定を読み込み"""
        self.webhooks = []
        try:
            path = self.settings_dir / "webhook_settings.json"
            if path.exists():
                with open(path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                # シグナルをブロックして保存が発火しないようにする
                self.webhook_switch.blockSignals(True)
                self.webhook_switch.setChecked(data.get("enabled", False))
                self.webhook_switch._update_style()  # 見た目を手動で更新
                self.webhook_switch.blockSignals(False)
                self.webhooks = data.get("webhooks", [])
        except Exception as e:
            print(f"Failed to load webhook settings: {e}")
    
    def _create_general_tab(self):
        """Generalタブを作成"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        # 実行設定グループ
        exec_group = QGroupBox("General Setting")
        exec_group.setStyleSheet(self._group_style())
        exec_layout = QVBoxLayout(exec_group)
        exec_layout.setContentsMargins(20, 25, 20, 20)
        exec_layout.setSpacing(12)
        
        # フォームレイアウト
        form_layout = QFormLayout()
        form_layout.setSpacing(12)
        
        # 並列数
        self.parallel_count = QSpinBox()
        self.parallel_count.setRange(1, 20)
        self.parallel_count.setValue(3)
        self.parallel_count.setStyleSheet(self._spinbox_style())
        form_layout.addRow("Thread:", self.parallel_count)
        
        # Task Delay
        self.task_delay = QSpinBox()
        self.task_delay.setRange(0, 60)
        self.task_delay.setValue(0)
        self.task_delay.setStyleSheet(self._spinbox_style())
        form_layout.addRow("Task Delay(s):", self.task_delay)
        
        # Fetch待機時間
        self.fetch_wait_time = QSpinBox()
        self.fetch_wait_time.setRange(1, 300)
        self.fetch_wait_time.setValue(60)
        self.fetch_wait_time.setStyleSheet(self._spinbox_style())
        form_layout.addRow("Fetch Timeout(s):", self.fetch_wait_time)
        
        # リトライ回数
        self.retry_count = QSpinBox()
        self.retry_count.setRange(0, 10)
        self.retry_count.setValue(3)
        self.retry_count.setStyleSheet(self._spinbox_style())
        form_layout.addRow("Retry:", self.retry_count)
        
        # 保存ボタン（Retryフォームの真下）
        save_btn = QPushButton("Save")
        save_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        save_btn.clicked.connect(self._save_general_settings)
        save_btn.setStyleSheet("""
            QPushButton { background-color: #27ae60; color: white; border: none;
                border-radius: 6px; padding: 10px 25px; font-size: 13px; font-weight: bold; }
            QPushButton:hover { background-color: #2ecc71; }
        """)
        
        # ボタンを左寄せにするためのコンテナ
        btn_container = QWidget()
        btn_layout = QHBoxLayout(btn_container)
        btn_layout.setContentsMargins(0, 0, 0, 0)
        btn_layout.addWidget(save_btn)
        btn_layout.addStretch()
        form_layout.addRow("", btn_container)
        
        exec_layout.addLayout(form_layout)
        
        layout.addWidget(exec_group)
        
        # 説明エリア（枠外）
        desc_label = QLabel("""
<b style="color: #b0b0b0;">Settings Description:</b><br><br>
<b style="color: #909090;">Thread:</b> <span style="color: #909090;">同時に実行するタスクの数を設定します。</span><br><br>
<b style="color: #909090;">Task Delay(s):</b> <span style="color: #909090;">キューで待機中のタスクが開始されるまでの遅延時間（秒）を設定します。</span><br><br>
<b style="color: #909090;">Fetch Timeout(s):</b> <span style="color: #909090;">メールからOTPコードを取得する際の最大待機時間（秒）を設定します。</span><br><br>
<b style="color: #909090;">Retry:</b> <span style="color: #909090;">タスクが失敗した場合の再試行回数を設定します。</span>
        """)
        desc_label.setStyleSheet("color: #909090; font-size: 12px; padding: 10px 5px;")
        desc_label.setWordWrap(True)
        layout.addWidget(desc_label)
        
        layout.addStretch()
        return widget
    
    def _save_general_settings(self):
        """General設定を保存"""
        settings = {
            "parallel_count": self.parallel_count.value(),
            "task_delay": self.task_delay.value(),
            "fetch_wait_time": self.fetch_wait_time.value(),
            "retry_count": self.retry_count.value()
        }
        try:
            with open(self.settings_dir / "general_settings.json", 'w', encoding='utf-8') as f:
                json.dump(settings, f, ensure_ascii=False, indent=2)
            self.toast.show_toast("Settings saved", "success", 2000)
        except Exception as e:
            self.toast.show_toast(f"Failed to save: {str(e)[:30]}", "error", 3000)
    
    def _load_general_settings(self):
        """General設定を読み込み"""
        try:
            path = self.settings_dir / "general_settings.json"
            if path.exists():
                with open(path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                self.parallel_count.setValue(data.get("parallel_count", 3))
                self.task_delay.setValue(data.get("task_delay", 0))
                self.fetch_wait_time.setValue(data.get("fetch_wait_time", 60))
                self.retry_count.setValue(data.get("retry_count", 3))
        except Exception as e:
            print(f"Failed to load general settings: {e}")
    
    def _create_fetch_tab(self):
        widget = QWidget()
        layout = QHBoxLayout(widget)  # 左右レイアウトに変更
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        # 左側: 入力フォーム
        left_container = QWidget()
        left_layout = QVBoxLayout(left_container)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(15)
        
        form_group = QGroupBox("Add Account")
        form_group.setStyleSheet(self._group_style())
        form_layout = QFormLayout(form_group)
        form_layout.setContentsMargins(20, 25, 20, 20)
        form_layout.setSpacing(12)
        
        self.fetch_title = QLineEdit()
        self.fetch_title.setPlaceholderText("Gmail - Main")
        self.fetch_title.setStyleSheet(self._input_style())
        form_layout.addRow("Title:", self.fetch_title)
        
        self.imap_server = QLineEdit()
        self.imap_server.setPlaceholderText("imap.gmail.com")
        self.imap_server.setStyleSheet(self._input_style())
        form_layout.addRow("IMAP Server:", self.imap_server)
        
        self.imap_port = QSpinBox()
        self.imap_port.setRange(1, 65535)
        self.imap_port.setValue(993)
        self.imap_port.setStyleSheet(self._spinbox_style())
        form_layout.addRow("Port:", self.imap_port)
        
        self.fetch_email = QLineEdit()
        self.fetch_email.setPlaceholderText("your-email@gmail.com")
        self.fetch_email.setStyleSheet(self._input_style())
        form_layout.addRow("Mail:", self.fetch_email)
        
        # パスワード(目アイコン付き)
        pass_container = QWidget()
        pass_layout = QHBoxLayout(pass_container)
        pass_layout.setContentsMargins(0, 0, 0, 0)
        pass_layout.setSpacing(0)
        
        self.fetch_password = QLineEdit()
        self.fetch_password.setEchoMode(QLineEdit.EchoMode.Password)
        self.fetch_password.setPlaceholderText("App password")
        self.fetch_password.setStyleSheet("""
            QLineEdit {
                background-color: #2a2a3a; border: 1px solid #404050;
                border-top-left-radius: 6px; border-bottom-left-radius: 6px;
                border-top-right-radius: 0px; border-bottom-right-radius: 0px;
                padding: 10px 15px; color: #ffffff; font-size: 13px; min-width: 150px; max-width: 210px;
            }
            QLineEdit:focus { border-color: #4a90d9; }
        """)
        pass_layout.addWidget(self.fetch_password)
        
        self.eye_btn = QPushButton()
        self.eye_btn.setFixedSize(42, 42)
        self.eye_btn.setCheckable(True)
        self.eye_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        # 初期アイコン設定（Base64から）
        view_icon = get_icon_from_base64("view")
        hide_icon = get_icon_from_base64("hide")
        if not view_icon.isNull():
            self.eye_btn.setIcon(view_icon)
            self.eye_btn.setIconSize(QSize(18, 18))
        else:
            self.eye_btn.setText("👁")
        self.eye_btn.setStyleSheet("""
            QPushButton { background-color: #2a2a3a; border: 1px solid #404050; border-left: none;
                border-top-right-radius: 6px; border-bottom-right-radius: 6px; font-size: 16px; }
            QPushButton:hover { background-color: #3a3a4a; }
        """)
        # トグル時のアイコン切り替え
        def toggle_eye(checked):
            self.fetch_password.setEchoMode(QLineEdit.EchoMode.Normal if checked else QLineEdit.EchoMode.Password)
            if not hide_icon.isNull() and not view_icon.isNull():
                self.eye_btn.setIcon(hide_icon if checked else view_icon)
            else:
                self.eye_btn.setText("🙈" if checked else "👁")
        self.eye_btn.toggled.connect(toggle_eye)
        pass_layout.addWidget(self.eye_btn)
        pass_layout.addStretch()
        form_layout.addRow("Password:", pass_container)
        
        # ボタン
        btn_layout = QHBoxLayout()
        save_btn = QPushButton("Save")
        save_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        save_btn.clicked.connect(self._save_account)
        save_btn.setStyleSheet("""
            QPushButton { background-color: #27ae60; color: white; border: none;
                border-radius: 6px; padding: 10px 25px; font-size: 13px; font-weight: bold; }
            QPushButton:hover { background-color: #2ecc71; }
        """)
        btn_layout.addWidget(save_btn)
        
        test_btn = QPushButton("Connect Test")
        test_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        test_btn.clicked.connect(self._test_connection)
        test_btn.setStyleSheet("""
            QPushButton { background-color: transparent; color: #4a90d9; border: 1px solid #4a90d9;
                border-radius: 6px; padding: 10px 20px; font-size: 13px; }
            QPushButton:hover { background-color: #4a90d922; }
        """)
        btn_layout.addWidget(test_btn)
        btn_layout.addStretch()
        form_layout.addRow("", btn_layout)
        
        left_layout.addWidget(form_group)
        left_layout.addStretch()
        
        # 右側: 保存済みアカウント一覧
        right_container = QWidget()
        right_layout = QVBoxLayout(right_container)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(15)
        
        accounts_group = QGroupBox("Saved Accounts")
        accounts_group.setStyleSheet(self._group_style())
        self.accounts_layout = QVBoxLayout(accounts_group)
        self.accounts_layout.setContentsMargins(20, 25, 20, 20)
        self.accounts_layout.setSpacing(8)
        
        right_layout.addWidget(accounts_group)
        right_layout.addStretch()
        
        # 左右を追加（6:4の比率）
        layout.addWidget(left_container, 6)
        layout.addWidget(right_container, 4)
        
        return widget
    
    def _save_account(self):
        title = self.fetch_title.text().strip()
        server = self.imap_server.text().strip()
        port = self.imap_port.value()
        email = self.fetch_email.text().strip()
        password = self.fetch_password.text()
        
        if not all([title, server, email, password]):
            self.toast.show_toast("Please fill in all fields", "warning", 3000)
            return
        
        account = {
            "title": title, "imap_server": server, "imap_port": port,
            "email": email, "password": password,
            "selected": len(self.fetch_accounts) == 0
        }
        self.fetch_accounts.append(account)
        self._save_settings()
        self._refresh_accounts()
        
        self.fetch_title.clear()
        self.imap_server.clear()
        self.imap_port.setValue(993)
        self.fetch_email.clear()
        self.fetch_password.clear()
        
        # トースト通知
        self.toast.show_toast(f"Account '{title}' saved", "success", 2000)
    
    def _refresh_accounts(self):
        while self.accounts_layout.count():
            item = self.accounts_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        
        for i, acc in enumerate(self.fetch_accounts):
            row = QWidget()
            row.setStyleSheet("background-color: #2a2a3a; border-radius: 8px;")
            row_layout = QHBoxLayout(row)
            row_layout.setContentsMargins(15, 10, 15, 10)
            row_layout.setSpacing(15)
            
            radio = QPushButton("●" if acc.get("selected") else "○")
            radio.setFixedSize(24, 24)
            radio.setCursor(Qt.CursorShape.PointingHandCursor)
            radio.setStyleSheet(f"background: transparent; color: {'#4a90d9' if acc.get('selected') else '#606060'}; border: none; font-size: 18px;")
            radio.clicked.connect(lambda _, idx=i: self._select_account(idx))
            row_layout.addWidget(radio)
            
            title_lbl = QLabel(acc.get("title", ""))
            title_lbl.setStyleSheet("color: #ffffff; font-weight: bold; background: transparent;")
            row_layout.addWidget(title_lbl)
            
            email_lbl = QLabel(acc.get("email", ""))
            email_lbl.setStyleSheet("color: #b0b0b0; background: transparent; margin-left: 10px;")
            row_layout.addWidget(email_lbl)
            row_layout.addStretch()
            
            edit_btn = QPushButton()
            edit_btn.setFixedSize(32, 32)
            edit_btn.setCursor(Qt.CursorShape.PointingHandCursor)
            edit_icon = get_icon_from_base64("edit")
            if not edit_icon.isNull():
                edit_btn.setIcon(edit_icon)
                edit_btn.setIconSize(QSize(16, 16))
            else:
                edit_btn.setText("✏️")
            edit_btn.setStyleSheet("background: transparent; border: none; font-size: 16px;")
            edit_btn.clicked.connect(lambda _, idx=i: self._edit_account(idx))
            row_layout.addWidget(edit_btn)
            
            del_btn = QPushButton()
            del_btn.setFixedSize(32, 32)
            del_btn.setCursor(Qt.CursorShape.PointingHandCursor)
            delete_icon = get_icon_from_base64("delete")
            if not delete_icon.isNull():
                del_btn.setIcon(delete_icon)
                del_btn.setIconSize(QSize(16, 16))
            else:
                del_btn.setText("🗑️")
            del_btn.setStyleSheet("background: transparent; border: none; font-size: 16px;")
            del_btn.clicked.connect(lambda _, idx=i: self._delete_account(idx))
            row_layout.addWidget(del_btn)
            
            self.accounts_layout.addWidget(row)
    
    def _select_account(self, index):
        for i, acc in enumerate(self.fetch_accounts):
            acc["selected"] = (i == index)
        self._save_settings()
        self._refresh_accounts()
    
    def _edit_account(self, index):
        acc = self.fetch_accounts[index]
        self.fetch_title.setText(acc.get("title", ""))
        self.imap_server.setText(acc.get("imap_server", ""))
        self.imap_port.setValue(acc.get("imap_port", 993))
        self.fetch_email.setText(acc.get("email", ""))
        self.fetch_password.setText(acc.get("password", ""))
        del self.fetch_accounts[index]
        self._save_settings()
        self._refresh_accounts()
    
    def _delete_account(self, index):
        acc = self.fetch_accounts[index]
        if QMessageBox.question(self, "確認", f"'{acc.get('title')}' を削除しますか？") == QMessageBox.StandardButton.Yes:
            title = acc.get('title')
            del self.fetch_accounts[index]
            if self.fetch_accounts and not any(a.get("selected") for a in self.fetch_accounts):
                self.fetch_accounts[0]["selected"] = True
            self._save_settings()
            self._refresh_accounts()
            self.toast.show_toast(f"'{title}' を削除しました", "success", 2000)
    
    def _test_connection(self):
        import imaplib
        server = self.imap_server.text().strip()
        port = self.imap_port.value()
        email = self.fetch_email.text().strip()
        password = self.fetch_password.text()
        if not all([server, email, password]):
            self.toast.show_toast("サーバー、メール、パスワードを入力してください", "warning", 3000)
            return
        try:
            mail = imaplib.IMAP4_SSL(server, port)
            mail.login(email, password)
            mail.logout()
            self.toast.show_toast("IMAP接続に成功しました!", "success", 3000)
        except Exception as e:
            self.toast.show_toast(f"接続失敗: {str(e)[:50]}", "error", 5000)
    
    def _save_settings(self):
        selected = next((a for a in self.fetch_accounts if a.get("selected")), None)
        data = {
            "accounts": self.fetch_accounts,
            "imap_server": selected.get("imap_server", "") if selected else "",
            "imap_port": selected.get("imap_port", 993) if selected else 993,
            "email": selected.get("email", "") if selected else "",
            "password": selected.get("password", "") if selected else ""
        }
        try:
            with open(self.settings_dir / "fetch_settings.json", 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Failed to save settings: {e}")
    
    def load_settings(self):
        try:
            path = self.settings_dir / "fetch_settings.json"
            if path.exists():
                with open(path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                self.fetch_accounts = data.get("accounts", [])
                self._refresh_accounts()
        except Exception as e:
            print(f"Failed to load settings: {e}")
        
        # General設定も読み込み
        self._load_general_settings()
        
        # Webhook設定も読み込み
        self._load_webhook_settings()
        self._refresh_webhooks()
        
        # SMS設定も読み込み
        self._load_sms_settings()
    
    def _tab_style(self):
        return """
            QTabWidget::pane { border: 1px solid #404050; border-radius: 10px; background-color: #1e1e2e; }
            QTabBar::tab { background-color: #2a2a3a; color: #b0b0b0; border: 1px solid #404050;
                border-bottom: none; border-top-left-radius: 8px; border-top-right-radius: 8px;
                padding: 10px 25px; margin-right: 2px; font-size: 13px; }
            QTabBar::tab:selected { background-color: #1e1e2e; color: #ffffff; border-bottom: 2px solid #4a90d9; }
            QTabBar::tab:hover:!selected { background-color: #3a3a4a; }
        """
    
    def _group_style(self):
        return """
            QGroupBox { font-size: 14px; font-weight: bold; color: #b0b0b0;
                border: 1px solid #404050; border-radius: 10px; margin-top: 10px; padding-top: 15px; }
            QGroupBox::title { subcontrol-origin: margin; left: 15px; padding: 0 5px; }
        """
    
    def _input_style(self):
        return """
            QLineEdit { background-color: #2a2a3a; border: 1px solid #404050; border-radius: 6px;
                padding: 10px 15px; color: #ffffff; font-size: 13px; min-width: 150px; max-width: 250px; }
            QLineEdit:focus { border-color: #4a90d9; }
        """
    
    def _spinbox_style(self):
        return """
            QSpinBox { background-color: #2a2a3a; border: 1px solid #404050; border-radius: 6px;
                padding: 10px 15px; color: #ffffff; font-size: 13px; min-width: 150px; max-width: 250px; min-height: 20px; }
            QSpinBox:focus { border-color: #4a90d9; }
        """


class ToolsPage(QWidget):
    """ツールページ（タブベース）"""
    
    # GitHubの郵便番号データURL（Raw URL）
    POSTAL_DATA_BASE_URL = "https://raw.githubusercontent.com/TestRaffle/projectwin-assets/main/data/postal_codes"
    
    def __init__(self):
        super().__init__()
        self.settings_dir = SETTINGS_DIR
        self.postal_data = {}  # 郵便番号データのメモリキャッシュ
        self.setup_ui()
    
    def _get_cache_dir(self):
        """キャッシュディレクトリのパスを取得（AppData/Local/ProjectWIN/postal_codes）"""
        if sys.platform == "win32":
            appdata = os.environ.get("LOCALAPPDATA", os.path.expanduser("~"))
            cache_dir = os.path.join(appdata, "ProjectWIN", "postal_codes")
        else:
            cache_dir = os.path.join(os.path.expanduser("~"), ".projectwin", "postal_codes")
        
        # ディレクトリが存在しなければ作成
        if not os.path.exists(cache_dir):
            os.makedirs(cache_dir, exist_ok=True)
        
        return cache_dir
    
    def _download_postal_data(self, prefecture):
        """GitHubから都道府県の郵便番号データをダウンロード"""
        import json
        import urllib.request
        import urllib.error
        
        cache_dir = self._get_cache_dir()
        cache_path = os.path.join(cache_dir, f"{prefecture}.json")
        
        # キャッシュが存在すればそれを使用
        if os.path.exists(cache_path):
            try:
                with open(cache_path, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception as e:
                print(f"キャッシュ読み込みエラー ({prefecture}): {e}")
                # キャッシュが壊れている場合は再ダウンロード
        
        # GitHubからダウンロード
        url = f"{self.POSTAL_DATA_BASE_URL}/{urllib.parse.quote(prefecture)}.json"
        print(f"郵便番号データをダウンロード中: {prefecture}")
        
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "ProjectWIN/1.0"})
            with urllib.request.urlopen(req, timeout=30) as response:
                data = json.loads(response.read().decode("utf-8"))
            
            # キャッシュに保存
            try:
                with open(cache_path, "w", encoding="utf-8") as f:
                    json.dump(data, f, ensure_ascii=False)
            except Exception as e:
                print(f"キャッシュ保存エラー ({prefecture}): {e}")
            
            return data
            
        except urllib.error.HTTPError as e:
            print(f"ダウンロードエラー ({prefecture}): HTTP {e.code}")
            return None
        except urllib.error.URLError as e:
            print(f"ダウンロードエラー ({prefecture}): {e.reason}")
            return None
        except Exception as e:
            print(f"ダウンロードエラー ({prefecture}): {e}")
            return None
    
    def _get_postal_data(self, prefectures):
        """指定された都道府県の郵便番号データを取得（キャッシュまたはダウンロード）"""
        missing_prefs = []
        
        # メモリキャッシュにない都道府県を特定
        for pref in prefectures:
            if pref not in self.postal_data:
                missing_prefs.append(pref)
        
        # 不足分をダウンロード
        for pref in missing_prefs:
            data = self._download_postal_data(pref)
            if data:
                self.postal_data[pref] = data
        
        # 利用可能な都道府県のリストを返す
        available = [p for p in prefectures if p in self.postal_data and len(self.postal_data[p]) > 0]
        return available
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)
        
        header = QLabel("Toolbox")
        header.setStyleSheet("font-size: 24px; font-weight: bold; color: #ffffff; padding-bottom: 10px;")
        layout.addWidget(header)
        
        # トースト通知
        self.toast = ToastNotification(self)
        
        self.tabs = QTabWidget()
        self.tabs.setStyleSheet(self._tab_style())
        
        # 個人情報生成タブ
        self.identity_tab = self._create_identity_tab()
        self.tabs.addTab(self.identity_tab, "Personal Generator")
        
        # パスワード生成タブ
        self.password_tab = self._create_password_tab()
        self.tabs.addTab(self.password_tab, "Password Generator")
        
        layout.addWidget(self.tabs, 1)  # stretch factor 1 でタブを広げる
    
    def _tab_style(self):
        return """
            QTabWidget::pane { border: 1px solid #404050; border-radius: 10px; background-color: #1e1e2e; }
            QTabBar::tab { background-color: #2a2a3a; color: #b0b0b0; border: 1px solid #404050;
                border-bottom: none; border-top-left-radius: 8px; border-top-right-radius: 8px;
                padding: 10px 25px; margin-right: 2px; font-size: 13px; }
            QTabBar::tab:selected { background-color: #1e1e2e; color: #ffffff; border-bottom: 2px solid #4a90d9; }
            QTabBar::tab:hover:!selected { background-color: #3a3a4a; }
        """
    
    def _create_identity_tab(self):
        """個人情報生成タブを作成"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        combo_style = """
            QComboBox { background-color: #2a2a3a; border: 1px solid #404050; border-radius: 6px;
                padding: 6px 10px; color: #ffffff; font-size: 12px; min-width: 140px; }
            QComboBox:focus { border-color: #4a90d9; }
            QComboBox::drop-down { border: none; width: 20px; }
            QComboBox::down-arrow { image: none; border-left: 4px solid transparent;
                border-right: 4px solid transparent; border-top: 5px solid #808080; margin-right: 5px; }
            QComboBox QAbstractItemView { background-color: #2a2a3a; border: 1px solid #404050;
                color: #ffffff; selection-background-color: #4a90d9; }
        """
        
        label_style = "color: #b0b0b0; font-size: 13px;"
        
        spinbox_style_fixed = """
            QSpinBox { background-color: #2a2a3a; border: 1px solid #404050; border-radius: 6px;
                padding: 8px 12px; color: #ffffff; font-size: 13px; min-width: 70px; max-width: 70px; }
            QSpinBox:focus { border-color: #4a90d9; }
            QSpinBox::up-button, QSpinBox::down-button { width: 20px; }
        """
        
        groupbox_style = """
            QGroupBox { font-size: 14px; font-weight: bold; color: #b0b0b0;
                border: 1px solid #404050; border-radius: 10px; margin-top: 10px; padding-top: 15px; }
            QGroupBox::title { subcontrol-origin: margin; left: 15px; padding: 0 5px; }
        """
        
        # ===== 上部エリア（左: Information to Generate、右: Generation Settings）=====
        top_layout = QHBoxLayout()
        top_layout.setSpacing(15)
        
        # --- 左側: Information to Generate ---
        info_group = QGroupBox("Information to Generate")
        info_group.setStyleSheet(groupbox_style)
        info_layout = QVBoxLayout(info_group)
        info_layout.setSpacing(12)
        
        # 氏名行
        name_row = QHBoxLayout()
        name_row.setSpacing(15)
        name_label = QLabel("氏名 :")
        name_label.setStyleSheet(label_style)
        name_label.setFixedWidth(50)
        name_row.addWidget(name_label)
        
        self.chk_name_kanji = TextCheckmarkCheckBox("漢字")
        name_row.addWidget(self.chk_name_kanji)
        
        self.chk_name_kana = TextCheckmarkCheckBox("カタカナ")
        name_row.addWidget(self.chk_name_kana)
        
        self.chk_name_romaji = TextCheckmarkCheckBox("ローマ字")
        name_row.addWidget(self.chk_name_romaji)
        
        self.name_format = QComboBox()
        self.name_format.addItem("区切りなし", "none")
        self.name_format.addItem("区切りあり", "separate")
        self.name_format.setStyleSheet(combo_style)
        name_row.addWidget(self.name_format)
        
        name_row.addStretch()
        info_layout.addLayout(name_row)
        
        # 住所行
        address_row = QHBoxLayout()
        address_row.setSpacing(15)
        self.chk_address = TextCheckmarkCheckBox("住所")
        address_row.addWidget(self.chk_address)
        
        # 市区町村フィルター
        city_filter_label = QLabel("詳細設定:")
        city_filter_label.setStyleSheet("color: #b0b0b0; font-size: 12px;")
        address_row.addWidget(city_filter_label)
        
        self.city_filter = QLineEdit()
        self.city_filter.setPlaceholderText("例: 渋谷区、大阪市（任意）")
        self.city_filter.setStyleSheet("""
            QLineEdit {
                background-color: #3a3a4a; border: 1px solid #404050; border-radius: 4px;
                padding: 4px 8px; color: #ffffff; font-size: 12px; min-width: 150px;
            }
            QLineEdit:focus { border-color: #4a90d9; }
        """)
        self.city_filter.setToolTip("指定した文字列を含む市区町村のみを使用（空欄で全て）")
        address_row.addWidget(self.city_filter)
        
        address_row.addStretch()
        info_layout.addLayout(address_row)
        
        # 電話番号行
        phone_row = QHBoxLayout()
        phone_row.setSpacing(15)
        self.chk_phone = TextCheckmarkCheckBox("電話番号")
        phone_row.addWidget(self.chk_phone)
        
        self.phone_format = QComboBox()
        self.phone_format.addItem("区切りあり (090-1234-5678)", "hyphen")
        self.phone_format.addItem("区切りなし (09012345678)", "none")
        self.phone_format.setStyleSheet(combo_style)
        phone_row.addWidget(self.phone_format)
        
        phone_row.addStretch()
        info_layout.addLayout(phone_row)
        
        # 性別行
        gender_row = QHBoxLayout()
        gender_row.setSpacing(15)
        self.chk_gender = TextCheckmarkCheckBox("性別")
        gender_row.addWidget(self.chk_gender)
        gender_row.addStretch()
        info_layout.addLayout(gender_row)
        
        # 生年月日行
        birthday_row = QHBoxLayout()
        birthday_row.setSpacing(15)
        self.chk_birthday = TextCheckmarkCheckBox("生年月日")
        birthday_row.addWidget(self.chk_birthday)
        
        self.birthday_format = QComboBox()
        self.birthday_format.addItem("YYYY/MM/DD", "slash")
        self.birthday_format.addItem("YYYYMMDD", "none")
        self.birthday_format.addItem("YYYY年MM月DD日", "japanese")
        self.birthday_format.setStyleSheet(combo_style)
        birthday_row.addWidget(self.birthday_format)
        
        birthday_row.addStretch()
        info_layout.addLayout(birthday_row)
        
        top_layout.addWidget(info_group, 1)  # stretch factor 1
        
        # --- 右側: Generation Settings ---
        gen_group = QGroupBox("Generation Settings")
        gen_group.setStyleSheet(groupbox_style)
        gen_layout = QVBoxLayout(gen_group)
        gen_layout.setSpacing(12)
        
        # 生成数
        count_row = QHBoxLayout()
        count_label = QLabel("生成数:")
        count_label.setStyleSheet(label_style)
        count_label.setFixedWidth(70)
        self.gen_count = QSpinBox()
        self.gen_count.setRange(1, 10000)
        self.gen_count.setValue(100)
        self.gen_count.setStyleSheet(spinbox_style_fixed)
        count_row.addWidget(count_label)
        count_row.addWidget(self.gen_count)
        count_row.addStretch()
        gen_layout.addLayout(count_row)
        
        # 年齢範囲
        age_row = QHBoxLayout()
        age_label = QLabel("年齢範囲:")
        age_label.setStyleSheet(label_style)
        age_label.setFixedWidth(70)
        self.age_min = QSpinBox()
        self.age_min.setRange(0, 120)
        self.age_min.setValue(20)
        self.age_min.setStyleSheet(spinbox_style_fixed)
        age_separator = QLabel("〜")
        age_separator.setStyleSheet("color: #ffffff; font-size: 14px;")
        self.age_max = QSpinBox()
        self.age_max.setRange(0, 120)
        self.age_max.setValue(40)
        self.age_max.setStyleSheet(spinbox_style_fixed)
        age_row.addWidget(age_label)
        age_row.addWidget(self.age_min)
        age_row.addWidget(age_separator)
        age_row.addWidget(self.age_max)
        age_row.addStretch()
        gen_layout.addLayout(age_row)
        
        # 男女比率
        ratio_label = QLabel("男女比率:")
        ratio_label.setStyleSheet(label_style)
        gen_layout.addWidget(ratio_label)
        
        # ラジオボタン行
        male_values = [100, 90, 80, 70, 60, 50, 40, 30, 20, 10, 0]
        self.ratio_buttons = []
        
        ratio_buttons_row = QHBoxLayout()
        ratio_buttons_row.setSpacing(4)
        ratio_buttons_row.addSpacing(30)  # 「男性」ラベル分のスペース
        for val in male_values:
            radio_btn = QPushButton("●" if val == 50 else "○")
            radio_btn.setFixedSize(20, 20)
            radio_btn.setCursor(Qt.CursorShape.PointingHandCursor)
            radio_btn.setStyleSheet(f"background: transparent; color: {'#4a90d9' if val == 50 else '#606060'}; border: none; font-size: 14px;")
            radio_btn.setProperty("ratio_value", val)
            radio_btn.clicked.connect(lambda _, v=val: self._select_ratio(v))
            ratio_buttons_row.addWidget(radio_btn)
            self.ratio_buttons.append(radio_btn)
        ratio_buttons_row.addStretch()
        gen_layout.addLayout(ratio_buttons_row)
        
        # 男性パーセント行
        male_row = QHBoxLayout()
        male_row.setSpacing(4)
        male_lbl = QLabel("男性")
        male_lbl.setStyleSheet("color: #b0b0b0; font-size: 10px;")
        male_lbl.setFixedWidth(30)
        male_row.addWidget(male_lbl)
        for val in male_values:
            pct = QLabel(f"{val}%")
            pct.setStyleSheet("color: #b0b0b0; font-size: 9px;")
            pct.setFixedWidth(20)
            pct.setAlignment(Qt.AlignmentFlag.AlignCenter)
            male_row.addWidget(pct)
        male_row.addStretch()
        gen_layout.addLayout(male_row)
        
        # 女性パーセント行
        female_row = QHBoxLayout()
        female_row.setSpacing(4)
        female_lbl = QLabel("女性")
        female_lbl.setStyleSheet("color: #b0b0b0; font-size: 10px;")
        female_lbl.setFixedWidth(30)
        female_row.addWidget(female_lbl)
        for val in male_values:
            pct = QLabel(f"{100-val}%")
            pct.setStyleSheet("color: #b0b0b0; font-size: 9px;")
            pct.setFixedWidth(20)
            pct.setAlignment(Qt.AlignmentFlag.AlignCenter)
            female_row.addWidget(pct)
        female_row.addStretch()
        gen_layout.addLayout(female_row)
        
        self.selected_ratio = 50
        
        # 出力形式
        output_row = QHBoxLayout()
        output_row.setSpacing(10)
        output_label = QLabel("出力形式:")
        output_label.setStyleSheet(label_style)
        output_label.setFixedWidth(70)
        
        self.radio_xlsx_btn = QPushButton("●")
        self.radio_xlsx_btn.setFixedSize(20, 20)
        self.radio_xlsx_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.radio_xlsx_btn.setStyleSheet("background: transparent; color: #4a90d9; border: none; font-size: 14px;")
        self.radio_xlsx_btn.clicked.connect(lambda: self._select_output_format("xlsx"))
        
        xlsx_label = QLabel("Excel")
        xlsx_label.setStyleSheet("color: #ffffff; font-size: 12px;")
        
        self.radio_csv_btn = QPushButton("○")
        self.radio_csv_btn.setFixedSize(20, 20)
        self.radio_csv_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.radio_csv_btn.setStyleSheet("background: transparent; color: #606060; border: none; font-size: 14px;")
        self.radio_csv_btn.clicked.connect(lambda: self._select_output_format("csv"))
        
        csv_label = QLabel("CSV")
        csv_label.setStyleSheet("color: #ffffff; font-size: 12px;")
        
        output_row.addWidget(output_label)
        output_row.addWidget(self.radio_xlsx_btn)
        output_row.addWidget(xlsx_label)
        output_row.addSpacing(10)
        output_row.addWidget(self.radio_csv_btn)
        output_row.addWidget(csv_label)
        output_row.addStretch()
        gen_layout.addLayout(output_row)
        
        self.selected_output_format = "xlsx"
        
        gen_layout.addStretch()
        top_layout.addWidget(gen_group, 1)  # stretch factor 1 で残りのスペースを使う
        
        layout.addLayout(top_layout)
        
        # ===== 都道府県選択 =====
        pref_group = QGroupBox("Prefecture (複数選択可)")
        pref_group.setStyleSheet(groupbox_style)
        pref_layout = QVBoxLayout(pref_group)
        
        # 全選択/全解除ボタン
        pref_btn_layout = QHBoxLayout()
        select_all_btn = QPushButton("全選択")
        select_all_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        select_all_btn.clicked.connect(self._select_all_prefectures)
        select_all_btn.setStyleSheet("""
            QPushButton { background-color: #3a3a4a; color: #ffffff; border: 1px solid #404050;
                border-radius: 4px; padding: 5px 15px; font-size: 12px; }
            QPushButton:hover { background-color: #4a4a5a; }
        """)
        deselect_all_btn = QPushButton("全解除")
        deselect_all_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        deselect_all_btn.clicked.connect(self._deselect_all_prefectures)
        deselect_all_btn.setStyleSheet("""
            QPushButton { background-color: #3a3a4a; color: #ffffff; border: 1px solid #404050;
                border-radius: 4px; padding: 5px 15px; font-size: 12px; }
            QPushButton:hover { background-color: #4a4a5a; }
        """)
        pref_btn_layout.addWidget(select_all_btn)
        pref_btn_layout.addWidget(deselect_all_btn)
        pref_btn_layout.addStretch()
        pref_layout.addLayout(pref_btn_layout)
        
        # 都道府県チェックボックス
        pref_grid = QGridLayout()
        pref_grid.setSpacing(5)
        
        prefectures = [
            "北海道", "青森県", "岩手県", "宮城県", "秋田県", "山形県", "福島県",
            "茨城県", "栃木県", "群馬県", "埼玉県", "千葉県", "東京都", "神奈川県",
            "新潟県", "富山県", "石川県", "福井県", "山梨県", "長野県", "岐阜県",
            "静岡県", "愛知県", "三重県", "滋賀県", "京都府", "大阪府", "兵庫県",
            "奈良県", "和歌山県", "鳥取県", "島根県", "岡山県", "広島県", "山口県",
            "徳島県", "香川県", "愛媛県", "高知県", "福岡県", "佐賀県", "長崎県",
            "熊本県", "大分県", "宮崎県", "鹿児島県", "沖縄県"
        ]
        
        self.pref_checkboxes = []
        
        for i, pref in enumerate(prefectures):
            chk = TextCheckmarkCheckBox(pref)
            chk.set_box_size(14)
            self.pref_checkboxes.append(chk)
            pref_grid.addWidget(chk, i // 8, i % 8)
        
        pref_layout.addLayout(pref_grid)
        layout.addWidget(pref_group, 1)
        
        # ===== Generateボタン =====
        gen_btn_layout = QHBoxLayout()
        generate_btn = QPushButton("Generate")
        generate_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        generate_btn.clicked.connect(self._generate_identities)
        generate_btn.setStyleSheet("""
            QPushButton { background-color: #27ae60; color: white; border: none;
                border-radius: 6px; padding: 12px 40px; font-size: 14px; font-weight: bold; }
            QPushButton:hover { background-color: #2ecc71; }
        """)
        gen_btn_layout.addWidget(generate_btn)
        gen_btn_layout.addStretch()
        layout.addLayout(gen_btn_layout)
        
        return widget
    
    def _select_ratio(self, value):
        """男女比率を選択"""
        self.selected_ratio = value
        male_values = [100, 90, 80, 70, 60, 50, 40, 30, 20, 10, 0]
        for i, btn in enumerate(self.ratio_buttons):
            is_selected = male_values[i] == value
            btn.setText("●" if is_selected else "○")
            btn.setStyleSheet(f"background: transparent; color: {'#4a90d9' if is_selected else '#606060'}; border: none; font-size: 18px;")
    
    def _select_output_format(self, fmt):
        """出力形式を選択"""
        self.selected_output_format = fmt
        is_xlsx = fmt == "xlsx"
        self.radio_xlsx_btn.setText("●" if is_xlsx else "○")
        self.radio_xlsx_btn.setStyleSheet(f"background: transparent; color: {'#4a90d9' if is_xlsx else '#606060'}; border: none; font-size: 18px;")
        self.radio_csv_btn.setText("●" if not is_xlsx else "○")
        self.radio_csv_btn.setStyleSheet(f"background: transparent; color: {'#4a90d9' if not is_xlsx else '#606060'}; border: none; font-size: 18px;")
    
    def _select_all_prefectures(self):
        """全都道府県を選択"""
        for chk in self.pref_checkboxes:
            chk.setChecked(True)
    
    def _deselect_all_prefectures(self):
        """全都道府県を解除"""
        for chk in self.pref_checkboxes:
            chk.setChecked(False)
    
    def _kana_to_romaji(self, kana):
        """カタカナをローマ字に変換"""
        kana_romaji_map = {
            'ア': 'A', 'イ': 'I', 'ウ': 'U', 'エ': 'E', 'オ': 'O',
            'カ': 'Ka', 'キ': 'Ki', 'ク': 'Ku', 'ケ': 'Ke', 'コ': 'Ko',
            'サ': 'Sa', 'シ': 'Shi', 'ス': 'Su', 'セ': 'Se', 'ソ': 'So',
            'タ': 'Ta', 'チ': 'Chi', 'ツ': 'Tsu', 'テ': 'Te', 'ト': 'To',
            'ナ': 'Na', 'ニ': 'Ni', 'ヌ': 'Nu', 'ネ': 'Ne', 'ノ': 'No',
            'ハ': 'Ha', 'ヒ': 'Hi', 'フ': 'Fu', 'ヘ': 'He', 'ホ': 'Ho',
            'マ': 'Ma', 'ミ': 'Mi', 'ム': 'Mu', 'メ': 'Me', 'モ': 'Mo',
            'ヤ': 'Ya', 'ユ': 'Yu', 'ヨ': 'Yo',
            'ラ': 'Ra', 'リ': 'Ri', 'ル': 'Ru', 'レ': 'Re', 'ロ': 'Ro',
            'ワ': 'Wa', 'ヲ': 'Wo', 'ン': 'N',
            'ガ': 'Ga', 'ギ': 'Gi', 'グ': 'Gu', 'ゲ': 'Ge', 'ゴ': 'Go',
            'ザ': 'Za', 'ジ': 'Ji', 'ズ': 'Zu', 'ゼ': 'Ze', 'ゾ': 'Zo',
            'ダ': 'Da', 'ヂ': 'Di', 'ヅ': 'Du', 'デ': 'De', 'ド': 'Do',
            'バ': 'Ba', 'ビ': 'Bi', 'ブ': 'Bu', 'ベ': 'Be', 'ボ': 'Bo',
            'パ': 'Pa', 'ピ': 'Pi', 'プ': 'Pu', 'ペ': 'Pe', 'ポ': 'Po',
            'キャ': 'Kya', 'キュ': 'Kyu', 'キョ': 'Kyo',
            'シャ': 'Sha', 'シュ': 'Shu', 'ショ': 'Sho',
            'チャ': 'Cha', 'チュ': 'Chu', 'チョ': 'Cho',
            'ニャ': 'Nya', 'ニュ': 'Nyu', 'ニョ': 'Nyo',
            'ヒャ': 'Hya', 'ヒュ': 'Hyu', 'ヒョ': 'Hyo',
            'ミャ': 'Mya', 'ミュ': 'Myu', 'ミョ': 'Myo',
            'リャ': 'Rya', 'リュ': 'Ryu', 'リョ': 'Ryo',
            'ギャ': 'Gya', 'ギュ': 'Gyu', 'ギョ': 'Gyo',
            'ジャ': 'Ja', 'ジュ': 'Ju', 'ジョ': 'Jo',
            'ビャ': 'Bya', 'ビュ': 'Byu', 'ビョ': 'Byo',
            'ピャ': 'Pya', 'ピュ': 'Pyu', 'ピョ': 'Pyo',
            'ッ': '', 'ー': '',
            'ァ': 'a', 'ィ': 'i', 'ゥ': 'u', 'ェ': 'e', 'ォ': 'o',
        }
        result = []
        i = 0
        while i < len(kana):
            # 2文字の組み合わせを先にチェック
            if i + 1 < len(kana) and kana[i:i+2] in kana_romaji_map:
                result.append(kana_romaji_map[kana[i:i+2]])
                i += 2
            elif kana[i] in kana_romaji_map:
                result.append(kana_romaji_map[kana[i]])
                i += 1
            else:
                result.append(kana[i])
                i += 1
        # 最初の文字だけ大文字、残りは小文字に
        romaji = ''.join(result)
        if romaji:
            romaji = romaji[0].upper() + romaji[1:].lower()
        return romaji
    
    def _get_name_data(self):
        """漢字-カタカナ対応の名前データを取得"""
        # 苗字（漢字, カタカナ）
        last_names = [
            ("佐藤", "サトウ"), ("鈴木", "スズキ"), ("高橋", "タカハシ"), ("田中", "タナカ"), ("伊藤", "イトウ"),
            ("渡辺", "ワタナベ"), ("山本", "ヤマモト"), ("中村", "ナカムラ"), ("小林", "コバヤシ"), ("加藤", "カトウ"),
            ("吉田", "ヨシダ"), ("山田", "ヤマダ"), ("佐々木", "ササキ"), ("山口", "ヤマグチ"), ("松本", "マツモト"),
            ("井上", "イノウエ"), ("木村", "キムラ"), ("林", "ハヤシ"), ("斎藤", "サイトウ"), ("清水", "シミズ"),
            ("山崎", "ヤマザキ"), ("森", "モリ"), ("池田", "イケダ"), ("橋本", "ハシモト"), ("阿部", "アベ"),
            ("石川", "イシカワ"), ("山下", "ヤマシタ"), ("中島", "ナカジマ"), ("石井", "イシイ"), ("小川", "オガワ"),
            ("前田", "マエダ"), ("岡田", "オカダ"), ("長谷川", "ハセガワ"), ("藤田", "フジタ"), ("後藤", "ゴトウ"),
            ("近藤", "コンドウ"), ("村上", "ムラカミ"), ("遠藤", "エンドウ"), ("青木", "アオキ"), ("坂本", "サカモト"),
            ("藤井", "フジイ"), ("西村", "ニシムラ"), ("福田", "フクダ"), ("太田", "オオタ"), ("三浦", "ミウラ"),
            ("藤原", "フジワラ"), ("岡本", "オカモト"), ("松田", "マツダ"), ("中野", "ナカノ"), ("原田", "ハラダ"),
            ("小野", "オノ"), ("田村", "タムラ"), ("竹内", "タケウチ"), ("金子", "カネコ"), ("和田", "ワダ"),
            ("中山", "ナカヤマ"), ("石田", "イシダ"), ("上田", "ウエダ"), ("森田", "モリタ"), ("原", "ハラ"),
            ("柴田", "シバタ"), ("酒井", "サカイ"), ("工藤", "クドウ"), ("横山", "ヨコヤマ"), ("宮崎", "ミヤザキ"),
            ("宮本", "ミヤモト"), ("内田", "ウチダ"), ("高木", "タカギ"), ("安藤", "アンドウ"), ("谷口", "タニグチ"),
            ("大野", "オオノ"), ("丸山", "マルヤマ"), ("今井", "イマイ"), ("河野", "コウノ"), ("藤本", "フジモト"),
            ("村田", "ムラタ"), ("武田", "タケダ"), ("上野", "ウエノ"), ("杉山", "スギヤマ"), ("増田", "マスダ"),
            ("平野", "ヒラノ"), ("大塚", "オオツカ"), ("千葉", "チバ"), ("久保", "クボ"), ("松井", "マツイ"),
            ("小島", "コジマ"), ("岩崎", "イワサキ"), ("野口", "ノグチ"), ("菊地", "キクチ"), ("木下", "キノシタ"),
            ("野村", "ノムラ"), ("新井", "アライ"), ("渡部", "ワタベ"), ("櫻井", "サクライ"), ("佐野", "サノ"),
            ("古川", "フルカワ"), ("熊谷", "クマガイ"), ("菅原", "スガワラ"), ("杉本", "スギモト"), ("市川", "イチカワ"),
        ]
        
        # 男性名（漢字, カタカナ）
        first_names_male = [
            ("大翔", "ヒロト"), ("蓮", "レン"), ("悠真", "ユウマ"), ("陽翔", "ハルト"), ("湊", "ミナト"),
            ("樹", "イツキ"), ("悠人", "ユウト"), ("朝陽", "アサヒ"), ("颯太", "ソウタ"), ("蒼", "アオイ"),
            ("陽太", "ヨウタ"), ("結翔", "ユイト"), ("翔", "ショウ"), ("大和", "ヤマト"), ("颯真", "ソウマ"),
            ("健太", "ケンタ"), ("拓海", "タクミ"), ("翔太", "ショウタ"), ("雄大", "ユウダイ"), ("太一", "タイチ"),
            ("翔平", "ショウヘイ"), ("海斗", "カイト"), ("健", "ケン"), ("大輝", "ダイキ"), ("直樹", "ナオキ"),
            ("隼人", "ハヤト"), ("駿", "シュン"), ("涼太", "リョウタ"), ("和也", "カズヤ"), ("拓也", "タクヤ"),
            ("達也", "タツヤ"), ("康平", "コウヘイ"), ("裕太", "ユウタ"), ("亮", "リョウ"), ("健太郎", "ケンタロウ"),
            ("誠", "マコト"), ("学", "マナブ"), ("充", "ミツル"), ("篤司", "アツシ"), ("慎吾", "シンゴ"),
            ("浩二", "コウジ"), ("哲也", "テツヤ"), ("秀樹", "ヒデキ"), ("正樹", "マサキ"), ("雅彦", "マサヒコ"),
            ("博", "ヒロシ"), ("修", "オサム"), ("淳", "アツシ"), ("剛", "ツヨシ"), ("豊", "ユタカ"),
        ]
        
        # 女性名（漢字, カタカナ）
        first_names_female = [
            ("陽葵", "ヒマリ"), ("芽依", "メイ"), ("凛", "リン"), ("詩", "ウタ"), ("結菜", "ユイナ"),
            ("葵", "アオイ"), ("紬", "ツムギ"), ("咲良", "サクラ"), ("結愛", "ユア"), ("莉子", "リコ"),
            ("美月", "ミヅキ"), ("心春", "コハル"), ("美桜", "ミオ"), ("一花", "イチカ"), ("杏", "アン"),
            ("美咲", "ミサキ"), ("さくら", "サクラ"), ("花", "ハナ"), ("彩", "アヤ"), ("愛", "アイ"),
            ("七海", "ナナミ"), ("美優", "ミユ"), ("真央", "マオ"), ("遥", "ハルカ"), ("楓", "カエデ"),
            ("千尋", "チヒロ"), ("麻衣", "マイ"), ("優花", "ユウカ"), ("彩花", "アヤカ"), ("菜々子", "ナナコ"),
            ("沙織", "サオリ"), ("理恵", "リエ"), ("絵美", "エミ"), ("明日香", "アスカ"), ("由美", "ユミ"),
            ("恵", "メグミ"), ("舞", "マイ"), ("瞳", "ヒトミ"), ("香織", "カオリ"), ("智子", "トモコ"),
            ("裕子", "ユウコ"), ("直美", "ナオミ"), ("美紀", "ミキ"), ("知美", "トモミ"), ("真由美", "マユミ"),
            ("京子", "キョウコ"), ("典子", "ノリコ"), ("洋子", "ヨウコ"), ("和子", "カズコ"), ("幸子", "サチコ"),
        ]
        
        return last_names, first_names_male, first_names_female
    
    def _generate_identities(self):
        """個人情報を生成"""
        try:
            from faker import Faker
            import random
            from datetime import datetime
            import os
            
            fake = Faker('ja_JP')
            
            # 設定を取得
            count = self.gen_count.value()
            min_age = self.age_min.value()
            max_age = self.age_max.value()
            male_ratio = self.selected_ratio  # ラジオボタンから取得
            
            # 住所が選択されている場合のみ都道府県チェック
            available_prefs = []
            filtered_postal_data = {}  # 市区町村フィルター適用後のデータ
            
            if self.chk_address.isChecked():
                # 選択された都道府県
                selected_prefs = [chk.text() for chk in self.pref_checkboxes if chk.isChecked()]
                if not selected_prefs:
                    self.toast.show_toast("都道府県を1つ以上選択してください", "error")
                    return
                
                # 郵便番号データを取得（キャッシュまたはGitHubからダウンロード）
                self.toast.show_toast(f"郵便番号データを準備中...", "info")
                QApplication.processEvents()  # UIを更新
                
                available_prefs = self._get_postal_data(selected_prefs)
                
                if not available_prefs:
                    self.toast.show_toast("郵便番号データの取得に失敗しました。インターネット接続を確認してください", "error")
                    return
                
                # 市区町村フィルターを取得
                city_filter_text = self.city_filter.text().strip()
                
                if city_filter_text:
                    # フィルター適用：指定した文字列を含む市区町村のみ抽出
                    for pref in available_prefs:
                        filtered_entries = [
                            entry for entry in self.postal_data[pref]
                            if city_filter_text in entry.get("city", "")
                        ]
                        if filtered_entries:
                            filtered_postal_data[pref] = filtered_entries
                    
                    # フィルター後にデータが残っている都道府県のみ使用
                    available_prefs = list(filtered_postal_data.keys())
                    
                    if not available_prefs:
                        self.toast.show_toast(f"「{city_filter_text}」を含む市区町村が見つかりません", "error")
                        return
                    
                    # フィルター後の件数を表示
                    total_entries = sum(len(v) for v in filtered_postal_data.values())
                    self.toast.show_toast(f"「{city_filter_text}」: {total_entries}件の住所が見つかりました", "info")
                else:
                    # フィルターなし：すべてのデータを使用
                    filtered_postal_data = {pref: self.postal_data[pref] for pref in available_prefs}
            
            # フォーマット
            phone_fmt = self.phone_format.currentData()
            birthday_fmt = self.birthday_format.currentData()
            output_fmt = self.selected_output_format  # ラジオボタンから取得
            
            # マンション名のリスト
            mansion_names = ["グランドメゾン", "パークハイツ", "ライオンズマンション", "プラウド", "ブリリア",
                            "シティタワー", "パークシティ", "ガーデンズ", "レジデンス", "コートハウス",
                            "グリーンハイツ", "サンライズ", "オーシャンビュー", "スカイタワー", "リバーサイド",
                            "フォレストヒルズ", "アーバンライフ", "センチュリー", "ロイヤルハイツ", "エスペランサ",
                            "クレストコート", "ベルメゾン", "シャトレー", "ドミール", "メゾンドール",
                            "ヴィラージュ", "アクアテラス", "ソレイユ", "ラ・メゾン", "エクセレント"]
            
            # データ生成
            data = []
            name_fmt = self.name_format.currentData()  # "none" or "separate"
            
            for _ in range(count):
                row = {}
                
                # 性別決定
                is_male = random.randint(1, 100) <= male_ratio
                gender = "男" if is_male else "女"
                
                # 氏名（漢字-カタカナ連動辞書から取得）
                last_names, first_names_male, first_names_female = self._get_name_data()
                
                # 苗字を選択（漢字, カタカナのペア）
                last_name_pair = random.choice(last_names)
                last_name = last_name_pair[0]
                last_kana = last_name_pair[1]
                
                # 名前を選択（性別に応じて）
                if is_male:
                    first_name_pair = random.choice(first_names_male)
                else:
                    first_name_pair = random.choice(first_names_female)
                first_name = first_name_pair[0]
                first_kana = first_name_pair[1]
                
                # ローマ字変換（カタカナベース）
                last_romaji = self._kana_to_romaji(last_kana)
                first_romaji = self._kana_to_romaji(first_kana)
                
                # === 氏名 ===
                if self.chk_name_kanji.isChecked():
                    if name_fmt == "separate":
                        row["姓_漢字"] = last_name
                        row["名_漢字"] = first_name
                    else:
                        row["氏名_漢字"] = f"{last_name} {first_name}"
                
                if self.chk_name_kana.isChecked():
                    if name_fmt == "separate":
                        row["姓_カタカナ"] = last_kana
                        row["名_カタカナ"] = first_kana
                    else:
                        row["氏名_カタカナ"] = f"{last_kana} {first_kana}"
                
                if self.chk_name_romaji.isChecked():
                    if name_fmt == "separate":
                        row["姓_ローマ字"] = last_romaji
                        row["名_ローマ字"] = first_romaji
                    else:
                        row["氏名_ローマ字"] = f"{last_romaji} {first_romaji}"
                
                # === 住所 ===
                if self.chk_address.isChecked():
                    # 利用可能な都道府県からランダムに選択
                    pref = random.choice(available_prefs)
                    
                    # フィルター済みの郵便番号データからランダムに選択
                    addr_entry = random.choice(filtered_postal_data[pref])
                    zipcode = addr_entry["zip"]
                    pref_name = addr_entry["pref"]
                    city = addr_entry["city"]
                    town = addr_entry["town"]
                    
                    # 番地を生成（1-30番地 1-20号）
                    banchi = random.randint(1, 30)
                    go = random.randint(1, 20)
                    banchi_str = f"{banchi}-{go}"
                    
                    # マンションをランダムで追加（50%の確率）
                    if random.random() < 0.5:
                        mansion = random.choice(mansion_names)
                        room = random.randint(101, 1505)
                        mansion_str = f"{mansion}{room}号室"
                    else:
                        mansion_str = ""
                    
                    # セル分割で出力
                    row["郵便番号"] = zipcode
                    row["都道府県"] = pref_name
                    row["市区町村"] = city
                    row["町名"] = town
                    row["番地"] = banchi_str
                    row["建物名"] = mansion_str
                
                # === 電話番号 ===
                if self.chk_phone.isChecked():
                    # 090/080/070のいずれかで始まる11桁の電話番号を生成
                    prefix = random.choice(["090", "080", "070"])
                    remaining = ''.join([str(random.randint(0, 9)) for _ in range(8)])
                    if phone_fmt == "hyphen":
                        phone = f"{prefix}-{remaining[:4]}-{remaining[4:]}"
                    else:  # none
                        phone = f"{prefix}{remaining}"
                    row["電話番号"] = phone
                
                # === 性別 ===
                if self.chk_gender.isChecked():
                    row["性別"] = gender
                
                # === 生年月日 ===
                if self.chk_birthday.isChecked():
                    birthday = fake.date_of_birth(minimum_age=min_age, maximum_age=max_age)
                    if birthday_fmt == "slash":
                        row["生年月日"] = birthday.strftime("%Y/%m/%d")
                    elif birthday_fmt == "none":
                        row["生年月日"] = birthday.strftime("%Y%m%d")
                    else:  # japanese
                        row["生年月日"] = birthday.strftime("%Y年%m月%d日")
                
                data.append(row)
            
            # 出力ディレクトリ（_internal/Export/IdentityGenerator）
            if getattr(sys, 'frozen', False):
                base_dir = os.path.dirname(sys.executable)
            else:
                base_dir = os.path.dirname(os.path.abspath(__file__))
            export_dir = os.path.join(base_dir, "_internal", "Export", "PersonalGenerator")
            
            os.makedirs(export_dir, exist_ok=True)
            
            # ファイル名（タイムスタンプ付き）
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            if output_fmt == "xlsx":
                # Excel出力
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Generated Data"
                
                # ヘッダー
                if data:
                    headers = list(data[0].keys())
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center")
                    
                    # データ
                    for row_idx, row_data in enumerate(data, 2):
                        for col_idx, header in enumerate(headers, 1):
                            ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
                    
                    # 列幅調整
                    for col in ws.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column].width = adjusted_width
                
                filepath = os.path.join(export_dir, f"personal_{timestamp}.xlsx")
                wb.save(filepath)
                
            else:
                # CSV出力
                import csv
                filepath = os.path.join(export_dir, f"personal_{timestamp}.csv")
                
                if data:
                    headers = list(data[0].keys())
                    with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
                        writer = csv.DictWriter(f, fieldnames=headers)
                        writer.writeheader()
                        writer.writerows(data)
            
            self.toast.show_toast(f"生成完了: {filepath}", "success")
            
            # フォルダを開く
            if os.path.exists(export_dir):
                os.startfile(export_dir)
                
        except ImportError as e:
            self.toast.show_toast(f"必要なライブラリがありません: {e}", "error")
        except Exception as e:
            self.toast.show_toast(f"生成エラー: {e}", "error")
            import traceback
            traceback.print_exc()
    
    def _create_password_tab(self):
        """パスワード生成タブを作成"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        # スクロールエリア
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea { border: none; background-color: transparent; }")
        
        scroll_content = QWidget()
        scroll_layout = QVBoxLayout(scroll_content)
        scroll_layout.setSpacing(20)
        
        label_style = "color: #b0b0b0; font-size: 13px;"
        
        spinbox_style = """
            QSpinBox { background-color: #2a2a3a; border: 1px solid #404050; border-radius: 6px;
                padding: 8px 12px; color: #ffffff; font-size: 13px; min-width: 70px; max-width: 70px; }
            QSpinBox:focus { border-color: #4a90d9; }
            QSpinBox::up-button, QSpinBox::down-button { width: 20px; }
        """
        
        # ===== 文字種の選択 =====
        char_group = QGroupBox("Character Types")
        char_group.setStyleSheet("""
            QGroupBox { font-size: 14px; font-weight: bold; color: #b0b0b0;
                border: 1px solid #404050; border-radius: 10px; margin-top: 10px; padding: 15px 10px 15px 10px; }
            QGroupBox::title { subcontrol-origin: margin; left: 15px; padding: 0 5px; }
        """)
        char_layout = QHBoxLayout(char_group)
        char_layout.setSpacing(20)
        char_layout.setContentsMargins(10, 10, 10, 10)
        
        self.chk_uppercase = TextCheckmarkCheckBox("英字(大文字)")
        char_layout.addWidget(self.chk_uppercase)
        
        self.chk_lowercase = TextCheckmarkCheckBox("英字(小文字)")
        char_layout.addWidget(self.chk_lowercase)
        
        self.chk_numbers = TextCheckmarkCheckBox("数字")
        char_layout.addWidget(self.chk_numbers)
        
        self.chk_symbols = TextCheckmarkCheckBox("記号")
        char_layout.addWidget(self.chk_symbols)
        
        char_layout.addStretch()
        scroll_layout.addWidget(char_group)
        
        # ===== 生成設定 =====
        gen_group = QGroupBox("Generation Settings")
        gen_group.setStyleSheet("""
            QGroupBox { font-size: 14px; font-weight: bold; color: #b0b0b0;
                border: 1px solid #404050; border-radius: 10px; margin-top: 10px; padding: 15px 10px 15px 10px; }
            QGroupBox::title { subcontrol-origin: margin; left: 15px; padding: 0 5px; }
        """)
        gen_layout = QVBoxLayout(gen_group)
        gen_layout.setSpacing(15)
        gen_layout.setContentsMargins(10, 10, 10, 10)
        
        # 文字数
        length_row = QHBoxLayout()
        length_label = QLabel("文字数:")
        length_label.setStyleSheet(label_style)
        length_label.setFixedWidth(60)
        self.pw_length = QSpinBox()
        self.pw_length.setRange(4, 128)
        self.pw_length.setValue(10)
        self.pw_length.setStyleSheet(spinbox_style)
        length_row.addWidget(length_label)
        length_row.addWidget(self.pw_length)
        length_row.addStretch()
        gen_layout.addLayout(length_row)
        
        # 作成数
        count_row = QHBoxLayout()
        count_label = QLabel("作成数:")
        count_label.setStyleSheet(label_style)
        count_label.setFixedWidth(60)
        self.pw_count = QSpinBox()
        self.pw_count.setRange(1, 10000)
        self.pw_count.setValue(10)
        self.pw_count.setStyleSheet(spinbox_style)
        count_row.addWidget(count_label)
        count_row.addWidget(self.pw_count)
        count_row.addStretch()
        gen_layout.addLayout(count_row)
        
        scroll_layout.addWidget(gen_group)
        
        # ===== Generateボタン（Generation Settingsのすぐ下）=====
        gen_btn_layout = QHBoxLayout()
        generate_btn = QPushButton("Generate")
        generate_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        generate_btn.clicked.connect(self._generate_passwords)
        generate_btn.setStyleSheet("""
            QPushButton { background-color: #27ae60; color: white; border: none;
                border-radius: 6px; padding: 12px 40px; font-size: 14px; font-weight: bold; }
            QPushButton:hover { background-color: #2ecc71; }
        """)
        gen_btn_layout.addWidget(generate_btn)
        gen_btn_layout.addStretch()
        scroll_layout.addLayout(gen_btn_layout)
        
        scroll_layout.addStretch()
        scroll.setWidget(scroll_content)
        layout.addWidget(scroll)
        
        return widget
    
    def _generate_passwords(self):
        """パスワードを生成"""
        import random
        import string
        from datetime import datetime
        
        try:
            # 文字種を取得
            chars = ""
            if self.chk_uppercase.isChecked():
                chars += string.ascii_uppercase
            if self.chk_lowercase.isChecked():
                chars += string.ascii_lowercase
            if self.chk_numbers.isChecked():
                chars += string.digits
            if self.chk_symbols.isChecked():
                chars += "!@#$%^&*()_+-=[]{}|;:,.<>?"
            
            if not chars:
                self.toast.show_toast("文字種を1つ以上選択してください", "error")
                return
            
            # 設定取得
            length = self.pw_length.value()
            count = self.pw_count.value()
            
            # パスワード生成
            passwords = []
            for _ in range(count):
                pw = ''.join(random.choice(chars) for _ in range(length))
                passwords.append(pw)
            
            # 出力ディレクトリ
            if getattr(sys, 'frozen', False):
                base_dir = os.path.dirname(sys.executable)
            else:
                base_dir = os.path.dirname(os.path.abspath(__file__))
            export_dir = os.path.join(base_dir, "_internal", "Export", "PasswordGenerator")
            
            os.makedirs(export_dir, exist_ok=True)
            
            # ファイル名（タイムスタンプ付き）
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = os.path.join(export_dir, f"passwords_{timestamp}.txt")
            
            with open(filepath, 'w', encoding='utf-8') as f:
                for pw in passwords:
                    f.write(pw + '\n')
            
            self.toast.show_toast(f"生成完了: {count}件", "success")
            
            # フォルダを開く
            if os.path.exists(export_dir):
                os.startfile(export_dir)
                
        except Exception as e:
            self.toast.show_toast(f"生成エラー: {e}", "error")
            import traceback
            traceback.print_exc()


class ProxyPage(QWidget):
    """プロキシ設定ページ"""
    
    def __init__(self):
        super().__init__()
        self.settings_dir = SETTINGS_DIR
        self.proxy_groups = []  # [{title, proxies: [], selected: bool}]
        self.setup_ui()
        self.load_settings()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)
        
        # ヘッダー
        header = QLabel("Proxies")
        header.setStyleSheet("""
            font-size: 24px;
            font-weight: bold;
            color: #ffffff;
            padding-bottom: 10px;
        """)
        layout.addWidget(header)
        
        # トースト通知
        self.toast = ToastNotification(self)
        
        # プロキシ有効化（スイッチボタン）
        enable_layout = QHBoxLayout()
        enable_label = QLabel("Use Proxy")
        enable_label.setStyleSheet("color: #e0e0e0; font-size: 14px;")
        enable_layout.addWidget(enable_label)
        
        # スイッチボタン
        self.proxy_switch = SwitchButton()
        self.proxy_switch.toggled.connect(self._save_settings)
        enable_layout.addWidget(self.proxy_switch)
        
        enable_layout.addStretch()
        layout.addLayout(enable_layout)
        
        # プロキシグループ追加
        add_group = QGroupBox("Add Proxy Group")
        add_group.setStyleSheet(self._group_style())
        add_layout = QVBoxLayout(add_group)
        add_layout.setContentsMargins(20, 25, 20, 20)
        add_layout.setSpacing(15)
        
        # タイトル入力
        title_layout = QHBoxLayout()
        title_label = QLabel("Title:")
        title_label.setStyleSheet("color: #b0b0b0; font-size: 13px;")
        title_label.setFixedWidth(80)
        title_layout.addWidget(title_label)
        self.group_title = QLineEdit()
        self.group_title.setPlaceholderText("e.g., US Proxies, Datacenter 1")
        self.group_title.setStyleSheet(self._input_style())
        title_layout.addWidget(self.group_title)
        add_layout.addLayout(title_layout)
        
        # プロキシ入力（複数行）
        proxy_label = QLabel("Proxies (one per line):")
        proxy_label.setStyleSheet("color: #b0b0b0; font-size: 13px;")
        add_layout.addWidget(proxy_label)
        
        self.proxy_input = QTextEdit()
        self.proxy_input.setPlaceholderText("ip:port or ip:port:user:pass\n192.168.1.1:8080\n10.0.0.1:3128:admin:password")
        self.proxy_input.setStyleSheet(self._textedit_style())
        self.proxy_input.setMinimumHeight(120)
        self.proxy_input.setMaximumHeight(150)
        add_layout.addWidget(self.proxy_input)
        
        # Saveボタン
        save_btn = QPushButton("Save")
        save_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        save_btn.clicked.connect(self._save_group)
        save_btn.setStyleSheet("""
            QPushButton { background-color: #27ae60; color: white; border: none;
                border-radius: 6px; padding: 10px 25px; font-size: 13px; font-weight: bold; }
            QPushButton:hover { background-color: #2ecc71; }
        """)
        add_layout.addWidget(save_btn, alignment=Qt.AlignmentFlag.AlignLeft)
        
        layout.addWidget(add_group)
        
        # 保存済みグループ一覧
        saved_group = QGroupBox("Saved Groups")
        saved_group.setStyleSheet(self._group_style())
        self.groups_layout = QVBoxLayout(saved_group)
        self.groups_layout.setContentsMargins(20, 25, 20, 20)
        self.groups_layout.setSpacing(10)
        layout.addWidget(saved_group)
        
        layout.addStretch()
    
    def _save_group(self):
        """プロキシグループを保存"""
        title = self.group_title.text().strip()
        proxy_text = self.proxy_input.toPlainText().strip()
        
        if not title:
            self.toast.show_toast("Please enter a title", "warning", 3000)
            return
        
        if not proxy_text:
            self.toast.show_toast("Please enter at least one proxy", "warning", 3000)
            return
        
        # プロキシをパース
        proxies = []
        for line in proxy_text.split('\n'):
            line = line.strip()
            if line:
                proxies.append(line)
        
        if not proxies:
            self.toast.show_toast("No valid proxies found", "warning", 3000)
            return
        
        group = {
            "title": title,
            "proxies": proxies,
            "selected": len(self.proxy_groups) == 0  # 最初のグループは自動選択
        }
        self.proxy_groups.append(group)
        self._save_settings()
        self._refresh_groups()
        
        # 入力クリア
        self.group_title.clear()
        self.proxy_input.clear()
        
        self.toast.show_toast(f"Group '{title}' saved ({len(proxies)} proxies)", "success", 2000)
    
    def _refresh_groups(self):
        """グループ一覧を更新"""
        # 既存のウィジェットを削除
        while self.groups_layout.count():
            item = self.groups_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        
        if not self.proxy_groups:
            empty_label = QLabel("No proxy groups saved")
            empty_label.setStyleSheet("color: #808080; font-size: 13px; padding: 20px;")
            empty_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self.groups_layout.addWidget(empty_label)
            return
        
        for i, group in enumerate(self.proxy_groups):
            row = QWidget()
            row.setStyleSheet("background-color: #2a2a3a; border-radius: 8px;")
            row_layout = QHBoxLayout(row)
            row_layout.setContentsMargins(15, 10, 15, 10)
            row_layout.setSpacing(15)
            
            # ラジオボタン（Fetchと同じスタイル）
            radio = QPushButton("●" if group.get("selected") else "○")
            radio.setFixedSize(24, 24)
            radio.setCursor(Qt.CursorShape.PointingHandCursor)
            radio.setStyleSheet(f"background: transparent; color: {'#4a90d9' if group.get('selected') else '#606060'}; border: none; font-size: 18px;")
            radio.clicked.connect(lambda _, idx=i: self._select_group(idx))
            row_layout.addWidget(radio)
            
            # タイトル
            title_label = QLabel(group["title"])
            title_label.setStyleSheet("color: #ffffff; font-weight: bold; background: transparent;")
            title_label.setFixedWidth(150)
            row_layout.addWidget(title_label)
            
            # プロキシ数（タイトルと同じフォント）
            count_label = QLabel(f"{len(group['proxies'])} Proxies")
            count_label.setStyleSheet("color: #b0b0b0; background: transparent;")
            row_layout.addWidget(count_label)
            
            row_layout.addStretch()
            
            # チェックボタン
            check_btn = QPushButton()
            check_btn.setFixedSize(32, 32)
            check_btn.setCursor(Qt.CursorShape.PointingHandCursor)
            check_icon = get_icon_from_base64("speedometer")
            if not check_icon.isNull():
                check_btn.setIcon(check_icon)
                check_btn.setIconSize(QSize(16, 16))
            else:
                check_btn.setText("🔍")
            check_btn.setToolTip("Check proxies")
            check_btn.setStyleSheet("background: transparent; border: none; font-size: 16px;")
            check_btn.clicked.connect(lambda _, idx=i: self._check_group(idx))
            row_layout.addWidget(check_btn)
            
            # 編集ボタン
            edit_btn = QPushButton()
            edit_btn.setFixedSize(32, 32)
            edit_btn.setCursor(Qt.CursorShape.PointingHandCursor)
            edit_icon = get_icon_from_base64("edit")
            if not edit_icon.isNull():
                edit_btn.setIcon(edit_icon)
                edit_btn.setIconSize(QSize(16, 16))
            else:
                edit_btn.setText("✏️")
            edit_btn.setStyleSheet("background: transparent; border: none; font-size: 16px;")
            edit_btn.clicked.connect(lambda _, idx=i: self._edit_group(idx))
            row_layout.addWidget(edit_btn)
            
            # 削除ボタン
            delete_btn = QPushButton()
            delete_btn.setFixedSize(32, 32)
            delete_btn.setCursor(Qt.CursorShape.PointingHandCursor)
            delete_icon = get_icon_from_base64("delete")
            if not delete_icon.isNull():
                delete_btn.setIcon(delete_icon)
                delete_btn.setIconSize(QSize(16, 16))
            else:
                delete_btn.setText("🗑️")
            delete_btn.setStyleSheet("background: transparent; border: none; font-size: 16px;")
            delete_btn.clicked.connect(lambda _, idx=i: self._delete_group(idx))
            row_layout.addWidget(delete_btn)
            
            self.groups_layout.addWidget(row)
    
    def _select_group(self, index):
        """グループを選択"""
        for i, group in enumerate(self.proxy_groups):
            group["selected"] = (i == index)
        self._save_settings()
        self._refresh_groups()
    
    def _on_radio_toggled(self, index, checked):
        """ラジオボタンの選択変更（互換性のため残す）"""
        if checked:
            self._select_group(index)
    
    def _edit_group(self, index):
        """グループを編集"""
        group = self.proxy_groups[index]
        self.group_title.setText(group["title"])
        self.proxy_input.setPlainText('\n'.join(group["proxies"]))
        
        # 編集モードとして古いグループを削除
        del self.proxy_groups[index]
        self._save_settings()
        self._refresh_groups()
        
        self.toast.show_toast("Edit mode - modify and save", "info", 2000)
    
    def _delete_group(self, index):
        """グループを削除"""
        group = self.proxy_groups[index]
        title = group["title"]
        
        if QMessageBox.question(self, "Confirm", f"Delete '{title}'?") == QMessageBox.StandardButton.Yes:
            del self.proxy_groups[index]
            # 選択されていたグループが削除された場合、最初のグループを選択
            if self.proxy_groups and not any(g.get("selected") for g in self.proxy_groups):
                self.proxy_groups[0]["selected"] = True
            self._save_settings()
            self._refresh_groups()
            self.toast.show_toast(f"'{title}' deleted", "success", 2000)
    
    def _check_group(self, index):
        """グループのプロキシをチェック"""
        group = self.proxy_groups[index]
        title = group["title"]
        proxies = group.get("proxies", [])
        
        if not proxies:
            self.toast.show_toast("No proxies in this group", "error")
            return
        
        # ProxyCheckerダイアログを表示
        dialog = ProxyCheckerDialog(title, proxies, self)
        dialog.show()
    
    def _save_settings(self):
        """設定を保存"""
        try:
            data = {
                "enabled": self.proxy_switch.isChecked(),
                "groups": self.proxy_groups
            }
            save_path = self.settings_dir / "proxy_settings.json"
            with open(save_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Failed to save proxy settings: {e}")
    
    def load_settings(self):
        """設定を読み込み"""
        try:
            path = self.settings_dir / "proxy_settings.json"
            if path.exists():
                with open(path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                # シグナルをブロックして保存が発火しないようにする
                self.proxy_switch.blockSignals(True)
                self.proxy_switch.setChecked(data.get("enabled", False))
                self.proxy_switch._update_style()  # 見た目を手動で更新
                self.proxy_switch.blockSignals(False)
                self.proxy_groups = data.get("groups", [])
                self._refresh_groups()
        except Exception as e:
            print(f"Failed to load proxy settings: {e}")
    
    def get_random_proxy(self):
        """選択されたグループからランダムにプロキシを取得"""
        if not self.proxy_switch.isChecked():
            return None
        
        selected_group = next((g for g in self.proxy_groups if g.get("selected")), None)
        if not selected_group or not selected_group.get("proxies"):
            return None
        
        import random
        return random.choice(selected_group["proxies"])
    
    def _group_style(self):
        return """
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #b0b0b0;
                border: 1px solid #404050;
                border-radius: 10px;
                margin-top: 10px;
                padding-top: 15px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 5px;
            }
        """
    
    def _input_style(self):
        return """
            QLineEdit {
                background-color: #2a2a3a;
                border: 1px solid #404050;
                border-radius: 6px;
                padding: 10px 15px;
                color: #ffffff;
                font-size: 13px;
            }
            QLineEdit:focus {
                border-color: #4a90d9;
            }
        """
    
    def _textedit_style(self):
        return """
            QTextEdit {
                background-color: #2a2a3a;
                border: 1px solid #404050;
                border-radius: 6px;
                padding: 10px 15px;
                color: #ffffff;
                font-size: 13px;
            }
            QTextEdit:focus {
                border-color: #4a90d9;
            }
        """


class UpdateDialog(QDialog):
    """アップデート確認・実行ダイアログ"""
    
    def __init__(self, latest_version, changelog="", parent=None):
        super().__init__(parent)
        self.latest_version = latest_version
        self.setWindowTitle("Update Available")
        self.setFixedSize(420, 220)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self._drag_pos = None
        self._setup_ui(changelog)
    
    def _setup_ui(self, changelog):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        
        container = QWidget()
        container.setObjectName("update_container")
        container.setStyleSheet("""
            #update_container {
                background-color: #252535;
                border-radius: 12px;
                border: 1px solid #303040;
            }
        """)
        outer.addWidget(container)
        
        layout = QVBoxLayout(container)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(14)
        
        # タイトル
        # vプレフィックスを除去して表示
        display_version = self.latest_version.lstrip('v')
        title = QLabel(f"🔄 アップデート v{display_version}")
        title.setStyleSheet("font-size: 17px; font-weight: bold; color: #ffffff; background: transparent;")
        layout.addWidget(title)
        
        # 説明
        desc = QLabel("新しいバージョンが利用可能です。\n今すぐアップデートしますか？")
        desc.setStyleSheet("color: #b0b0b0; font-size: 13px; background: transparent;")
        desc.setWordWrap(True)
        layout.addWidget(desc)
        
        # プログレスバー（初期は非表示）
        self.progress = QProgressBar()
        self.progress.setStyleSheet("""
            QProgressBar {
                border: 1px solid #303040;
                border-radius: 6px;
                background-color: #1e1e2e;
                height: 18px;
                text-align: center;
                color: #ffffff;
                font-size: 11px;
            }
            QProgressBar::chunk {
                background-color: #3498db;
                border-radius: 5px;
            }
        """)
        self.progress.hide()
        layout.addWidget(self.progress)
        
        # ステータスラベル（初期は非表示）
        self.status_label = QLabel("")
        self.status_label.setStyleSheet("color: #b0b0b0; font-size: 12px; background: transparent;")
        self.status_label.hide()
        layout.addWidget(self.status_label)
        
        layout.addStretch()
        
        # ボタン
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        self.later_btn = QPushButton("あとで")
        self.later_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.later_btn.setStyleSheet("""
            QPushButton {
                background-color: #404050; color: #b0b0b0;
                border: none; border-radius: 8px;
                padding: 9px 24px; font-size: 13px;
            }
            QPushButton:hover { background-color: #505060; }
        """)
        self.later_btn.clicked.connect(self.reject)
        btn_layout.addWidget(self.later_btn)
        
        self.update_btn = QPushButton("アップデート")
        self.update_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.update_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db; color: white;
                border: none; border-radius: 8px;
                padding: 9px 24px; font-size: 13px; font-weight: bold;
            }
            QPushButton:hover { background-color: #2980b9; }
            QPushButton:disabled { background-color: #404050; color: #808080; }
        """)
        self.update_btn.clicked.connect(self._on_update)
        btn_layout.addWidget(self.update_btn)
        layout.addLayout(btn_layout)
    
    def _on_update(self):
        """アップデート実行"""
        self.update_btn.setEnabled(False)
        self.later_btn.setEnabled(False)
        self.progress.show()
        self.status_label.show()
        self.status_label.setText("ダウンロード中...")
        self.progress.setValue(0)
        QApplication.processEvents()
        
        # ダウンロード
        needs, latest, url, _ = check_for_update()
        if not url:
            self.status_label.setText("⚠ ダウンロードURLが見つかりません")
            self.update_btn.setEnabled(True)
            self.later_btn.setEnabled(True)
            return
        
        def dl_progress(percent):
            self.progress.setValue(percent)
            self.status_label.setText(f"ダウンロード中... {percent}%")
            QApplication.processEvents()
        
        zip_path = download_update(url, dl_progress)
        
        if not zip_path:
            self.status_label.setText("⚠ ダウンロードに失敗しました")
            self.update_btn.setEnabled(True)
            self.later_btn.setEnabled(True)
            return
        
        # 適用
        self.status_label.setText("更新を適用中...")
        self.progress.setValue(0)
        QApplication.processEvents()
        
        def apply_progress(percent):
            self.progress.setValue(percent)
            QApplication.processEvents()
        
        success = apply_update(zip_path, apply_progress)
        
        if success:
            save_version(latest)
            self.status_label.setText("✅ 更新完了！再起動します...")
            self.progress.setValue(100)
            QApplication.processEvents()
            
            import time
            time.sleep(1)
            restart_app()
        else:
            self.status_label.setText("⚠ 更新に失敗しました")
            self.update_btn.setEnabled(True)
            self.later_btn.setEnabled(True)
    
    # ドラッグ移動
    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            self._drag_pos = event.globalPosition().toPoint() - self.frameGeometry().topLeft()
    
    def mouseMoveEvent(self, event):
        if event.buttons() == Qt.MouseButton.LeftButton and self._drag_pos:
            self.move(event.globalPosition().toPoint() - self._drag_pos)
    
    def mouseReleaseEvent(self, event):
        self._drag_pos = None


class ProxyCheckerDialog(QDialog):
    """プロキシチェッカーダイアログ"""
    
    TEST_SITES = {
        "Google": "http://www.google.com",
        "Amazon": "https://www.amazon.co.jp/",
        "Custom": ""
    }
    
    def __init__(self, group_name, proxies, parent=None):
        super().__init__(parent)
        self.group_name = group_name
        self.proxies = proxies
        self.check_threads = []
        self.results = {}
        self.is_checking = False
        self.proxy_checkboxes = []
        
        self.setWindowTitle(f"Proxy Checker - {group_name}")
        self.setMinimumSize(800, 550)
        self.setWindowFlags(Qt.WindowType.Window)
        self._setup_ui()
    
    def _setup_ui(self):
        self.setStyleSheet("""
            QDialog {
                background-color: #1e1e2e;
            }
            QTableWidget {
                background-color: #1e1e2e;
                border: none;
                color: #ffffff;
                outline: none;
                selection-background-color: transparent;
                gridline-color: transparent;
            }
            QTableWidget::item {
                padding: 0px;
                margin: 0px;
                border: none;
                outline: none;
            }
            QTableWidget::item:selected {
                background-color: transparent;
                border: none;
            }
            QHeaderView::section {
                background-color: #1e1e2e;
                color: #808080;
                padding: 10px;
                border: none;
                border-bottom: 1px solid #303040;
                font-weight: bold;
                font-size: 12px;
            }
            QScrollBar:vertical {
                background: #252535;
                width: 10px;
                border-radius: 5px;
            }
            QScrollBar::handle:vertical {
                background: #404050;
                border-radius: 5px;
                min-height: 30px;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                height: 0;
            }
            QComboBox {
                background-color: #3a3a4a;
                border: 1px solid #404050;
                border-radius: 4px;
                padding: 6px 12px;
                color: #ffffff;
                font-size: 12px;
                min-width: 120px;
            }
            QComboBox:hover { border-color: #4a90d9; }
            QComboBox::drop-down {
                border: none;
                width: 20px;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 6px solid #b0b0b0;
                margin-right: 8px;
            }
            QComboBox QAbstractItemView {
                background-color: #2a2a3a;
                border: 1px solid #404050;
                color: #ffffff;
                selection-background-color: #3a3a4a;
            }
            QLineEdit {
                background-color: #3a3a4a;
                border: 1px solid #404050;
                border-radius: 4px;
                padding: 6px 12px;
                color: #ffffff;
                font-size: 12px;
            }
            QLineEdit:focus { border-color: #4a90d9; }
            QCheckBox {
                color: #ffffff;
                spacing: 6px;
                background: transparent;
            }
            QCheckBox::indicator {
                width: 14px;
                height: 14px;
                border-radius: 3px;
                border: 2px solid #404050;
                background-color: #252535;
            }
            QCheckBox::indicator:checked {
                background-color: #4a90d9;
                border-color: #4a90d9;
            }
            QLabel {
                background: transparent;
            }
        """)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        # ヘッダー
        header_layout = QHBoxLayout()
        
        # タイトルアイコン
        title_icon_label = QLabel()
        speedometer_icon = get_icon_from_base64("speedometer")
        if not speedometer_icon.isNull():
            title_icon_label.setPixmap(speedometer_icon.pixmap(20, 20))
        else:
            title_icon_label.setText("🔍")
        title_icon_label.setStyleSheet("background: transparent;")
        header_layout.addWidget(title_icon_label)
        
        title = QLabel(f"{self.group_name} ({len(self.proxies)} proxies)")
        title.setStyleSheet("font-size: 18px; font-weight: bold; color: #ffffff;")
        header_layout.addWidget(title)
        
        header_layout.addStretch()
        
        # Test Site選択
        site_label = QLabel("Test site:")
        site_label.setStyleSheet("color: #b0b0b0; font-size: 12px; background: transparent;")
        header_layout.addWidget(site_label)
        
        self.site_combo = QComboBox()
        self.site_combo.addItems(["Google", "Amazon", "Custom"])
        self.site_combo.currentTextChanged.connect(self._on_site_changed)
        header_layout.addWidget(self.site_combo)
        
        # Custom URL入力
        self.custom_url = QLineEdit()
        self.custom_url.setPlaceholderText("https://example.com")
        self.custom_url.setFixedWidth(200)
        self.custom_url.setVisible(False)
        header_layout.addWidget(self.custom_url)
        
        layout.addLayout(header_layout)
        
        # テーブル
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setShowGrid(False)
        self.table.setMouseTracking(True)
        self.table.viewport().setMouseTracking(True)
        
        # マスターチェックボックス
        self.master_checkbox = QCheckBox()
        self.master_checkbox.setChecked(True)
        self.master_checkbox.setStyleSheet("""
            QCheckBox {
                background: transparent;
            }
            QCheckBox::indicator {
                width: 14px;
                height: 14px;
                border-radius: 3px;
                border: 2px solid #404050;
                background-color: #252535;
            }
            QCheckBox::indicator:checked {
                background-color: #4a90d9;
                border-color: #4a90d9;
            }
        """)
        self.master_checkbox.stateChanged.connect(self._on_master_checkbox_changed)
        
        # ヘッダーラベル設定
        self.table.setHorizontalHeaderLabels(["", "Proxy", "Status", "Speed", "Location"])
        
        # ヘッダーの最初の列にマスターチェックボックスを配置
        header_checkbox_widget = QWidget()
        header_checkbox_layout = QHBoxLayout(header_checkbox_widget)
        header_checkbox_layout.setContentsMargins(0, 0, 0, 0)
        header_checkbox_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        header_checkbox_layout.addWidget(self.master_checkbox)
        self.table.horizontalHeader().setMinimumSectionSize(50)
        
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(0, 50)
        self.table.setColumnWidth(2, 100)
        self.table.setColumnWidth(3, 100)
        self.table.setColumnWidth(4, 150)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QTableWidget.SelectionMode.NoSelection)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.verticalHeader().setVisible(False)
        
        # 行ホバー用の変数
        self.hovered_row = -1
        
        # チェックボックスリストを初期化
        self.proxy_checkboxes = [None] * len(self.proxies)
        self.proxy_checked = [True] * len(self.proxies)
        
        # 行数を設定
        self.table.setRowCount(len(self.proxies))
        for i in range(len(self.proxies)):
            self.table.setRowHeight(i, 45)
        
        # バッチ処理用の変数
        self._batch_index = 0
        self._batch_size = 50  # 一度に作成する行数
        
        # テーブルのヘッダーを非表示にし、カスタムヘッダーを作成
        self.table.horizontalHeader().setVisible(False)
        
        # カスタムヘッダー行
        header_row = QHBoxLayout()
        header_row.setContentsMargins(0, 0, 0, 5)
        header_row.setSpacing(0)
        
        # マスターチェックボックス（50px幅に合わせる）
        master_cb_container = QWidget()
        master_cb_container.setFixedWidth(50)
        master_cb_container.setStyleSheet("background: transparent;")
        master_cb_layout = QHBoxLayout(master_cb_container)
        master_cb_layout.setContentsMargins(0, 0, 0, 0)
        master_cb_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        master_cb_layout.addWidget(self.master_checkbox)
        header_row.addWidget(master_cb_container)
        
        # Proxyラベル（ストレッチ）
        proxy_label = QLabel("Proxy")
        proxy_label.setStyleSheet("color: #808080; font-weight: bold; font-size: 12px; padding-left: 8px;")
        header_row.addWidget(proxy_label, 1)
        
        # Statusラベル（100px）
        status_label = QLabel("Status")
        status_label.setFixedWidth(100)
        status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        status_label.setStyleSheet("color: #808080; font-weight: bold; font-size: 12px;")
        header_row.addWidget(status_label)
        
        # Speedラベル（100px）
        speed_label = QLabel("Speed")
        speed_label.setFixedWidth(100)
        speed_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        speed_label.setStyleSheet("color: #808080; font-weight: bold; font-size: 12px;")
        header_row.addWidget(speed_label)
        
        # Locationラベル（150px）
        location_label = QLabel("Location")
        location_label.setFixedWidth(150)
        location_label.setStyleSheet("color: #808080; font-weight: bold; font-size: 12px; padding-left: 8px;")
        header_row.addWidget(location_label)
        
        layout.addLayout(header_row)
        
        layout.addWidget(self.table)
        
        # プログレスバー（テーブルの下）
        self.progress = QProgressBar()
        self.progress.setFixedHeight(12)
        self.progress.setStyleSheet("""
            QProgressBar {
                border: 1px solid #303040;
                border-radius: 4px;
                background-color: #252535;
                text-align: center;
                color: #ffffff;
                font-size: 9px;
            }
            QProgressBar::chunk {
                background-color: #27ae60;
                border-radius: 3px;
            }
        """)
        self.progress.setVisible(False)
        layout.addWidget(self.progress)
        
        # フッター（統計とStart Checkボタン）
        footer_layout = QHBoxLayout()
        
        self.stats_label = QLabel("Select proxies and click 'Start Check' to begin")
        self.stats_label.setStyleSheet("color: #b0b0b0; font-size: 12px;")
        footer_layout.addWidget(self.stats_label)
        
        footer_layout.addStretch()
        
        # チェックボタン（右下）
        self.check_btn = QPushButton("▶ Start Check")
        self.check_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.check_btn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60; color: #ffffff;
                border: none; border-radius: 8px;
                padding: 10px 24px; font-size: 13px; font-weight: bold;
            }
            QPushButton:hover { background-color: #2ecc71; }
            QPushButton:disabled { background-color: #404050; color: #808080; }
        """)
        self.check_btn.clicked.connect(self._start_check)
        footer_layout.addWidget(self.check_btn)
        
        layout.addLayout(footer_layout)
        
        # ウィンドウ表示後にバッチ処理で行を作成
        from PySide6.QtCore import QTimer
        self._batch_timer = QTimer()
        self._batch_timer.timeout.connect(self._create_batch_rows)
        self._batch_timer.start(0)  # できるだけ速く
    
    def _create_batch_rows(self):
        """バッチで行を作成"""
        end_idx = min(self._batch_index + self._batch_size, len(self.proxies))
        
        for i in range(self._batch_index, end_idx):
            self._create_row(i)
        
        self._batch_index = end_idx
        
        # 全て作成完了したらタイマーを停止
        if self._batch_index >= len(self.proxies):
            self._batch_timer.stop()
    
    def _create_row(self, row_idx):
        """単一行を作成"""
        if self.proxy_checkboxes[row_idx] is not None:
            return  # 既に作成済み
        
        proxy = self.proxies[row_idx]
        
        # チェックボックス（列0）
        cb = QCheckBox()
        cb.setChecked(self.proxy_checked[row_idx])
        cb.stateChanged.connect(lambda state, idx=row_idx: self._on_proxy_checkbox_changed(idx, state))
        cb_widget = QWidget()
        cb_widget.setStyleSheet("background-color: transparent;")
        cb_layout = QHBoxLayout(cb_widget)
        cb_layout.setContentsMargins(0, 0, 0, 0)
        cb_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        cb_layout.addWidget(cb)
        cb_widget.setMouseTracking(True)
        cb_widget.enterEvent = lambda event, r=row_idx: self._update_row_hover(r)
        cb.setMouseTracking(True)
        cb.enterEvent = lambda event, r=row_idx: self._update_row_hover(r)
        self.table.setCellWidget(row_idx, 0, cb_widget)
        
        self.proxy_checkboxes[row_idx] = cb
        
        # キャッシュされた結果を確認
        cached = self.results.get(row_idx)
        
        # テキストセル作成用ヘルパー
        def create_text_cell(text, color="#ffffff", center=False):
            widget = QWidget()
            widget.setStyleSheet("background-color: transparent;")
            wlayout = QHBoxLayout(widget)
            wlayout.setContentsMargins(8, 0, 8, 0)
            wlayout.setSpacing(0)
            label = QLabel(text)
            label.setStyleSheet(f"color: {color}; background-color: transparent;")
            label.setObjectName("cell_label")
            wlayout.addWidget(label)
            if center:
                wlayout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            else:
                wlayout.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
            widget.setMouseTracking(True)
            widget.enterEvent = lambda event, r=row_idx: self._update_row_hover(r)
            return widget
        
        # Proxy（列1）
        self.table.setCellWidget(row_idx, 1, create_text_cell(proxy))
        
        # Status（列2）- 中央揃え
        if cached:
            self.table.setCellWidget(row_idx, 2, create_text_cell(cached["status_text"], cached["status_color"], center=True))
        else:
            self.table.setCellWidget(row_idx, 2, create_text_cell("Waiting", "#808080", center=True))
        
        # Speed（列3）- 中央揃え
        if cached:
            self.table.setCellWidget(row_idx, 3, create_text_cell(cached["speed_text"], cached["speed_color"], center=True))
        else:
            self.table.setCellWidget(row_idx, 3, create_text_cell("-", "#808080", center=True))
        
        # Location（列4）
        if cached:
            self.table.setCellWidget(row_idx, 4, create_text_cell(cached["location_text"], cached["location_color"]))
        else:
            self.table.setCellWidget(row_idx, 4, create_text_cell("-", "#808080"))
    
    def _on_proxy_checkbox_changed(self, idx, state):
        """プロキシのチェックボックス変更時"""
        self.proxy_checked[idx] = (state == Qt.CheckState.Checked.value)
    
    def _update_row_hover(self, new_row):
        """行のホバー状態を更新"""
        if new_row == self.hovered_row:
            return
        
        # 前のホバー行の背景色をリセット
        if 0 <= self.hovered_row < self.table.rowCount():
            self._set_row_background(self.hovered_row, "transparent")
        
        # 新しいホバー行の背景色を設定
        if 0 <= new_row < self.table.rowCount():
            self._set_row_background(new_row, "#2a2a3a")
        
        self.hovered_row = new_row
    
    def _set_row_background(self, row, color):
        """行の背景色を設定（全てウィジェットベース）"""
        col_count = self.table.columnCount()
        for col in range(col_count):
            widget = self.table.cellWidget(row, col)
            if widget:
                if color == "transparent":
                    widget.setStyleSheet("background-color: transparent; border-radius: 0px;")
                else:
                    # 最初のセルは左側に角丸、最後のセルは右側に角丸、中間は角丸なし
                    if col == 0:
                        widget.setStyleSheet(f"background-color: {color}; border-top-left-radius: 6px; border-bottom-left-radius: 6px; border-top-right-radius: 0px; border-bottom-right-radius: 0px;")
                    elif col == col_count - 1:
                        widget.setStyleSheet(f"background-color: {color}; border-top-left-radius: 0px; border-bottom-left-radius: 0px; border-top-right-radius: 6px; border-bottom-right-radius: 6px;")
                    else:
                        widget.setStyleSheet(f"background-color: {color}; border-radius: 0px;")
    
    def _on_site_changed(self, site):
        """テストサイト変更時"""
        self.custom_url.setVisible(site == "Custom")
    
    def _on_master_checkbox_changed(self, state):
        """マスターチェックボックス変更時"""
        checked = state == Qt.CheckState.Checked.value
        # 全プロキシのチェック状態を更新
        for i in range(len(self.proxies)):
            self.proxy_checked[i] = checked
        # 作成済みのチェックボックスを更新
        for cb in self.proxy_checkboxes:
            if cb is not None:
                cb.setChecked(checked)
    
    def _get_test_url(self):
        """テストURLを取得"""
        site = self.site_combo.currentText()
        if site == "Custom":
            url = self.custom_url.text().strip()
            if not url:
                return None
            if not url.startswith("http"):
                url = "https://" + url
            return url
        return self.TEST_SITES.get(site, "http://www.google.com")
    
    def _start_check(self):
        """プロキシチェックを開始"""
        if self.is_checking:
            return
        
        # テストURL確認
        test_url = self._get_test_url()
        if not test_url:
            return
        
        # 選択されたプロキシを取得（proxy_checkedリストを使用）
        selected_proxies = []
        selected_indices = []
        for i in range(len(self.proxies)):
            if self.proxy_checked[i]:
                selected_proxies.append(self.proxies[i])
                selected_indices.append(i)
        
        if not selected_proxies:
            self.stats_label.setText("⚠ No proxies selected")
            self.stats_label.setStyleSheet("color: #e74c3c; font-size: 12px;")
            return
        
        self.is_checking = True
        self.check_btn.setEnabled(False)
        self.check_btn.setText("Checking...")
        self.progress.setVisible(True)
        self.progress.setMaximum(len(selected_proxies))
        self.progress.setValue(0)
        self.results = {}
        self.check_threads = []
        
        # ステータスをリセット（ウィジェットベース）- 行が作成済みの場合のみ
        for i in selected_indices:
            if self.proxy_checkboxes[i] is not None:
                self._update_cell_text(i, 2, "Checking...", "#f39c12")
                self._update_cell_text(i, 3, "-", "#808080")
                self._update_cell_text(i, 4, "-", "#808080")
        
        # 各プロキシをチェック（スレッドで実行）
        self.completed_count = 0
        self.total_to_check = len(selected_proxies)
        
        for i, proxy in zip(selected_indices, selected_proxies):
            thread = ProxyCheckThread(proxy, i, test_url)
            thread.result_ready.connect(self._on_result)
            thread.start()
            self.check_threads.append(thread)
    
    def _update_cell_text(self, row, col, text, color="#ffffff"):
        """セルのテキストを更新（ウィジェットベース）"""
        widget = self.table.cellWidget(row, col)
        if widget:
            label = widget.findChild(QLabel, "cell_label")
            if label:
                label.setText(text)
                label.setStyleSheet(f"color: {color}; background-color: transparent;")
    
    def _on_result(self, index, proxy, success, speed_ms, location, error):
        """チェック結果を受信"""
        self.completed_count += 1
        self.progress.setValue(self.completed_count)
        
        # テーブルを更新（ウィジェットベース）
        if success:
            status_text = "✓ OK"
            status_color = "#27ae60"
            
            if speed_ms < 500:
                speed_text = f"{speed_ms}ms"
                speed_color = "#27ae60"  # 緑（高速）
            elif speed_ms < 1500:
                speed_text = f"{speed_ms}ms"
                speed_color = "#f39c12"  # 黄（普通）
            else:
                speed_text = f"{speed_ms}ms"
                speed_color = "#e74c3c"  # 赤（遅い）
            
            location_text = location if location else "OK"
            location_color = "#808080"
        else:
            status_text = "✗ Failed"
            status_color = "#e74c3c"
            speed_text = "-"
            speed_color = "#808080"
            location_text = error[:25] if error else "Connection failed"
            location_color = "#e74c3c"
        
        # 結果をキャッシュ（行が後から作成されたときに使用）
        self.results[index] = {
            "proxy": proxy, 
            "success": success, 
            "speed": speed_ms, 
            "location": location,
            "status_text": status_text,
            "status_color": status_color,
            "speed_text": speed_text,
            "speed_color": speed_color,
            "location_text": location_text,
            "location_color": location_color
        }
        
        # 行が作成済みならUIを更新
        if self.proxy_checkboxes[index] is not None:
            self._update_cell_text(index, 2, status_text, status_color)
            self._update_cell_text(index, 3, speed_text, speed_color)
            self._update_cell_text(index, 4, location_text, location_color)
        
        # 全て完了したら統計を表示
        if self.completed_count >= self.total_to_check:
            self._show_stats()
    
    def _show_stats(self):
        """統計を表示"""
        self.is_checking = False
        self.check_btn.setEnabled(True)
        self.check_btn.setText("▶ Recheck")
        
        working = sum(1 for r in self.results.values() if r.get("success", False))
        failed = len(self.results) - working
        
        speeds = [r["speed"] for r in self.results.values() if r.get("success", False) and r.get("speed", 0) > 0]
        avg_speed = int(sum(speeds) / len(speeds)) if speeds else 0
        
        self.stats_label.setText(
            f"✓ Working: {working}  |  ✗ Failed: {failed}  |  ⚡ Avg Speed: {avg_speed}ms"
        )
        self.stats_label.setStyleSheet("color: #ffffff; font-size: 13px; font-weight: bold;")
    
    def closeEvent(self, event):
        """ウィンドウを閉じる時にスレッドを停止"""
        # バッチタイマーを停止
        if hasattr(self, '_batch_timer'):
            self._batch_timer.stop()
        for thread in self.check_threads:
            if thread.isRunning():
                thread.terminate()
                thread.wait()
        event.accept()


class ProxyCheckThread(QThread):
    """プロキシをチェックするスレッド"""
    result_ready = Signal(int, str, bool, int, str, str)  # index, proxy, success, speed_ms, location, error
    
    def __init__(self, proxy, index, test_url="http://www.google.com"):
        super().__init__()
        self.proxy = proxy
        self.index = index
        self.test_url = test_url
    
    def run(self):
        import time
        
        try:
            # requestsを使用（より信頼性が高い）
            import requests
            from requests.auth import HTTPProxyAuth
            
            proxy_str = self.proxy.strip()
            
            # プロキシ形式をパース
            proxy_url = None
            auth = None
            
            if '@' in proxy_str:
                # user:pass@host:port 形式
                auth_part, host_part = proxy_str.rsplit('@', 1)
                if ':' in auth_part:
                    user, password = auth_part.split(':', 1)
                    proxy_url = f"http://{host_part}"
                    auth = HTTPProxyAuth(user, password)
                else:
                    proxy_url = f"http://{proxy_str}"
            elif proxy_str.count(':') >= 3:
                # host:port:user:pass 形式
                parts = proxy_str.split(':')
                host = parts[0]
                port = parts[1]
                user = parts[2]
                password = ':'.join(parts[3:])
                proxy_url = f"http://{host}:{port}"
                auth = HTTPProxyAuth(user, password)
            else:
                # host:port 形式
                proxy_url = f"http://{proxy_str}"
            
            proxies = {
                'http': proxy_url,
                'https': proxy_url
            }
            
            # タイムアウト設定
            start_time = time.time()
            
            # テストリクエスト
            response = requests.get(
                self.test_url,
                proxies=proxies,
                auth=auth,
                timeout=15,
                headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
            )
            
            elapsed_ms = int((time.time() - start_time) * 1000)
            
            if response.status_code == 200:
                # ロケーション情報を取得（オプション）
                location = ""
                try:
                    loc_response = requests.get(
                        "http://ip-api.com/json/?fields=country,city",
                        proxies=proxies,
                        auth=auth,
                        timeout=5
                    )
                    if loc_response.status_code == 200:
                        import json
                        info = loc_response.json()
                        city = info.get('city', '')
                        country = info.get('country', '')
                        if city or country:
                            location = f"{city}, {country}".strip(", ")
                except:
                    pass
                
                self.result_ready.emit(self.index, self.proxy, True, elapsed_ms, location, "")
            else:
                self.result_ready.emit(self.index, self.proxy, False, 0, "", f"HTTP {response.status_code}")
                
        except requests.exceptions.ProxyError as e:
            self.result_ready.emit(self.index, self.proxy, False, 0, "", "Proxy error")
        except requests.exceptions.ConnectTimeout:
            self.result_ready.emit(self.index, self.proxy, False, 0, "", "Connection timeout")
        except requests.exceptions.ReadTimeout:
            self.result_ready.emit(self.index, self.proxy, False, 0, "", "Read timeout")
        except requests.exceptions.ConnectionError:
            self.result_ready.emit(self.index, self.proxy, False, 0, "", "Connection failed")
        except ImportError:
            # requestsがない場合はurllibにフォールバック
            self._fallback_urllib()
        except Exception as e:
            error_msg = str(e)[:30]
            self.result_ready.emit(self.index, self.proxy, False, 0, "", error_msg)
    
    def _fallback_urllib(self):
        """requestsがない場合のフォールバック"""
        import urllib.request
        import time
        
        try:
            proxy_str = self.proxy.strip()
            
            if '@' in proxy_str:
                auth_part, host_part = proxy_str.rsplit('@', 1)
                proxy_url = f"http://{auth_part}@{host_part}"
            elif proxy_str.count(':') >= 3:
                parts = proxy_str.split(':')
                host = parts[0]
                port = parts[1]
                user = parts[2]
                password = ':'.join(parts[3:])
                proxy_url = f"http://{user}:{password}@{host}:{port}"
            else:
                proxy_url = f"http://{proxy_str}"
            
            proxy_handler = urllib.request.ProxyHandler({
                'http': proxy_url,
                'https': proxy_url
            })
            opener = urllib.request.build_opener(proxy_handler)
            
            start_time = time.time()
            
            req = urllib.request.Request(
                self.test_url,
                headers={"User-Agent": "Mozilla/5.0"}
            )
            
            with opener.open(req, timeout=15) as response:
                elapsed_ms = int((time.time() - start_time) * 1000)
                if response.status == 200:
                    self.result_ready.emit(self.index, self.proxy, True, elapsed_ms, "", "")
                else:
                    self.result_ready.emit(self.index, self.proxy, False, 0, "", f"HTTP {response.status}")
                    
        except Exception as e:
            error_msg = str(e)[:30]
            self.result_ready.emit(self.index, self.proxy, False, 0, "", error_msg)


class LicenseDialog(QDialog):
    """ライセンスキー入力ダイアログ"""
    
    def __init__(self, license_manager, parent=None):
        super().__init__(parent)
        self.license_manager = license_manager
        self.authenticated = False
        self.setWindowTitle("License Activation")
        self.setFixedSize(480, 340)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self._drag_pos = None
        self._setup_ui()
    
    def _setup_ui(self):
        # 外枠（角丸コンテナ）
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        
        container = QWidget()
        container.setObjectName("license_container")
        container.setStyleSheet("""
            #license_container {
                background-color: #252535;
                border-radius: 12px;
                border: 1px solid #303040;
            }
        """)
        outer.addWidget(container)
        
        layout = QVBoxLayout(container)
        layout.setContentsMargins(32, 28, 32, 28)
        layout.setSpacing(16)
        
        # タイトルバー（ドラッグ用 + 閉じるボタン）
        title_bar = QHBoxLayout()
        title_bar.setSpacing(4)
        
        # ロゴ画像を表示（GitHubからダウンロード）
        title_icon = QLabel()
        title_icon.setFixedSize(32, 32)
        title_icon.setStyleSheet("background: transparent; margin-top: 2px;")
        logo_icon = get_icon_from_base64("Logo")
        if not logo_icon.isNull():
            title_icon.setPixmap(logo_icon.pixmap(32, 32))
        title_bar.addWidget(title_icon)
        
        title = QLabel("License Activation")
        title.setStyleSheet("font-size: 18px; font-weight: bold; color: #ffffff; background: transparent;")
        title_bar.addWidget(title)
        title_bar.addStretch()
        
        close_btn = QPushButton("✕")
        close_btn.setFixedSize(28, 28)
        close_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: transparent; color: #808080;
                border: none; font-size: 16px; border-radius: 14px;
            }
            QPushButton:hover { background-color: #e74c3c; color: white; }
        """)
        close_btn.clicked.connect(self.reject)
        title_bar.addWidget(close_btn)
        layout.addLayout(title_bar)
        
        # 説明文
        desc = QLabel("Enter your license key to activate.")
        desc.setStyleSheet("color: #b0b0b0; font-size: 13px; background: transparent;")
        layout.addWidget(desc)
        
        # ライセンスキー入力
        self.key_input = QLineEdit()
        self.key_input.setPlaceholderText("XXXX-XXXX-XXXX-XXXX")
        self.key_input.setStyleSheet("""
            QLineEdit {
                background-color: #1e1e2e;
                border: 1px solid #303040;
                border-radius: 8px;
                padding: 12px 16px;
                color: #ffffff;
                font-size: 15px;
                font-family: 'Consolas', 'Monaco', monospace;
                letter-spacing: 2px;
            }
            QLineEdit:focus {
                border-color: #3498db;
            }
        """)
        # キャッシュがあれば自動入力
        cached_key = self.license_manager.cached_license_key
        if cached_key:
            self.key_input.setText(cached_key)
        self.key_input.returnPressed.connect(self._on_activate)
        layout.addWidget(self.key_input)
        
        # エラーメッセージ
        self.error_label = QLabel("")
        self.error_label.setStyleSheet("color: #e74c3c; font-size: 12px; background: transparent;")
        self.error_label.setWordWrap(True)
        self.error_label.hide()
        layout.addWidget(self.error_label)
        
        layout.addStretch()
        
        # ボタンエリア
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        self.activate_btn = QPushButton("  Activate  ")
        self.activate_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.activate_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 8px;
                padding: 10px 32px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:disabled {
                background-color: #404050;
                color: #808080;
            }
        """)
        self.activate_btn.clicked.connect(self._on_activate)
        btn_layout.addWidget(self.activate_btn)
        layout.addLayout(btn_layout)
    
    def _on_activate(self):
        """Activateボタンクリック"""
        key = self.key_input.text().strip()
        if not key:
            self._show_error("Please enter a license key")
            return
        
        self.activate_btn.setEnabled(False)
        self.activate_btn.setText("  Verifying...  ")
        QApplication.processEvents()
        
        success, message = self.license_manager.activate(key)
        
        if success:
            self.authenticated = True
            self.accept()
        else:
            self._show_error(message)
            self.activate_btn.setEnabled(True)
            self.activate_btn.setText("  Activate  ")
    
    def _show_error(self, message):
        """エラーメッセージを表示"""
        self.error_label.setText(f"⚠ {message}")
        self.error_label.show()
    
    # ドラッグ移動
    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            self._drag_pos = event.globalPosition().toPoint() - self.frameGeometry().topLeft()
    
    def mouseMoveEvent(self, event):
        if event.buttons() == Qt.MouseButton.LeftButton and self._drag_pos:
            self.move(event.globalPosition().toPoint() - self._drag_pos)
    
    def mouseReleaseEvent(self, event):
        self._drag_pos = None


class MainWindow(QMainWindow):
    """メインウィンドウ"""
    
    RESIZE_MARGIN = 8  # リサイズ可能な端の幅
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Project WIN")
        self.setMinimumSize(1200, 800)
        self.resize(1320, 820)  # デフォルトサイズ
        # タイトルバーを非表示にしてフレームレスに（UpdateDialogと同じ）
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint)
        # 角を丸くするために背景を透明に
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        # リサイズ端でカーソル変更のためHoverイベントを有効化
        self.setAttribute(Qt.WidgetAttribute.WA_Hover)
        self.setMouseTracking(True)
        self.sidebar_expanded = True
        self._drag_pos = None
        self._resizing = False
        self._resize_direction = None
        
        self.setup_ui()
        self.apply_dark_theme()
    
    def mousePressEvent(self, event):
        """ドラッグ/リサイズ開始位置を記録"""
        if event.button() == Qt.MouseButton.LeftButton:
            pos = event.position().toPoint()
            self._resize_direction = self._get_resize_direction(pos)
            
            if self._resize_direction:
                self._resizing = True
            else:
                self._drag_pos = event.globalPosition().toPoint() - self.frameGeometry().topLeft()
            event.accept()
    
    def mouseMoveEvent(self, event):
        """ウィンドウをドラッグで移動またはリサイズ"""
        if self._resizing and self._resize_direction:
            self._do_resize(event.globalPosition().toPoint())
            event.accept()
        elif event.buttons() == Qt.MouseButton.LeftButton and self._drag_pos is not None:
            self.move(event.globalPosition().toPoint() - self._drag_pos)
            event.accept()
    
    def mouseReleaseEvent(self, event):
        """ドラッグ/リサイズ終了"""
        self._drag_pos = None
        self._resizing = False
        self._resize_direction = None
        self.setCursor(Qt.CursorShape.ArrowCursor)
    
    def _update_cursor_for_pos(self, local_pos):
        """ウィンドウ座標でのマウス位置からリサイズカーソルを更新"""
        direction = self._get_resize_direction(local_pos)
        if direction in ('top-left', 'bottom-right'):
            self.setCursor(Qt.CursorShape.SizeFDiagCursor)
        elif direction in ('top-right', 'bottom-left'):
            self.setCursor(Qt.CursorShape.SizeBDiagCursor)
        elif direction in ('left', 'right'):
            self.setCursor(Qt.CursorShape.SizeHorCursor)
        elif direction in ('top', 'bottom'):
            self.setCursor(Qt.CursorShape.SizeVerCursor)
        else:
            self.unsetCursor()
    
    def event(self, event):
        """全イベントを監視してリサイズ端でカーソルを変更"""
        from PySide6.QtCore import QEvent
        if event.type() == QEvent.Type.HoverMove and not self._resizing:
            pos = event.position().toPoint()
            self._update_cursor_for_pos(pos)
        return super().event(event)
    
    def _get_resize_direction(self, pos):
        """マウス位置からリサイズ方向を判定"""
        rect = self.rect()
        x, y = pos.x(), pos.y()
        m = self.RESIZE_MARGIN
        
        on_left = x < m
        on_right = x > rect.width() - m
        on_top = y < m
        on_bottom = y > rect.height() - m
        
        if on_top and on_left:
            return 'top-left'
        elif on_top and on_right:
            return 'top-right'
        elif on_bottom and on_left:
            return 'bottom-left'
        elif on_bottom and on_right:
            return 'bottom-right'
        elif on_left:
            return 'left'
        elif on_right:
            return 'right'
        elif on_top:
            return 'top'
        elif on_bottom:
            return 'bottom'
        return None
    
    def _do_resize(self, global_pos):
        """リサイズを実行"""
        geo = self.frameGeometry()
        min_w, min_h = self.minimumWidth(), self.minimumHeight()
        
        if 'left' in self._resize_direction:
            new_left = global_pos.x()
            new_width = geo.right() - new_left
            if new_width >= min_w:
                geo.setLeft(new_left)
        
        if 'right' in self._resize_direction:
            new_width = global_pos.x() - geo.left()
            if new_width >= min_w:
                geo.setWidth(new_width)
        
        if 'top' in self._resize_direction:
            new_top = global_pos.y()
            new_height = geo.bottom() - new_top
            if new_height >= min_h:
                geo.setTop(new_top)
        
        if 'bottom' in self._resize_direction:
            new_height = global_pos.y() - geo.top()
            if new_height >= min_h:
                geo.setHeight(new_height)
        
        self.setGeometry(geo)
    
    def setup_ui(self):
        # メインウィジェット（角丸のコンテナ）
        main_widget = QWidget()
        main_widget.setObjectName("main_container")
        main_widget.setMouseTracking(True)
        main_widget.setStyleSheet("""
            #main_container {
                background-color: #252535;
                border-radius: 10px;
            }
        """)
        self.setCentralWidget(main_widget)
        main_layout = QHBoxLayout(main_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        
        # サイドバーコンテナ（折りたたみ用）- 左側の角丸
        self.sidebar_container = QFrame()
        self.sidebar_container.setStyleSheet("""
            QFrame {
                background-color: #1e1e2e;
                border-top-left-radius: 10px;
                border-bottom-left-radius: 10px;
            }
        """)
        self.sidebar_width_expanded = 130
        self.sidebar_width_collapsed = 80
        self.sidebar_container.setFixedWidth(self.sidebar_width_expanded)
        
        sidebar_layout = QVBoxLayout(self.sidebar_container)
        sidebar_layout.setContentsMargins(12, 15, 0, 20)
        sidebar_layout.setSpacing(0)
        
        # ロゴ（GitHubからダウンロード）
        self.logo_icon = QLabel()
        self.logo_icon.setFixedSize(72, 72)
        self.logo_icon.setStyleSheet("margin-left: -10px;")
        logo_icon = get_icon_from_base64("Logo")
        if not logo_icon.isNull():
            logo_pixmap = logo_icon.pixmap(72, 72)
            self.logo_icon.setPixmap(logo_pixmap)
        sidebar_layout.addWidget(self.logo_icon)

        
        # タイトル（ロゴのすぐ下、間隔を狭く）
        self.title_label = QLabel("Project WIN")
        self.title_label.setStyleSheet("""
            font-size: 11px;
            font-weight: bold;
            color: #ffffff;
            padding: 0px;
            margin-top: 0px;
            margin-bottom: 15px;
        """)
        sidebar_layout.addWidget(self.title_label)
        
        # ナビゲーションボタン（Base64アイコン使用）
        self.nav_buttons = []
        
        self.task_btn = SidebarButton("Tasks", "task")
        self.task_btn.setChecked(True)
        self.task_btn.clicked.connect(lambda: self.switch_page(0))
        sidebar_layout.addWidget(self.task_btn)
        self.nav_buttons.append(self.task_btn)
        
        self.setting_btn = SidebarButton("Settings", "settings")
        self.setting_btn.clicked.connect(lambda: self.switch_page(1))
        sidebar_layout.addWidget(self.setting_btn)
        self.nav_buttons.append(self.setting_btn)
        
        self.proxy_btn = SidebarButton("Proxies", "web")
        self.proxy_btn.clicked.connect(lambda: self.switch_page(2))
        sidebar_layout.addWidget(self.proxy_btn)
        self.nav_buttons.append(self.proxy_btn)
        
        self.tools_btn = SidebarButton("Toolbox", "box")
        self.tools_btn.clicked.connect(lambda: self.switch_page(3))
        sidebar_layout.addWidget(self.tools_btn)
        self.nav_buttons.append(self.tools_btn)
        
        # スペーサー（ボタンと折りたたみボタンの間隔）
        sidebar_layout.addSpacing(80)
        
        # 折りたたみボタン
        self.toggle_btn = QPushButton("◀")
        self.toggle_btn.setFixedSize(24, 48)
        self.toggle_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.toggle_btn.setStyleSheet("""
            QPushButton {
                background-color: #252535;
                color: #606070;
                border: 1px solid #303040;
                border-right: none;
                border-top-left-radius: 6px;
                border-bottom-left-radius: 6px;
                border-top-right-radius: 0px;
                border-bottom-right-radius: 0px;
                font-size: 10px;
            }
            QPushButton:hover {
                background-color: #353545;
                color: #ffffff;
            }
        """)
        self.toggle_btn.clicked.connect(self.toggle_sidebar)
        sidebar_layout.addWidget(self.toggle_btn, alignment=Qt.AlignmentFlag.AlignRight)
        
        # バージョン情報を下に押し下げるstretch
        sidebar_layout.addStretch()
        
        # バージョン情報（動的に読み取り）
        _app_version = "0.0.0"
        try:
            import json as _json
            # is_compiled()を使用（PyInstaller/Nuitka両対応）
            if is_compiled():
                _base = os.path.dirname(sys.executable)
                _vpath = os.path.join(_base, "_internal", "version.json")
            else:
                _base = os.path.dirname(os.path.abspath(__file__))
                _vpath = os.path.join(_base, "version.json")
            if os.path.exists(_vpath):
                with open(_vpath, "r", encoding="utf-8") as _vf:
                    _app_version = _json.load(_vf).get("version", "0.0.0")
        except Exception:
            pass
        self.version_label = QLabel(f"v{_app_version}")
        self.version_label.setStyleSheet("color: #606070; font-size: 12px; padding: 10px; margin-right: 12px;")
        sidebar_layout.addWidget(self.version_label)
        
        main_layout.addWidget(self.sidebar_container)
        
        # メインコンテンツコンテナ（コントロールボタン + コンテンツ）- 右側の角丸
        content_container = QWidget()
        content_container.setStyleSheet("""
            background-color: #252535;
            border-top-right-radius: 10px;
            border-bottom-right-radius: 10px;
        """)
        content_layout = QVBoxLayout(content_container)
        content_layout.setContentsMargins(0, 0, 0, 0)
        content_layout.setSpacing(0)
        
        # ウィンドウコントロール（右上）
        control_bar = QWidget()
        control_bar.setFixedHeight(32)
        control_bar.setStyleSheet("background-color: transparent;")
        control_bar_layout = QHBoxLayout(control_bar)
        control_bar_layout.setContentsMargins(0, 4, 10, 0)
        control_bar_layout.setSpacing(0)
        control_bar_layout.addStretch()
        
        # ボタン共通スタイル
        control_btn_style = """
            QPushButton {
                background-color: transparent;
                color: #b0b0b0;
                border: none;
                font-size: 18px;
                font-family: monospace;
            }
            QPushButton:hover {
                background-color: #404050;
            }
        """
        
        # 最小化ボタン
        minimize_btn = QPushButton("−")
        minimize_btn.setFixedSize(46, 28)
        minimize_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        minimize_btn.setStyleSheet(control_btn_style)
        minimize_btn.clicked.connect(self.showMinimized)
        control_bar_layout.addWidget(minimize_btn)
        
        # 最大化ボタン
        self.maximize_btn = QPushButton("□")
        self.maximize_btn.setFixedSize(46, 28)
        self.maximize_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.maximize_btn.setStyleSheet(control_btn_style)
        self.maximize_btn.clicked.connect(self.toggle_maximize)
        control_bar_layout.addWidget(self.maximize_btn)
        
        # 閉じるボタン
        close_btn = QPushButton("×")
        close_btn.setFixedSize(46, 28)
        close_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                color: #b0b0b0;
                border: none;
                font-size: 18px;
                font-family: monospace;
            }
            QPushButton:hover {
                background-color: #e74c3c;
                color: white;
            }
        """)
        close_btn.clicked.connect(self.close)
        control_bar_layout.addWidget(close_btn)
        
        content_layout.addWidget(control_bar)
        
        # メインコンテンツエリア
        self.content_stack = QStackedWidget()
        self.content_stack.setStyleSheet("background-color: #252535;")
        
        # 各ページを追加（参照を保持）
        self.proxy_page = ProxyPage()
        self.setting_page = SettingPage()
        self.tools_page = ToolsPage()
        self.task_page = TaskPage(self.proxy_page)
        self.task_page.settings_page = self.setting_page  # Webhook送信用に参照を設定
        
        self.content_stack.addWidget(self.task_page)
        self.content_stack.addWidget(self.setting_page)
        self.content_stack.addWidget(self.proxy_page)
        self.content_stack.addWidget(self.tools_page)
        
        content_layout.addWidget(self.content_stack)
        
        main_layout.addWidget(content_container)
    
    def toggle_maximize(self):
        """最大化/通常サイズを切り替える"""
        if self.isMaximized():
            self.showNormal()
            self.maximize_btn.setText("□")
        else:
            self.showMaximized()
            self.maximize_btn.setText("❐")
    
    def toggle_sidebar(self):
        """サイドバーの展開/折りたたみを切り替える"""
        self.sidebar_expanded = not self.sidebar_expanded
        
        if self.sidebar_expanded:
            # 展開
            self.sidebar_container.setFixedWidth(self.sidebar_width_expanded)
            self.toggle_btn.setText("◀")
            self.title_label.show()
            self.version_label.show()
            self.logo_icon.setStyleSheet("margin-left: -10px;")
        else:
            # 折りたたみ（アイコンは表示、タイトルは非表示）
            self.sidebar_container.setFixedWidth(self.sidebar_width_collapsed)
            self.toggle_btn.setText("▶")
            self.title_label.hide()
            self.version_label.hide()
            # タイトル分の高さ（約30px）を補う
            self.logo_icon.setStyleSheet("margin-left: -10px;")
        
        # ナビゲーションボタンの表示を更新
        for btn in self.nav_buttons:
            btn.set_expanded(self.sidebar_expanded)
    
    def switch_page(self, index):
        """ページを切り替える"""
        self.content_stack.setCurrentIndex(index)
        
        # ボタンの選択状態を更新
        for i, btn in enumerate(self.nav_buttons):
            btn.setChecked(i == index)
    
    def apply_dark_theme(self):
        """ダークテーマを適用"""
        self.setStyleSheet("""
            QMainWindow {
                background-color: #252535;
            }
            QLabel {
                color: #e0e0e0;
            }
            QScrollBar:vertical {
                background-color: #2a2a3a;
                width: 12px;
                border-radius: 6px;
            }
            QScrollBar::handle:vertical {
                background-color: #404050;
                border-radius: 6px;
                min-height: 30px;
            }
            QScrollBar::handle:vertical:hover {
                background-color: #505060;
            }
        """)


class SplashScreen(QWidget):
    """起動時のローディングスプラッシュスクリーン"""
    
    def __init__(self):
        super().__init__()
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.WindowStaysOnTopHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setFixedSize(350, 250)
        
        # 画面中央に配置
        screen = QApplication.primaryScreen().geometry()
        self.move(
            (screen.width() - self.width()) // 2,
            (screen.height() - self.height()) // 2
        )
        
        self._logo_pixmap = None
        self._load_logo()
        self._setup_ui()
        self._angle = 0
        
        # アニメーションタイマー
        self._timer = QTimer()
        self._timer.timeout.connect(self._rotate)
        self._timer.start(50)
    
    def _load_logo(self):
        """GitHubからロゴをダウンロード"""
        import urllib.request
        try:
            url = f"{ICONS_BASE_URL}Logo.png"
            with urllib.request.urlopen(url, timeout=5) as response:
                data = response.read()
                self._logo_pixmap = QPixmap()
                self._logo_pixmap.loadFromData(data)
        except Exception as e:
            print(f"Logo download error: {e}")
            self._logo_pixmap = None
    
    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        
        # メインコンテナ
        container = QWidget()
        container.setStyleSheet("""
            QWidget {
                background-color: #1e1e2e;
                border-radius: 16px;
                border: 1px solid #3a3a4a;
            }
        """)
        container_layout = QVBoxLayout(container)
        container_layout.setContentsMargins(40, 35, 40, 35)
        container_layout.setSpacing(20)
        container_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        # ロゴ表示用ラベル
        self.logo_label = QLabel()
        self.logo_label.setFixedSize(100, 100)
        self.logo_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        if self._logo_pixmap and not self._logo_pixmap.isNull():
            # ロゴがあれば表示
            scaled = self._logo_pixmap.scaled(90, 90, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
            self.logo_label.setPixmap(scaled)
            self.logo_label.setStyleSheet("background: transparent; border: none;")
        else:
            # ロゴがなければ四角い枠を表示
            self.logo_label.setStyleSheet("""
                background: transparent;
                border: 3px solid #3a3a4a;
                border-radius: 16px;
            """)
        
        container_layout.addWidget(self.logo_label, alignment=Qt.AlignmentFlag.AlignCenter)
        
        # テキスト
        self.label = QLabel("Starting...")
        self.label.setStyleSheet("color: #ffffff; font-size: 16px; font-weight: bold; background: transparent; border: none;")
        self.label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        container_layout.addWidget(self.label)
        
        layout.addWidget(container)
    
    def _rotate(self):
        self._angle = (self._angle + 8) % 360
        self.update()
    
    def paintEvent(self, event):
        super().paintEvent(event)
        
        # ロゴがない場合のみ回転する四角を描画
        if self._logo_pixmap is None or self._logo_pixmap.isNull():
            painter = QPainter(self)
            painter.setRenderHint(QPainter.RenderHint.Antialiasing)
            
            # 四角の位置を計算
            logo_rect = self.logo_label.geometry()
            center_x = logo_rect.x() + logo_rect.width() // 2 + 40
            center_y = logo_rect.y() + logo_rect.height() // 2 + 35
            
            # 回転する四角を描画
            pen = QPen(QColor("#4a90d9"), 3)
            pen.setCapStyle(Qt.PenCapStyle.RoundCap)
            painter.setPen(pen)
            painter.setBrush(Qt.BrushStyle.NoBrush)
            painter.translate(center_x, center_y)
            painter.rotate(self._angle)
            painter.drawRoundedRect(-35, -35, 70, 70, 12, 12)
            
            painter.end()
    
    def set_message(self, text):
        self.label.setText(text)
    
    def finish(self):
        self._timer.stop()
        self.close()


def main():
    # ===== 多重起動防止（exe同士のみ） =====
    try:
        import ctypes
        
        # exe か py かを判定
        is_frozen = getattr(sys, 'frozen', False)
        
        if is_frozen:
            # exe起動の場合のみ多重起動を防止
            mutex_name = "ProjectWIN_EXE_SingleInstance_Mutex"
            kernel32 = ctypes.windll.kernel32
            mutex = kernel32.CreateMutexW(None, False, mutex_name)
            last_error = kernel32.GetLastError()
            
            # ERROR_ALREADY_EXISTS (183) = 既に別のインスタンスが起動中
            if last_error == 183:
                ctypes.windll.user32.MessageBoxW(
                    None,
                    "Project WIN is already running.\nOnly one instance is allowed.",
                    "Project WIN",
                    0x40  # MB_ICONINFORMATION
                )
                return
        # py起動の場合は多重起動チェックをスキップ（開発用）
    except:
        pass  # ctypesが使えない環境ではスキップ
    
    app = QApplication(sys.argv)
    
    # アプリケーションフォント設定
    font = QFont("Segoe UI", 10)
    app.setFont(font)
    
    # ===== スプラッシュスクリーン表示 =====
    splash = SplashScreen()
    splash.show()
    app.processEvents()
    
    # ===== Playwrightを事前初期化（QThread問題を回避） =====
    splash.set_message("Initializing...")
    app.processEvents()
    try:
        from playwright.sync_api import sync_playwright
        _pw = sync_playwright().start()
        _pw.stop()
        print("Playwright initialized")
    except Exception as e:
        print(f"Playwright init warning: {e}")
    
    # ===== ライセンス認証 =====
    splash.set_message("Checking license...")
    app.processEvents()
    if LicenseManager is not None:
        lm = LicenseManager()
        
        # キャッシュがあれば自動検証
        if lm.cached_license_key:
            success, message = lm.verify()
            if success:
                print(f"License: {message}")
            else:
                # 検証失敗 → ダイアログ表示
                splash.finish()
                dialog = LicenseDialog(lm)
                if dialog.exec() != QDialog.DialogCode.Accepted:
                    sys.exit(0)
                splash = SplashScreen()
                splash.show()
                app.processEvents()
        else:
            # キャッシュなし → ダイアログ表示（初回起動）
            splash.finish()
            dialog = LicenseDialog(lm)
            if dialog.exec() != QDialog.DialogCode.Accepted:
                sys.exit(0)
            splash = SplashScreen()
            splash.show()
            app.processEvents()
        
        print(f"License authenticated")
    else:
        print("License manager not found - running without license check")
    
    # ===== アップデートチェック =====
    splash.set_message("Checking updates...")
    app.processEvents()
    print(f"check_for_update function: {check_for_update}")
    if check_for_update is not None:
        try:
            print("Calling check_for_update()...")
            needs_update, latest_version, download_url, changelog = check_for_update()
            print(f"Result: needs_update={needs_update}, latest={latest_version}, url={download_url}")
            if needs_update and download_url:
                splash.finish()
                update_dialog = UpdateDialog(latest_version, changelog)
                update_dialog.exec()
                # ユーザーが「あとで」を選んだ場合はそのまま続行
                splash = SplashScreen()
                splash.show()
                app.processEvents()
        except Exception as e:
            print(f"Update check failed: {e}")
            import traceback
            traceback.print_exc()
    else:
        print("check_for_update is None - skipping update check")
    
    # ===== メインウィンドウ起動 =====
    splash.set_message("Starting...")
    app.processEvents()
    window = MainWindow()
    
    # スプラッシュを閉じてメインウィンドウを表示
    splash.finish()
    window.show()
    
    # ハートビート（ライセンス有効性の定期チェック）
    if LicenseManager is not None:
        heartbeat_timer = QTimer()
        heartbeat_timer.timeout.connect(lm.heartbeat)
        heartbeat_timer.start(LicenseManager.HEARTBEAT_INTERVAL * 1000)
    
    sys.exit(app.exec())


if __name__ == "__main__":
    main()